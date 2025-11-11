from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, F, Q, Func, Value, Avg, Case, When, FloatField, DecimalField
from django.db.models.functions import TruncMonth, TruncYear
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.urls import reverse
from datetime import datetime, timedelta
import json
import csv
from django.db import transaction

from .models import (
    Web_User, Pharmacy_Details, ProductMaster, SupplierMaster, CustomerMaster,
    InvoiceMaster, InvoicePaid, PurchaseMaster, SalesInvoiceMaster, SalesMaster,
    SalesInvoicePaid, ProductRateMaster, ReturnInvoiceMaster, PurchaseReturnInvoicePaid,
    ReturnPurchaseMaster, ReturnSalesInvoiceMaster, ReturnSalesInvoicePaid, ReturnSalesMaster,
    SaleRateMaster, PaymentMaster, ReceiptMaster
)
from .forms import (
    LoginForm, UserRegistrationForm, UserUpdateForm, PharmacyDetailsForm, ProductForm,
    SupplierForm, CustomerForm, InvoiceForm, InvoicePaymentForm, PurchaseForm,
    SalesInvoiceForm, SalesForm, SalesPaymentForm, ProductRateForm,
    PurchaseReturnInvoiceForm, PurchaseReturnForm, SalesReturnInvoiceForm, SalesReturnForm,
    SaleRateForm, SalesReturnPaymentForm, PaymentForm, ReceiptForm
)
from .utils import get_stock_status, get_batch_stock_status, generate_invoice_pdf, generate_sales_invoice_pdf, get_avg_mrp, parse_expiry_date, generate_sales_invoice_number
from .date_utils import parse_ddmmyyyy_date, format_date_for_display, format_date_for_backend, convert_legacy_dates
from .low_stock_views import low_stock_update, update_low_stock_item, bulk_update_low_stock
# Authentication views
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f"Welcome back, {user.get_full_name()}!")
                return redirect('dashboard')
            else:
                messages.error(request, "Invalid username or password.")
        else:
            messages.error(request, "Invalid username or password.")
    else:
        form = LoginForm()
        
    context = {
        'form': form,
        'title': 'Login'
    }
    return render(request, 'login.html', context)
def logout_view(request):
    logout(request)
    messages.info(request, "You have successfully logged out.")
    return redirect('login')
@login_required
def register_user(request):
    if not request.user.user_type == 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('dashboard')
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()
            messages.success(request, f"Account created for {user.username}!")
            return redirect('user_list')
    else:
        form = UserRegistrationForm()
    
    context = {
        'form': form,
        'title': 'Register User'
    }
    return render(request, 'profile.html', context)

@login_required
def user_list(request):
    if not request.user.user_type == 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('dashboard')
    
    users = Web_User.objects.all().order_by('username')
    
    context = {
        'users': users,
        'title': 'User List'
    }
    return render(request, 'user_list.html', context)

@login_required
def update_user(request, pk):
    if not request.user.user_type == 'admin' and not request.user.id == pk:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('dashboard')
    
    user = get_object_or_404(Web_User, id=pk)
    
    if request.method == 'POST':
        form = UserUpdateForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, f"Account updated for {user.username}!")
            if request.user.user_type == 'admin':
                return redirect('user_list')
            else:
                return redirect('profile')
    else:
        form = UserUpdateForm(instance=user)
    
    context = {
        'form': form,
        'user': user,
        'title': 'Update User'
    }
    return render(request, 'profile.html', context)

@login_required
def profile(request):
    return update_user(request, request.user.id)

# Dashboard
@login_required
def dashboard(request):
    # Get counts for dashboard cards
    product_count = ProductMaster.objects.count()
    supplier_count = SupplierMaster.objects.count()
    customer_count = CustomerMaster.objects.count()
    
    # Get recent sales
    recent_sales = SalesInvoiceMaster.objects.order_by('-sales_invoice_date')[:5]
    
    # Get recent purchases
    recent_purchases = InvoiceMaster.objects.order_by('-invoice_date')[:5]
    
    # Get low stock products (simplified for dashboard)
    low_stock_products = []
    
    try:
        products = ProductMaster.objects.all()[:20]  # Limit to first 20 for performance
        print(f"Dashboard: Checking {products.count()} products for low stock")
        
        for product in products:
            try:
                stock_info = get_stock_status(product.productid)
                current_stock = stock_info.get('current_stock', 0)
                
                if current_stock <= 10 and current_stock > 0:
                    low_stock_products.append({
                        'product': product,
                        'current_stock': current_stock
                    })
                    print(f"Dashboard: Added low stock product {product.product_name} - Stock: {current_stock}")
                    
                    if len(low_stock_products) >= 10:  # Limit to 10 for dashboard
                        break
            except Exception as e:
                print(f"Dashboard: Error processing stock for {product.product_name}: {e}")
                continue
        
        print(f"Dashboard: Total low stock products found: {len(low_stock_products)}")
        
    except Exception as e:
        print(f"Dashboard: Error in low stock section: {e}")
        low_stock_products = []
    
    # Get expired/expiring soon products
    expired_products = []
    from datetime import datetime, timedelta
    
    try:
        # Get products expiring in next 30 days or already expired
        current_date = datetime.now().date()
        warning_date = current_date + timedelta(days=30)
        
        print(f"Dashboard: Looking for products expiring before {warning_date}")
        
        # Check purchase records for expiry dates
        purchases_with_expiry = PurchaseMaster.objects.filter(
            product_expiry__isnull=False
        ).exclude(product_expiry='').select_related('productid')[:50]  # Limit for performance
        
        print(f"Dashboard: Found {purchases_with_expiry.count()} purchases with expiry dates")
        
        for purchase in purchases_with_expiry:
            try:
                # Parse expiry date
                expiry_str = str(purchase.product_expiry)
                expiry_date = None
                
                # Handle different date formats
                if len(expiry_str) == 10 and '-' in expiry_str:  # YYYY-MM-DD
                    expiry_date = datetime.strptime(expiry_str, '%Y-%m-%d').date()
                elif len(expiry_str) == 7 and '-' in expiry_str:  # MM-YYYY
                    month, year = expiry_str.split('-')
                    import calendar
                    last_day = calendar.monthrange(int(year), int(month))[1]
                    expiry_date = datetime(int(year), int(month), last_day).date()
                
                if expiry_date and expiry_date <= warning_date:
                    # Check if product has stock
                    stock_info = get_batch_stock_status(purchase.productid.productid, purchase.product_batch_no)
                    if stock_info[1] and stock_info[0] > 0:  # Has stock
                        expired_products.append({
                            'product': purchase.productid,
                            'batch_no': purchase.product_batch_no,
                            'expiry_date': expiry_date,
                            'current_stock': stock_info[0],
                            'days_to_expiry': (expiry_date - current_date).days
                        })
                        
                        print(f"Dashboard: Added expiring product {purchase.productid.product_name} - {purchase.product_batch_no}")
                        
                        if len(expired_products) >= 10:  # Limit to 10 for dashboard
                            break
            except Exception as e:
                print(f"Dashboard: Error processing expiry for {purchase.productid.product_name}: {e}")
                continue  # Skip invalid dates
        
        print(f"Dashboard: Total expired/expiring products found: {len(expired_products)}")
        
    except Exception as e:
        print(f"Dashboard: Error in expired products section: {e}")
        expired_products = []
    
    today = timezone.now().date()
    
    # Calculate financial overview
    current_month_start = today.replace(day=1)
    
    # Monthly sales
    monthly_sales_invoices = SalesInvoiceMaster.objects.filter(
        sales_invoice_date__gte=current_month_start
    )
    monthly_sales = SalesMaster.objects.filter(
        sales_invoice_no__in=monthly_sales_invoices
    ).aggregate(total=Sum('sale_total_amount'))['total'] or 0
    
    # Monthly purchases
    monthly_purchases = InvoiceMaster.objects.filter(
        invoice_date__gte=current_month_start
    ).aggregate(total=Sum('invoice_total'))['total'] or 0
    
    # Total outstanding payments from customers
    # Calculate total sales amounts
    sales_totals = SalesMaster.objects.values('sales_invoice_no').annotate(
        invoice_total=Sum('sale_total_amount')
    )
    
    # Create a dictionary mapping invoice numbers to their total amounts
    invoice_totals = {item['sales_invoice_no']: item['invoice_total'] for item in sales_totals}
    
    # Calculate total receivable properly
    total_receivable = 0
    for invoice in SalesInvoiceMaster.objects.all():
        # Get actual invoice total from sales items
        invoice_total = SalesMaster.objects.filter(
            sales_invoice_no=invoice.sales_invoice_no
        ).aggregate(Sum('sale_total_amount'))['sale_total_amount__sum'] or 0
        
        balance = invoice_total - invoice.sales_invoice_paid
        if balance > 0:
            total_receivable += balance
    
    # Total outstanding payments to suppliers
    total_payable = InvoiceMaster.objects.aggregate(
        total=Sum(F('invoice_total') - F('invoice_paid'))
    )['total'] or 0
    
    # Debug output
    print(f"Dashboard context: low_stock={len(low_stock_products)}, expired={len(expired_products)}")
    
    context = {
        'title': 'Dashboard',
        'product_count': product_count,
        'supplier_count': supplier_count,
        'customer_count': customer_count,
        'recent_sales': recent_sales,
        'recent_purchases': recent_purchases,
        'low_stock_products': low_stock_products,
        'low_stock_count': len(low_stock_products),
        'expired_products': expired_products,
        'monthly_sales': monthly_sales,
        'monthly_purchases': monthly_purchases,
        'total_receivable': total_receivable,
        'total_payable': total_payable
    }
    return render(request, 'dashboard.html', context)

# Pharmacy Details
@login_required
def pharmacy_details(request):
    if not request.user.user_type == 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('dashboard')
        
    try:
        pharmacy = Pharmacy_Details.objects.first()
        form = PharmacyDetailsForm(instance=pharmacy)
    except Pharmacy_Details.DoesNotExist:
        pharmacy = None
        form = PharmacyDetailsForm()
    
    if request.method == 'POST':
        if pharmacy:
            form = PharmacyDetailsForm(request.POST, instance=pharmacy)
        else:
            form = PharmacyDetailsForm(request.POST)
        
        if form.is_valid():
            form.save()
            messages.success(request, "Pharmacy details updated successfully!")
            return redirect('pharmacy_details')
    
    context = {
        'form': form,
        'pharmacy': pharmacy,
        'title': 'Pharmacy Details'
    }
    return render(request, 'pharmacy_details.html', context)


# Product views
@login_required
def product_list(request):
    # Get sorting parameter (default: productid to show newest last)
    sort_by = request.GET.get('sort', 'productid')
    
    if sort_by == 'name':
        products = ProductMaster.objects.all().order_by('product_name')
    else:
        products = ProductMaster.objects.all().order_by('productid')

    # Enhanced search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        # Split search query into words for better matching
        search_words = search_query.split()
        
        # Start with all products
        search_filter = Q()
        
        # For each word, create OR conditions across all fields
        for word in search_words:
            word_filter = (
                Q(product_name__icontains=word) |
                Q(product_company__icontains=word) |
                Q(product_salt__icontains=word) |
                Q(product_packing__icontains=word) |
                Q(product_category__icontains=word) |
                Q(product_barcode__icontains=word)
            )
            search_filter &= word_filter  # AND between words
        
        products = products.filter(search_filter)
        
        # Sort by relevance: exact matches first, then startswith, then contains
        products = sorted(products, key=lambda p: (
            not p.product_name.lower().startswith(search_query.lower()),
            not search_query.lower() in p.product_name.lower(),
            p.product_name.lower()
        ))
        
        # Convert back to queryset for pagination
        product_ids = [p.productid for p in products]
        products = ProductMaster.objects.filter(productid__in=product_ids)
        # Preserve the sorted order
        products = sorted(products, key=lambda p: product_ids.index(p.productid))
    
    # Pagination first to limit the number of products processed
    paginator = Paginator(products, 30)  # 30 products per page
    page_number = request.GET.get('page')
    products_page = paginator.get_page(page_number)
    
    # Add stock and batch information to products
    for product in products_page:
        try:
            from .utils import get_inventory_batches_info
            
            # Get stock status
            stock_info = get_stock_status(product.productid)
            product.current_stock = stock_info.get('current_stock', 0)
            
            # Get all batches information with rates and MRP
            product.batches_info = get_inventory_batches_info(product.productid)
            
            # Set primary batch info for backward compatibility
            if product.batches_info:
                first_batch = product.batches_info[0]
                product.batch_no = first_batch['batch_no']
                product.expiry_date = first_batch['expiry']
                product.avg_mrp = first_batch['mrp']
                product.batch_rates = first_batch['rates']
            else:
                product.batch_no = 'No Batch'
                product.expiry_date = None
                product.avg_mrp = 0
                product.batch_rates = {'rate_A': 0, 'rate_B': 0, 'rate_C': 0}
                
        except Exception as e:
            print(f"Error calculating stock for {product.product_name}: {e}")
            product.current_stock = 0
            product.batch_no = 'Error'
            product.expiry_date = None
            product.batches_info = []
            product.avg_mrp = 0
            product.batch_rates = {'rate_A': 0, 'rate_B': 0, 'rate_C': 0}
    
    context = {
        'products': products_page,
        'search_query': search_query,
        'sort_by': sort_by,
        'title': 'Product List'
    }
    return render(request, 'products/product_list.html', context)

@login_required
def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save()
            messages.success(request, f"Product '{product.product_name}' added successfully!")
            return redirect('product_list')
    else:
        form = ProductForm()
    
    context = {
        'form': form,
        'title': 'Add Product'
    }
    return render(request, 'products/product_form.html', context)

@login_required
def update_product(request, pk):
    product = get_object_or_404(ProductMaster, productid=pk)
    
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, f"Product '{product.product_name}' updated successfully!")
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)
    
    context = {
        'form': form,
        'product': product,
        'title': 'Update Product'
    }
    return render(request, 'products/product_form.html', context)

@login_required
def product_detail(request, pk):
    product = get_object_or_404(ProductMaster, productid=pk)
    
    # Get stock status
    stock_info = get_stock_status(pk)
    
    # Get purchase history
    purchases = PurchaseMaster.objects.filter(productid=pk).order_by('-purchase_entry_date')
    
    # Get sales history
    sales = SalesMaster.objects.filter(productid=pk).order_by('-sale_entry_date')
    
    # Get rate history
    rates = ProductRateMaster.objects.filter(rate_productid=pk).order_by('-rate_date')
    
    context = {
        'product': product,
        'stock_info': stock_info,
        'purchases': purchases,
        'sales': sales,
        'rates': rates,
        'title': f'Product: {product.product_name}'
    }
    return render(request, 'products/product_detail.html', context)

@login_required
def bulk_upload_products(request):
    # Check if user is admin (case-insensitive)
    if not request.user.user_type.lower() in ['admin']:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('product_list')
    
    if request.method == 'POST':
        csv_file = request.FILES.get('csv_file')
        
        if not csv_file:
            messages.error(request, "Please select a CSV file.")
            return redirect('bulk_upload_products')
        
        if not csv_file.name.endswith('.csv'):
            messages.error(request, "File must be a CSV.")
            return redirect('bulk_upload_products')
        
        # Process CSV file
        try:
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            
            success_count = 0
            error_count = 0
            errors = []
            
            for row in reader:
                try:
                    # Create or update product
                    product, created = ProductMaster.objects.update_or_create(
                        product_name=row.get('product_name', '').strip(),
                        product_company=row.get('product_company', '').strip(),
                        defaults={
                            'product_packing': row.get('product_packing', '').strip(),
                            'product_salt': row.get('product_salt', '').strip(),
                            'product_category': row.get('product_category', '').strip(),
                            'product_hsn': row.get('product_hsn', '').strip(),
                            'product_hsn_percent': row.get('product_hsn_percent', '').strip(),
                        }
                    )
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    errors.append(f"Error on row {reader.line_num}: {str(e)}")
            
            if success_count > 0:
                messages.success(request, f"Successfully processed {success_count} products.")
            
            if error_count > 0:
                messages.warning(request, f"Encountered {error_count} errors during import.")
                for error in errors[:10]:  # Show first 10 errors
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.error(request, f"... and {len(errors) - 10} more errors.")
            
            return redirect('product_list')
            
        except Exception as e:
            messages.error(request, f"Error processing CSV file: {str(e)}")
            return redirect('bulk_upload_products')
    
    # Generate a sample CSV for download
    if request.GET.get('sample'):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="product_template.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'product_name', 'product_company', 'product_packing', 
            'product_salt', 'product_category', 'product_hsn', 
            'product_hsn_percent'
        ])
        
        # Add sample row
        writer.writerow([
            'Paracetamol 500mg', 'ABC Pharma', '10x10',
            'Paracetamol', 'Analgesic', '30049099',
            '12'
        ])
        
        return response
    
    context = {
        'title': 'Bulk Upload Products'
    }
    return render(request, 'products/bulk_upload.html', context)

@login_required
def delete_product(request, pk):
    # Check if user is admin (case-insensitive)
    if not request.user.user_type.lower() in ['admin']:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': "You don't have permission to perform this action."})
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('product_list')
        
    product = get_object_or_404(ProductMaster, productid=pk)
    
    if request.method == 'POST':
        product_name = product.product_name
        try:
            product.delete()
            
            # Handle AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True, 
                    'message': f"Product '{product_name}' deleted successfully!"
                })
            
            messages.success(request, f"Product '{product_name}' deleted successfully!")
        except Exception as e:
            # Handle AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False, 
                    'error': f"Cannot delete product. It is referenced by other records. Error: {str(e)}"
                })
            
            messages.error(request, f"Cannot delete product. It is referenced by other records. Error: {str(e)}")
        return redirect('product_list')
    
    context = {
        'product': product,
        'title': 'Delete Product'
    }
    return render(request, 'products/product_confirm_delete.html', context)

# Supplier views
@login_required
def supplier_list(request):
    suppliers = SupplierMaster.objects.all().order_by('supplier_name')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        suppliers = suppliers.filter(
            Q(supplier_name__icontains=search_query) | 
            Q(supplier_mobile__icontains=search_query) |
            Q(supplier_emailid__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(suppliers, 10)  # 10 suppliers per page
    page_number = request.GET.get('page')
    suppliers = paginator.get_page(page_number)
    
    context = {
        'suppliers': suppliers,
        'search_query': search_query,
        'title': 'Supplier List'
    }
    return render(request, 'suppliers/supplier_list.html', context)

@login_required
def add_supplier(request):
    if request.method == 'POST':
        # Handle AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                # Create supplier from POST data
                supplier = SupplierMaster.objects.create(
                    supplier_name=request.POST.get('supplier_name'),
                    supplier_type=request.POST.get('supplier_type'),
                    supplier_mobile=request.POST.get('supplier_mobile'),
                    supplier_whatsapp=request.POST.get('supplier_whatsapp'),
                    supplier_emailid=request.POST.get('supplier_emailid'),
                    supplier_spoc=request.POST.get('supplier_spoc'),
                    supplier_address=request.POST.get('supplier_address'),
                    supplier_dlno=request.POST.get('supplier_dlno'),
                    supplier_gstno=request.POST.get('supplier_gstno'),
                    supplier_bank=request.POST.get('supplier_bank'),
                    supplier_bankaccountno=request.POST.get('supplier_bankaccountno'),
                    supplier_bankifsc=request.POST.get('supplier_bankifsc'),
                    supplier_upi=request.POST.get('supplier_upi', '')
                )
                
                return JsonResponse({
                    'success': True,
                    'supplier_id': supplier.supplierid,
                    'supplier_name': supplier.supplier_name,
                    'message': f"Supplier '{supplier.supplier_name}' added successfully!"
                })
                
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                })
        
        # Handle regular form submission
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save()
            messages.success(request, f"Supplier '{supplier.supplier_name}' added successfully!")
            return redirect('supplier_list')
    else:
        form = SupplierForm()
    
    context = {
        'form': form,
        'title': 'Add Supplier'
    }
    return render(request, 'suppliers/supplier_form.html', context)

@login_required
def update_supplier(request, pk):
    supplier = get_object_or_404(SupplierMaster, supplierid=pk)
    
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, f"Supplier '{supplier.supplier_name}' updated successfully!")
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    
    context = {
        'form': form,
        'supplier': supplier,
        'title': 'Update Supplier'
    }
    return render(request, 'suppliers/supplier_form.html', context)

@login_required
def supplier_detail(request, pk):
    supplier = get_object_or_404(SupplierMaster, supplierid=pk)
    
    # Get invoices for this supplier
    invoices = InvoiceMaster.objects.filter(supplierid=pk).order_by('-invoice_date')
    
    # Calculate total purchase and payment amounts
    total_purchase = invoices.aggregate(Sum('invoice_total'))['invoice_total__sum'] or 0
    total_paid = invoices.aggregate(Sum('invoice_paid'))['invoice_paid__sum'] or 0
    balance = total_purchase - total_paid
    
    context = {
        'supplier': supplier,
        'invoices': invoices,
        'total_purchase': total_purchase,
        'total_paid': total_paid,
        'balance': balance,
        'title': f'Supplier: {supplier.supplier_name}'
    }
    return render(request, 'suppliers/supplier_detail.html', context)

@login_required
def delete_supplier(request, pk):
    # Check if user is admin (case-insensitive)
    if not request.user.user_type.lower() in ['admin']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('supplier_list')
        
    supplier = get_object_or_404(SupplierMaster, supplierid=pk)
    
    if request.method == 'POST':
        supplier_name = supplier.supplier_name
        try:
            supplier.delete()
            messages.success(request, f"Supplier '{supplier_name}' deleted successfully!")
        except Exception as e:
            messages.error(request, f"Cannot delete supplier. It is referenced by other records. Error: {str(e)}")
        return redirect('supplier_list')
    
    context = {
        'supplier': supplier,
        'title': 'Delete Supplier'
    }
    return render(request, 'suppliers/supplier_confirm_delete.html', context)

# Customer views
@login_required
def customer_list(request):
    customers = CustomerMaster.objects.all().order_by('customer_name')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        customers = customers.filter(
            Q(customer_name__icontains=search_query) | 
            Q(customer_mobile__icontains=search_query) |
            Q(customer_emailid__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(customers, 10)  # 10 customers per page
    page_number = request.GET.get('page')
    customers = paginator.get_page(page_number)
    
    context = {
        'customers': customers,
        'search_query': search_query,
        'title': 'Customer List'
    }
    return render(request, 'customers/customer_list.html', context)

@login_required
def add_customer(request):
    if request.method == 'POST':
        # Handle AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                # Create customer from POST data
                customer = CustomerMaster.objects.create(
                    customer_name=request.POST.get('customer_name'),
                    customer_type=request.POST.get('customer_type'),
                    customer_mobile=request.POST.get('customer_mobile'),
                    customer_whatsapp=request.POST.get('customer_whatsapp'),
                    customer_emailid=request.POST.get('customer_emailid'),
                    customer_spoc=request.POST.get('customer_spoc'),
                    customer_address=request.POST.get('customer_address'),
                    customer_dlno=request.POST.get('customer_dlno'),
                    customer_gstno=request.POST.get('customer_gstno'),
                    customer_food_license_no=request.POST.get('customer_food_license_no'),
                    customer_credit_days=int(request.POST.get('customer_credit_days', 0)),
                    customer_bank=request.POST.get('customer_bank'),
                    customer_bankaccountno=request.POST.get('customer_bankaccountno'),
                    customer_bankifsc=request.POST.get('customer_bankifsc'),
                    customer_upi=request.POST.get('customer_upi', '')
                )
                
                return JsonResponse({
                    'success': True,
                    'customer_id': customer.customerid,
                    'customer_name': customer.customer_name,
                    'message': f"Customer '{customer.customer_name}' added successfully!"
                })
                
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                })
        
        # Handle regular form submission
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save()
            messages.success(request, f"Customer '{customer.customer_name}' added successfully!")
            return redirect('customer_list')
    else:
        form = CustomerForm()
    
    context = {
        'form': form,
        'title': 'Add Customer'
    }
    return render(request, 'customers/customer_form.html', context)

@login_required
def update_customer(request, pk):
    customer = get_object_or_404(CustomerMaster, customerid=pk)
    
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, f"Customer '{customer.customer_name}' updated successfully!")
            return redirect('customer_list')
    else:
        form = CustomerForm(instance=customer)
    
    context = {
        'form': form,
        'customer': customer,
        'title': 'Update Customer'
    }
    return render(request, 'customers/customer_form.html', context)

@login_required
def customer_detail(request, pk):
    customer = get_object_or_404(CustomerMaster, customerid=pk)
    
    # Get invoices for this customer
    invoices = SalesInvoiceMaster.objects.filter(customerid=pk).order_by('-sales_invoice_date')
    
    # Calculate total sales and payment amounts
    # We need to calculate sales total through SalesMaster since sales_invoice_total is a property
    total_sales = 0
    total_paid = 0
    
    # Get all invoice numbers
    invoice_numbers = invoices.values_list('sales_invoice_no', flat=True)
    
    # Calculate total sales from SalesMaster
    total_sales = SalesMaster.objects.filter(
        sales_invoice_no__in=invoice_numbers
    ).aggregate(total=Sum('sale_total_amount'))['total'] or 0
    
    # Calculate total paid directly from SalesInvoiceMaster
    total_paid = invoices.aggregate(total=Sum('sales_invoice_paid'))['total'] or 0
    
    # Calculate balance
    balance = total_sales - total_paid
    
    context = {
        'customer': customer,
        'invoices': invoices,
        'total_sales': total_sales,
        'total_paid': total_paid,
        'balance': balance,
        'title': f'Customer: {customer.customer_name}'
    }
    return render(request, 'customers/customer_detail.html', context)

@login_required
def delete_customer(request, pk):
    # Check if user is admin (case-insensitive)
    if not request.user.user_type.lower() in ['admin']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('customer_list')
        
    customer = get_object_or_404(CustomerMaster, customerid=pk)
    
    if request.method == 'POST':
        customer_name = customer.customer_name
        try:
            customer.delete()
            messages.success(request, f"Customer '{customer_name}' deleted successfully!")
        except Exception as e:
            messages.error(request, f"Cannot delete customer. It is referenced by other records. Error: {str(e)}")
        return redirect('customer_list')
    
    context = {
        'customer': customer,
        'title': 'Delete Customer'
    }
    return render(request, 'customers/customer_confirm_delete.html', context)

# Purchase Invoice views
@login_required
def invoice_list(request):
    invoices = InvoiceMaster.objects.all().order_by('-invoice_date')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        invoices = invoices.filter(
            Q(invoice_no__icontains=search_query) | 
            Q(supplierid__supplier_name__icontains=search_query)
        )
    
    # Filter by date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            invoices = invoices.filter(invoice_date__range=[start_date, end_date])
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
    
    # Pagination
    paginator = Paginator(invoices, 10)
    page_number = request.GET.get('page')
    invoices = paginator.get_page(page_number)
    
    context = {
        'invoices': invoices,
        'search_query': search_query,
        'start_date': start_date if 'start_date' in locals() else '',
        'end_date': end_date if 'end_date' in locals() else '',
        'title': 'Purchase Invoice List'
    }
    return render(request, 'purchases/invoice_list.html', context)

@login_required
def add_invoice(request):
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        if form.is_valid():
            invoice = form.save(commit=False)
            invoice.invoice_paid = 0  # Initialize paid amount to 0
            invoice.save()
            messages.success(request, f"Purchase Invoice #{invoice.invoice_no} added successfully!")
            return redirect('invoice_detail', pk=invoice.invoiceid)
    else:
        form = InvoiceForm()
    
    context = {
        'form': form,
        'title': 'Add Purchase Invoice'
    }
    return render(request, 'purchases/invoice_form.html', context)

@login_required
def edit_invoice(request, pk):
    # Only allow POST requests to prevent accidental data deletion on page refresh
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    invoice = get_object_or_404(InvoiceMaster, invoiceid=pk)
    
    try:
        # Update invoice basic details
        invoice.invoice_no = request.POST.get('invoice_no')
        invoice.invoice_date = request.POST.get('invoice_date')
        invoice.supplierid_id = request.POST.get('supplierid')
        invoice.scroll_no = request.POST.get('scroll_no') or ''
        invoice.transport_charges = float(request.POST.get('transport_charges', 0))
        
        # Process products data if provided
        products_data = request.POST.get('products_data')
        if products_data:
            try:
                products = json.loads(products_data)
                
                # Get existing products for this invoice
                existing_products = list(PurchaseMaster.objects.filter(product_invoiceid=invoice))
                
                # Clear all existing products first
                for purchase in existing_products:
                    purchase.delete()
                
                # Add all products from the form (both existing and new)
                for product_data in products:
                    try:
                        product = ProductMaster.objects.get(productid=product_data['productid'])
                        
                        # Create purchase entry
                        purchase = PurchaseMaster(
                            product_supplierid=invoice.supplierid,
                            product_invoiceid=invoice,
                            product_invoice_no=invoice.invoice_no,
                            productid=product,
                            product_name=product.product_name,
                            product_company=product.product_company,
                            product_packing=product.product_packing,
                            product_batch_no=product_data.get('batch_no', ''),
                            product_expiry=product_data.get('expiry', ''),
                            product_MRP=float(product_data.get('mrp', 0)),
                            product_purchase_rate=float(product_data.get('purchase_rate', 0)),
                            product_quantity=float(product_data.get('quantity', 0)),
                            product_scheme=float(product_data.get('scheme', 0)),
                            product_discount_got=float(product_data.get('discount', 0)),
                            IGST=float(product_data.get('igst', 0)),
                            purchase_calculation_mode=product_data.get('calculation_mode', 'flat'),
                            product_transportation_charges=0
                        )
                        
                        # Calculate actual rate and total
                        if purchase.purchase_calculation_mode == 'flat':
                            purchase.actual_rate_per_qty = purchase.product_purchase_rate - (purchase.product_discount_got / purchase.product_quantity) if purchase.product_quantity > 0 else purchase.product_purchase_rate
                        else:
                            purchase.actual_rate_per_qty = purchase.product_purchase_rate * (1 - (purchase.product_discount_got / 100))
                        
                        purchase.product_actual_rate = purchase.actual_rate_per_qty
                        purchase.total_amount = purchase.product_actual_rate * purchase.product_quantity
                        purchase.save()
                        
                        # Update sale rates if provided
                        rate_A = product_data.get('rate_A')
                        rate_B = product_data.get('rate_B')
                        rate_C = product_data.get('rate_C')
                        
                        if any([rate_A, rate_B, rate_C]):
                            SaleRateMaster.objects.update_or_create(
                                productid=purchase.productid,
                                product_batch_no=purchase.product_batch_no,
                                defaults={
                                    'rate_A': float(rate_A) if rate_A else 0,
                                    'rate_B': float(rate_B) if rate_B else 0,
                                    'rate_C': float(rate_C) if rate_C else 0
                                }
                            )
                            
                    except ProductMaster.DoesNotExist:
                        continue

                
                # Recalculate invoice total
                new_total = PurchaseMaster.objects.filter(product_invoiceid=invoice).aggregate(
                    total=Sum('total_amount')
                )['total'] or 0
                
                invoice.invoice_total = new_total
                
            except json.JSONDecodeError:
                pass  # If products_data is invalid, just update basic fields
        
        invoice.save()
        
        messages.success(request, f'Invoice #{invoice.invoice_no} updated successfully!')
        
        return JsonResponse({
            'success': True,
            'message': f'Invoice #{invoice.invoice_no} updated successfully!'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def invoice_detail(request, pk):
    invoice = get_object_or_404(InvoiceMaster, invoiceid=pk)
    
    # Get all purchases under this invoice
    purchases = PurchaseMaster.objects.filter(product_invoiceid=pk)
    
    # Calculate the sum of all purchase entries
    purchases_total = purchases.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Calculate the difference between invoice total and sum of purchases
    # Transport charges should NOT be included in the invoice total
    invoice_pending = invoice.invoice_total - purchases_total
    
    # Get all payments for this invoice
    payments = InvoicePaid.objects.filter(ip_invoiceid=pk).order_by('-payment_date')
    
    # Get suppliers and products for modal
    suppliers = SupplierMaster.objects.all().order_by('supplier_name')
    products = ProductMaster.objects.all().order_by('product_name')
    
    context = {
        'invoice': invoice,
        'purchases': purchases,
        'payments': payments,
        'purchases_total': purchases_total,
        'invoice_pending': invoice_pending,
        'has_pending_entries': abs(invoice_pending) > 0.01,  # Using a small threshold to account for floating-point errors
        'suppliers': suppliers,
        'products': products,
        'title': f'Purchase Invoice #{invoice.invoice_no}'
    }
    return render(request, 'purchases/invoice_detail.html', context)

@login_required
def add_purchase(request, invoice_id):
    invoice = get_object_or_404(InvoiceMaster, invoiceid=invoice_id)
    
    if request.method == 'POST':
        form = PurchaseForm(request.POST)
        if form.is_valid():
            purchase = form.save(commit=False)
            
            # Set additional fields
            purchase.product_supplierid = invoice.supplierid
            purchase.product_invoiceid = invoice
            purchase.product_invoice_no = invoice.invoice_no
            
            # Get product details from the selected product
            product = purchase.productid
            purchase.product_name = product.product_name
            purchase.product_company = product.product_company
            purchase.product_packing = product.product_packing
            
            # Calculate actual rate based on discount and quantity
            if purchase.purchase_calculation_mode == 'flat':
                # Flat discount amount
                purchase.actual_rate_per_qty = purchase.product_purchase_rate - (purchase.product_discount_got / purchase.product_quantity)
            else:
                # Percentage discount
                purchase.actual_rate_per_qty = purchase.product_purchase_rate * (1 - (purchase.product_discount_got / 100))
            
            purchase.product_actual_rate = purchase.actual_rate_per_qty
            
            # Calculate total amount before transport charges
            purchase.total_amount = purchase.product_actual_rate * purchase.product_quantity
            
            # Check if adding this product would exceed the invoice total
            existing_purchases_total = PurchaseMaster.objects.filter(
                product_invoiceid=invoice
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            
            new_total = existing_purchases_total + purchase.total_amount
            
            if new_total > invoice.invoice_total:
                messages.error(
                    request, 
                    f"Adding this product would exceed the invoice total amount of ₹{invoice.invoice_total}. "
                    f"Current total: ₹{existing_purchases_total:.2f}, This product: ₹{purchase.total_amount:.2f}, "
                    f"New total would be: ₹{new_total:.2f}"
                )
                context = {
                    'form': form,
                    'invoice': invoice,
                    'title': 'Add Purchase'
                }
                return render(request, 'purchases/purchase_form.html', context)
            
            # Calculate and distribute transport charges
            if invoice.transport_charges > 0:
                # Count existing products plus this new one
                existing_purchases = list(PurchaseMaster.objects.filter(product_invoiceid=invoice))
                total_products = len(existing_purchases) + 1
                
                # Calculate transport share per product
                transport_share_per_product = invoice.transport_charges / total_products
                
                # Update this product's transport charges
                purchase.product_transportation_charges = transport_share_per_product
                
                # Add the transport share to the actual rate
                transport_per_unit = transport_share_per_product / purchase.product_quantity
                purchase.product_actual_rate = purchase.actual_rate_per_qty + transport_per_unit
                
                # Recalculate total amount with transport charges included
                purchase.total_amount = purchase.product_actual_rate * purchase.product_quantity
                
                # Update existing products to redistribute transport charges
                for prev_purchase in existing_purchases:
                    prev_purchase.product_transportation_charges = transport_share_per_product
                    transport_per_unit = transport_share_per_product / prev_purchase.product_quantity
                    prev_purchase.product_actual_rate = prev_purchase.actual_rate_per_qty + transport_per_unit
                    prev_purchase.total_amount = prev_purchase.product_actual_rate * prev_purchase.product_quantity
                    prev_purchase.save()
            else:
                purchase.product_transportation_charges = 0
            
            purchase.save()
            
            # Save batch-specific sale rates to SaleRateMaster
            rate_A = form.cleaned_data.get('rate_A')
            rate_B = form.cleaned_data.get('rate_B')
            rate_C = form.cleaned_data.get('rate_C')
            
            # Check if any of the rates were specified
            if rate_A is not None or rate_B is not None or rate_C is not None:
                
                # Check if a rate entry already exists for this product batch
                try:
                    batch_rate = SaleRateMaster.objects.get(
                        productid=product,
                        product_batch_no=purchase.product_batch_no
                    )
                    # Update existing entry
                    batch_rate.rate_A = rate_A
                    batch_rate.rate_B = rate_B
                    batch_rate.rate_C = rate_C
                    batch_rate.save()
                except SaleRateMaster.DoesNotExist:
                    # Create new entry
                    SaleRateMaster.objects.create(
                        productid=product,
                        product_batch_no=purchase.product_batch_no,
                        rate_A=rate_A,
                        rate_B=rate_B,
                        rate_C=rate_C
                    )
            
            messages.success(request, f"Purchase for {purchase.product_name} added successfully!")
            return redirect('invoice_detail', pk=invoice_id)
    else:
        form = PurchaseForm()
    
    context = {
        'form': form,
        'invoice': invoice,
        'title': 'Add Purchase'
    }
    return render(request, 'purchases/purchase_form.html', context)

@login_required
def edit_purchase(request, invoice_id, purchase_id):
    invoice = get_object_or_404(InvoiceMaster, invoiceid=invoice_id)
    purchase = get_object_or_404(PurchaseMaster, purchaseid=purchase_id)
    
    # Ensure this purchase belongs to the specified invoice
    if purchase.product_invoiceid.invoiceid != invoice.invoiceid:
        messages.error(request, "This purchase does not belong to the specified invoice.")
        return redirect('invoice_detail', pk=invoice_id)
    
    if request.method == 'POST':
        form = PurchaseForm(request.POST, instance=purchase)
        if form.is_valid():
            purchase = form.save(commit=False)
            
            # Get product details from the selected product
            product = purchase.productid
            purchase.product_name = product.product_name
            purchase.product_company = product.product_company
            purchase.product_packing = product.product_packing
            
            # Store old quantity for comparison
            old_purchase = PurchaseMaster.objects.get(purchaseid=purchase_id)
            old_total = old_purchase.total_amount
            
            # Calculate actual rate based on discount and quantity
            if purchase.purchase_calculation_mode == 'flat':
                # Flat discount amount
                purchase.actual_rate_per_qty = purchase.product_purchase_rate - (purchase.product_discount_got / purchase.product_quantity)
            else:
                # Percentage discount
                purchase.actual_rate_per_qty = purchase.product_purchase_rate * (1 - (purchase.product_discount_got / 100))
            
            purchase.product_actual_rate = purchase.actual_rate_per_qty
            
            # Calculate total amount without transport charges
            base_total = purchase.product_actual_rate * purchase.product_quantity
            
            # Check if updating this product would exceed the invoice total
            # Get the sum of all purchases for this invoice excluding the current one
            other_purchases_total = PurchaseMaster.objects.filter(
                product_invoiceid=invoice
            ).exclude(
                purchaseid=purchase_id
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            
            new_total = other_purchases_total + base_total
            
            if new_total > invoice.invoice_total:
                messages.error(
                    request, 
                    f"Updating this product would exceed the invoice total amount of ₹{invoice.invoice_total}. "
                    f"Other products total: ₹{other_purchases_total:.2f}, This product: ₹{base_total:.2f}, "
                    f"New total would be: ₹{new_total:.2f}"
                )
                # Restore form with current values
                try:
                    batch_rate = SaleRateMaster.objects.get(
                        productid=old_purchase.productid,
                        product_batch_no=old_purchase.product_batch_no
                    )
                    form = PurchaseForm(instance=old_purchase, initial={
                        'rate_A': batch_rate.rate_A,
                        'rate_B': batch_rate.rate_B,
                        'rate_C': batch_rate.rate_C
                    })
                except SaleRateMaster.DoesNotExist:
                    form = PurchaseForm(instance=old_purchase)
                
                context = {
                    'form': form,
                    'invoice': invoice,
                    'purchase': old_purchase,
                    'title': 'Edit Purchase',
                    'is_edit': True
                }
                return render(request, 'purchases/purchase_form.html', context)
            
            # Calculate and distribute transport charges
            if invoice.transport_charges > 0:
                # Get all purchases for this invoice excluding the current one
                other_purchases = list(PurchaseMaster.objects.filter(
                    product_invoiceid=invoice
                ).exclude(
                    purchaseid=purchase_id
                ))
                
                total_products = len(other_purchases) + 1
                
                # Calculate transport share per product
                transport_share_per_product = invoice.transport_charges / total_products
                
                # Update this product's transport charges
                purchase.product_transportation_charges = transport_share_per_product
                
                # Add transport share to actual rate
                transport_per_unit = transport_share_per_product / purchase.product_quantity
                purchase.product_actual_rate = purchase.actual_rate_per_qty + transport_per_unit
                
                # Recalculate total amount with transport charges included
                purchase.total_amount = purchase.product_actual_rate * purchase.product_quantity
                
                # Update other products to redistribute transport charges
                for other_purchase in other_purchases:
                    other_purchase.product_transportation_charges = transport_share_per_product
                    other_transport_per_unit = transport_share_per_product / other_purchase.product_quantity
                    other_purchase.product_actual_rate = other_purchase.actual_rate_per_qty + other_transport_per_unit
                    other_purchase.total_amount = other_purchase.product_actual_rate * other_purchase.product_quantity
                    other_purchase.save()
            else:
                purchase.product_transportation_charges = 0
                purchase.total_amount = base_total
            
            purchase.save()
            
            # Only recalculate invoice total if quantity or rate changed, not for expiry date changes
            # Check if this is just an expiry date change
            old_purchase_data = PurchaseMaster.objects.get(purchaseid=purchase_id)
            is_expiry_only_change = (
                old_purchase_data.product_quantity == purchase.product_quantity and
                old_purchase_data.product_purchase_rate == purchase.product_purchase_rate and
                old_purchase_data.product_discount_got == purchase.product_discount_got and
                old_purchase_data.product_expiry != purchase.product_expiry
            )
            
            # Only update invoice total if it's not just an expiry date change
            if not is_expiry_only_change:
                all_purchases = PurchaseMaster.objects.filter(product_invoiceid=invoice)
                new_invoice_total = sum(p.total_amount for p in all_purchases)
                invoice.invoice_total = new_invoice_total
                invoice.save()
            
            # Save batch-specific sale rates to SaleRateMaster
            rate_A = form.cleaned_data.get('rate_A')
            rate_B = form.cleaned_data.get('rate_B')
            rate_C = form.cleaned_data.get('rate_C')
            
            # Check if any of the rates were specified
            if rate_A is not None or rate_B is not None or rate_C is not None:
                # Default to product master rates if not specified
                if rate_A is None:
                    # Get rates from SaleRateMaster or default values
                    rate_A = 0.0
                    rate_B = 0.0
                    rate_C = 0.0
                
                # Check if a rate entry already exists for this product batch
                try:
                    batch_rate = SaleRateMaster.objects.get(
                        productid=product,
                        product_batch_no=purchase.product_batch_no
                    )
                    # Update existing entry
                    batch_rate.rate_A = rate_A
                    batch_rate.rate_B = rate_B
                    batch_rate.rate_C = rate_C
                    batch_rate.save()
                except SaleRateMaster.DoesNotExist:
                    # Create new entry
                    SaleRateMaster.objects.create(
                        productid=product,
                        product_batch_no=purchase.product_batch_no,
                        rate_A=rate_A,
                        rate_B=rate_B,
                        rate_C=rate_C
                    )
            
            messages.success(request, f"Purchase for {purchase.product_name} updated successfully!")
            return redirect('invoice_detail', pk=invoice_id)
    else:
        # Try to get current rates for this product and batch
        try:
            batch_rate = SaleRateMaster.objects.get(
                productid=purchase.productid,
                product_batch_no=purchase.product_batch_no
            )
            
            # Initialize the form with current values including rates
            form = PurchaseForm(instance=purchase, initial={
                'rate_A': batch_rate.rate_A,
                'rate_B': batch_rate.rate_B,
                'rate_C': batch_rate.rate_C
            })
        except SaleRateMaster.DoesNotExist:
            # No batch-specific rates found, use default form
            form = PurchaseForm(instance=purchase)
    
    context = {
        'form': form,
        'invoice': invoice,
        'purchase': purchase,
        'title': 'Edit Purchase',
        'is_edit': True
    }
    return render(request, 'purchases/purchase_form.html', context)

@login_required
def delete_purchase(request, invoice_id, purchase_id):
    invoice = get_object_or_404(InvoiceMaster, invoiceid=invoice_id)
    purchase = get_object_or_404(PurchaseMaster, purchaseid=purchase_id)
    
    # Ensure this purchase belongs to the specified invoice
    if purchase.product_invoiceid.invoiceid != invoice.invoiceid:
        messages.error(request, "This purchase does not belong to the specified invoice.")
        return redirect('invoice_detail', pk=invoice_id)
    
    # Check if user is admin (case-insensitive)
    if not request.user.user_type.lower() in ['admin']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('invoice_detail', pk=invoice_id)
    
    if request.method == 'POST':
        product_name = purchase.product_name
        try:
            # Before deleting, check if we need to redistribute transport charges
            if invoice.transport_charges > 0:
                # Get all other purchases for this invoice
                other_purchases = list(PurchaseMaster.objects.filter(
                    product_invoiceid=invoice
                ).exclude(
                    purchaseid=purchase_id
                ))
                
                # If there are other purchases, redistribute transport charges
                if other_purchases:
                    # Calculate new transport share per product
                    transport_share_per_product = invoice.transport_charges / len(other_purchases)
                    
                    # Update all other products with new transport charges
                    for other_purchase in other_purchases:
                        other_purchase.product_transportation_charges = transport_share_per_product
                        transport_per_unit = transport_share_per_product / other_purchase.product_quantity
                        other_purchase.product_actual_rate = other_purchase.actual_rate_per_qty + transport_per_unit
                        other_purchase.total_amount = other_purchase.product_actual_rate * other_purchase.product_quantity
                        other_purchase.save()
            
            # Now delete the purchase
            purchase.delete()
            
            # Recalculate and update invoice total
            remaining_purchases = PurchaseMaster.objects.filter(product_invoiceid=invoice)
            new_total = sum(p.total_amount for p in remaining_purchases)
            invoice.invoice_total = new_total
            invoice.save()
            
            messages.success(request, f"Purchase for {product_name} deleted successfully! Invoice total updated to ₹{new_total:.2f}")
        except Exception as e:
            messages.error(request, f"Cannot delete purchase. Error: {str(e)}")
        return redirect('invoice_detail', pk=invoice_id)
    
    context = {
        'invoice': invoice,
        'purchase': purchase,
        'title': 'Delete Purchase'
    }
    return render(request, 'purchases/purchase_confirm_delete.html', context)

@login_required
def add_invoice_payment(request, invoice_id):
    invoice = get_object_or_404(InvoiceMaster, invoiceid=invoice_id)
    
    if request.method == 'POST':
        # Handle AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                # Validate required fields
                payment_date = request.POST.get('payment_date')
                payment_amount_str = request.POST.get('payment_amount')
                payment_mode = request.POST.get('payment_mode')
                payment_ref_no = request.POST.get('payment_ref_no', '')
                
                if not payment_date:
                    return JsonResponse({
                        'success': False,
                        'error': 'Payment date is required'
                    })
                
                if not payment_amount_str:
                    return JsonResponse({
                        'success': False,
                        'error': 'Payment amount is required'
                    })
                
                if not payment_mode:
                    return JsonResponse({
                        'success': False,
                        'error': 'Payment mode is required'
                    })
                
                try:
                    payment_amount = float(payment_amount_str)
                except (ValueError, TypeError):
                    return JsonResponse({
                        'success': False,
                        'error': 'Invalid payment amount'
                    })
                
                # Validate payment amount
                balance = float(invoice.invoice_total) - float(invoice.invoice_paid)
                if payment_amount > balance + 0.01:  # Add small tolerance for floating point
                    return JsonResponse({
                        'success': False,
                        'error': f'Payment amount cannot exceed balance due of ₹{balance:.2f}'
                    })
                
                if payment_amount <= 0:
                    return JsonResponse({
                        'success': False,
                        'error': 'Payment amount must be greater than 0'
                    })
                
                # Parse date
                from datetime import datetime
                try:
                    parsed_date = datetime.strptime(payment_date, '%Y-%m-%d').date()
                except ValueError:
                    return JsonResponse({
                        'success': False,
                        'error': 'Invalid date format'
                    })
                
                # Create payment record
                from .models import InvoicePaid
                payment = InvoicePaid.objects.create(
                    ip_invoiceid=invoice,
                    payment_date=parsed_date,
                    payment_amount=payment_amount,
                    payment_mode=payment_mode,
                    payment_ref_no=payment_ref_no
                )
                
                # Update invoice paid amount
                invoice.invoice_paid += payment_amount
                invoice.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Payment of ₹{payment_amount:.2f} added successfully!'
                })
                
            except Exception as e:
                import traceback
                print(f"Payment error: {str(e)}")
                print(traceback.format_exc())
                return JsonResponse({
                    'success': False,
                    'error': f'Server error: {str(e)}'
                })
        
        # Handle regular form submission
        form = InvoicePaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.ip_invoiceid = invoice
            
            # Check if payment amount is valid
            if payment.payment_amount > (invoice.invoice_total - invoice.invoice_paid):
                messages.error(request, "Payment amount cannot exceed the remaining balance.")
                return redirect('add_invoice_payment', invoice_id=invoice_id)
            
            payment.save()
            
            # Update invoice paid amount
            invoice.invoice_paid += payment.payment_amount
            invoice.save()
            
            messages.success(request, f"Payment of {payment.payment_amount} added successfully!")
            return redirect('invoice_detail', pk=invoice_id)
    else:
        # Initialize form with current datetime
        from django.utils import timezone
        initial_data = {
            'payment_date': timezone.now().date()
        }
        form = InvoicePaymentForm(initial=initial_data)
    
    context = {
        'form': form,
        'invoice': invoice,
        'balance': invoice.invoice_total - invoice.invoice_paid,
        'title': 'Add Invoice Payment'
    }
    return render(request, 'purchases/payment_form.html', context)

@login_required
def edit_invoice_payment(request, invoice_id, payment_id):
    invoice = get_object_or_404(InvoiceMaster, invoiceid=invoice_id)
    payment = get_object_or_404(InvoicePaid, payment_id=payment_id, ip_invoiceid=invoice_id)
    
    # Store original payment amount to calculate difference
    original_amount = payment.payment_amount
    
    if request.method == 'POST':
        form = InvoicePaymentForm(request.POST, instance=payment)
        if form.is_valid():
            # Calculate the difference between new and old payment amount
            new_payment = form.save(commit=False)
            difference = new_payment.payment_amount - original_amount
            
            # Check if new amount would exceed the invoice total
            if invoice.invoice_paid + difference > invoice.invoice_total:
                messages.error(request, "Payment amount cannot exceed the invoice total.")
                return redirect('edit_invoice_payment', invoice_id=invoice_id, payment_id=payment_id)
            
            # Update payment
            new_payment.save()
            
            # Update invoice paid amount
            invoice.invoice_paid += difference
            invoice.save()
            
            messages.success(request, f"Payment updated successfully!")
            return redirect('invoice_detail', pk=invoice_id)
    else:
        form = InvoicePaymentForm(instance=payment)
    
    context = {
        'form': form,
        'invoice': invoice,
        'payment': payment,
        'balance': invoice.invoice_total - invoice.invoice_paid + payment.payment_amount,
        'is_edit': True,
        'title': 'Edit Invoice Payment'
    }
    return render(request, 'purchases/payment_form.html', context)

@login_required
def delete_invoice_payment(request, invoice_id, payment_id):
    invoice = get_object_or_404(InvoiceMaster, invoiceid=invoice_id)
    payment = get_object_or_404(InvoicePaid, payment_id=payment_id, ip_invoiceid=invoice_id)
    
    if request.method == 'POST':
        # Update invoice paid amount
        invoice.invoice_paid -= payment.payment_amount
        invoice.save()
        
        # Delete payment
        payment.delete()
        
        messages.success(request, "Payment deleted successfully!")
        return redirect('invoice_detail', pk=invoice_id)
    
    context = {
        'payment': payment,
        'invoice': invoice,
        'title': 'Delete Invoice Payment'
    }
    return render(request, 'purchases/payment_confirm_delete.html', context)

@login_required
def delete_invoice(request, pk):
    # Check if user is admin (case-insensitive)
    if not request.user.user_type.lower() in ['admin']:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': "You don't have permission to perform this action."})
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('invoice_list')
        
    invoice = get_object_or_404(InvoiceMaster, invoiceid=pk)
    
    if request.method == 'POST':
        try:
            invoice_no = invoice.invoice_no
            invoice.delete()
            
            # Handle AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True, 
                    'message': f"Purchase Invoice #{invoice_no} deleted successfully!"
                })
            
            messages.success(request, f"Purchase Invoice #{invoice_no} deleted successfully!")
        except Exception as e:
            # Handle AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False, 
                    'error': f"Cannot delete invoice. Error: {str(e)}"
                })
            
            messages.error(request, f"Cannot delete invoice. Error: {str(e)}")
        return redirect('invoice_list')
    
    context = {
        'invoice': invoice,
        'title': 'Delete Purchase Invoice'
    }
    return render(request, 'purchases/invoice_confirm_delete.html', context)

# Sales Invoice views
@login_required
def sales_invoice_list(request):
    invoices = SalesInvoiceMaster.objects.all().order_by('-sales_invoice_date')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        invoices = invoices.filter(
            Q(sales_invoice_no__icontains=search_query) | 
            Q(customerid__customer_name__icontains=search_query)
        )
    
    # Filter by date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            invoices = invoices.filter(sales_invoice_date__range=[start_date, end_date])
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
    
    # Pagination
    paginator = Paginator(invoices, 10)
    page_number = request.GET.get('page')
    invoices = paginator.get_page(page_number)
    
    context = {
        'invoices': invoices,
        'search_query': search_query,
        'start_date': start_date if 'start_date' in locals() else '',
        'end_date': end_date if 'end_date' in locals() else '',
        'title': 'Sales Invoice List'
    }
    return render(request, 'sales/sales_invoice_list.html', context)

@login_required
def add_sales_invoice(request):
    # Generate the preview invoice number in ABC format
    preview_invoice_no = generate_sales_invoice_number()
    
    if request.method == 'POST':
        form = SalesInvoiceForm(request.POST)
        if form.is_valid():
            invoice = form.save(commit=False)
            
            # Set the generated invoice number
            invoice.sales_invoice_no = preview_invoice_no
            
            # Initialize paid amount to 0
            invoice.sales_invoice_paid = 0
            
            # Note: We don't need to set sales_invoice_total anymore as it's now calculated dynamically from sales items
            
            invoice.save()
            messages.success(request, f"Sales Invoice #{invoice.sales_invoice_no} added successfully!")
            return redirect('sales_invoice_detail', pk=invoice.sales_invoice_no)
    else:
        form = SalesInvoiceForm()
    
    context = {
        'form': form,
        'title': 'Add Sales Invoice',
        'preview_invoice_no': preview_invoice_no
    }
    return render(request, 'sales/sales_invoice_form.html', context)

@login_required
def sales_invoice_detail(request, pk):
    invoice = get_object_or_404(SalesInvoiceMaster, sales_invoice_no=pk)
    
    # Get all sales under this invoice
    sales = SalesMaster.objects.filter(sales_invoice_no=pk)
    
    # Get all payments for this invoice
    payments = SalesInvoicePaid.objects.filter(sales_ip_invoice_no=pk).order_by('-sales_payment_date')
    
    # Get customers and products for edit modal
    customers = CustomerMaster.objects.all().order_by('customer_name')
    products = ProductMaster.objects.all().order_by('product_name')
    
    context = {
        'invoice': invoice,
        'sales': sales,
        'payments': payments,
        'customers': customers,
        'products': products,
        'title': f'Sales Invoice #{invoice.sales_invoice_no}'
    }
    return render(request, 'sales/sales_invoice_detail.html', context)

@login_required
def print_sales_bill(request, pk):
    invoice = get_object_or_404(SalesInvoiceMaster, sales_invoice_no=pk)
    
    # Get all sales under this invoice
    sales = SalesMaster.objects.filter(sales_invoice_no=pk)
    
    # Get all payments for this invoice
    payments = SalesInvoicePaid.objects.filter(sales_ip_invoice_no=pk).order_by('-sales_payment_date')
    
    # Get pharmacy details for the bill header
    try:
        pharmacy = Pharmacy_Details.objects.first()
    except Pharmacy_Details.DoesNotExist:
        pharmacy = None
    
    # Calculate totals and tax amounts
    subtotal = 0
    total_tax = 0
    
    for sale in sales:
        # Base price before tax
        base_price = sale.sale_rate * sale.sale_quantity
        
        # Apply discount
        if sale.sale_calculation_mode == 'flat':
            # Flat discount in rupees
            base_price_after_discount = base_price - sale.sale_discount
        else:
            # Percentage discount
            base_price_after_discount = base_price - (base_price * sale.sale_discount / 100)
        
        # Calculate tax amount
        tax_amount = base_price_after_discount * (sale.sale_igst / 100)
        
        subtotal += base_price_after_discount
        total_tax += tax_amount
    
    context = {
        'invoice': invoice,
        'sales': sales,
        'payments': payments,
        'pharmacy': pharmacy,
        'subtotal': subtotal,
        'total_tax': total_tax,
        'total': invoice.sales_invoice_total,
        'balance': invoice.balance_due,
        'title': f'Print Bill: {invoice.sales_invoice_no}'
    }
    return render(request, 'sales/print_sales_bill.html', context)

@login_required
def print_receipt(request, pk):
    """Print receipt in landscape format with pharmacy details"""
    invoice = get_object_or_404(SalesInvoiceMaster, sales_invoice_no=pk)
    
    # Get all sales under this invoice
    sales = SalesMaster.objects.filter(sales_invoice_no=pk)
    
    # Get pharmacy details - fetch all fields explicitly
    pharmacy = None
    try:
        pharmacy = Pharmacy_Details.objects.first()
        if pharmacy:
            print(f"Pharmacy found: {pharmacy.pharmaname}")
            print(f"Proprietor: {pharmacy.proprietorname}")
            print(f"Contact: {pharmacy.proprietorcontact}")
    except Pharmacy_Details.DoesNotExist:
        print("No pharmacy details found in database")
        pharmacy = None
    except Exception as e:
        print(f"Error fetching pharmacy details: {e}")
        pharmacy = None
    
    context = {
        'invoice': invoice,
        'sales': sales,
        'pharmacy': pharmacy,
        'title': f'Receipt - Invoice #{invoice.sales_invoice_no}'
    }
    return render(request, 'sales/print_receipt.html', context)

@login_required
def add_sale(request, invoice_id):
    invoice = get_object_or_404(SalesInvoiceMaster, sales_invoice_no=invoice_id)
    
    if request.method == 'POST':
        form = SalesForm(request.POST)
        if form.is_valid():
            sale = form.save(commit=False)
            
            # Set additional fields
            sale.sales_invoice_no = invoice
            sale.customerid = invoice.customerid
            
            # Check stock availability first
            from .utils import get_batch_stock_status
            batch_quantity, is_available = get_batch_stock_status(
                sale.productid.productid, sale.product_batch_no
            )
            
            # If product is out of stock, show an error and don't save
            if not is_available:
                messages.error(request, f"Cannot add sale. Product {sale.productid.product_name} with batch {sale.product_batch_no} is out of stock.")
                context = {
                    'form': form,
                    'invoice': invoice,
                    'title': 'Add Sale'
                }
                return render(request, 'sales/sales_form.html', context)
            
            # If not enough quantity available, show error and don't save
            if batch_quantity < sale.sale_quantity:
                messages.error(request, f"Cannot add sale. Only {batch_quantity} units available for product {sale.productid.product_name} with batch {sale.product_batch_no}.")
                context = {
                    'form': form,
                    'invoice': invoice,
                    'title': 'Add Sale'
                }
                return render(request, 'sales/sales_form.html', context)
            
            # Get product details from the selected product
            product = sale.productid
            sale.product_name = product.product_name
            sale.product_company = product.product_company
            sale.product_packing = product.product_packing
            
            # Ensure expiry date is in MM-YYYY format
            if sale.product_expiry:
                try:
                    # If it's in YYYY-MM-DD format, convert to MM-YYYY
                    if len(sale.product_expiry) == 10 and '-' in sale.product_expiry:
                        parts = sale.product_expiry.split('-')
                        if len(parts[0]) == 4:  # YYYY-MM-DD format
                            sale.product_expiry = f"{parts[1]}-{parts[0]}"
                except:
                    pass  # Keep original format if conversion fails
            
            # Get batch-specific rates if available
            batch_specific_rate = None
            if sale.product_batch_no:
                try:
                    batch_specific_rate = SaleRateMaster.objects.get(
                        productid=product, 
                        product_batch_no=sale.product_batch_no
                    )
                except SaleRateMaster.DoesNotExist:
                    batch_specific_rate = None
            
            # Get MRP from purchase record for this batch
            try:
                purchase_record = PurchaseMaster.objects.filter(
                    productid=product,
                    product_batch_no=sale.product_batch_no
                ).first()
                fallback_mrp = purchase_record.product_MRP if purchase_record else 0.0
            except:
                fallback_mrp = 0.0
            
            # Set appropriate rate based on customer type and selected rate type
            if form.cleaned_data.get('custom_rate') and sale.rate_applied == 'custom':
                # Use the custom rate provided
                sale.sale_rate = form.cleaned_data.get('custom_rate')
            elif sale.rate_applied == 'A':
                if batch_specific_rate:
                    sale.sale_rate = batch_specific_rate.rate_A
                else:
                    sale.sale_rate = fallback_mrp  # Fallback to MRP
            elif sale.rate_applied == 'B':
                if batch_specific_rate:
                    sale.sale_rate = batch_specific_rate.rate_B
                else:
                    sale.sale_rate = fallback_mrp  # Fallback to MRP
            elif sale.rate_applied == 'C':
                if batch_specific_rate:
                    sale.sale_rate = batch_specific_rate.rate_C
                else:
                    sale.sale_rate = fallback_mrp  # Fallback to MRP
            
            # Calculate base price for all units
            base_price = sale.sale_rate * sale.sale_quantity
            
            # Apply discount first
            if sale.sale_calculation_mode == 'flat':
                # Flat discount amount
                discounted_amount = base_price - sale.sale_discount
            else:
                # Percentage discount
                discounted_amount = base_price * (1 - (sale.sale_discount / 100))
            
            # Then add GST to the discounted amount
            sale.sale_total_amount = discounted_amount * (1 + (sale.sale_igst / 100))
            
            sale.save()
            
            messages.success(request, f"Sale for {sale.product_name} added successfully!")
            return redirect('sales_invoice_detail', pk=invoice_id)
    else:
        form = SalesForm()
    
    context = {
        'form': form,
        'invoice': invoice,
        'title': 'Add Sale'
    }
    return render(request, 'sales/sales_form.html', context)

@login_required
def edit_sale(request, invoice_id, sale_id):
    invoice = get_object_or_404(SalesInvoiceMaster, sales_invoice_no=invoice_id)
    sale = get_object_or_404(SalesMaster, id=sale_id)
    
    # Ensure this sale belongs to the specified invoice
    if sale.sales_invoice_no.sales_invoice_no != invoice.sales_invoice_no:
        messages.error(request, "This sale does not belong to the specified invoice.")
        return redirect('sales_invoice_detail', pk=invoice_id)
    
    if request.method == 'POST':
        form = SalesForm(request.POST, instance=sale)
        if form.is_valid():
            sale = form.save(commit=False)
            
            # Validate stock for edit operation
            from .stock_validation import validate_edit_sale_stock
            
            stock_validation = validate_edit_sale_stock(
                sale_id,
                sale.productid.productid,
                sale.product_batch_no,
                sale.sale_quantity
            )
            
            if not stock_validation['valid']:
                messages.error(request, f"Cannot update sale. {stock_validation['message']}")
                context = {
                    'form': form,
                    'invoice': invoice,
                    'sale': sale,
                    'title': 'Edit Sale',
                    'is_edit': True
                }
                return render(request, 'sales/sales_form.html', context)
            
            # Get product details from the selected product
            product = sale.productid
            sale.product_name = product.product_name
            sale.product_company = product.product_company
            sale.product_packing = product.product_packing
            
            # Ensure expiry date is in MM-YYYY format
            if sale.product_expiry:
                try:
                    # If it's in YYYY-MM-DD format, convert to MM-YYYY
                    if len(sale.product_expiry) == 10 and '-' in sale.product_expiry:
                        parts = sale.product_expiry.split('-')
                        if len(parts[0]) == 4:  # YYYY-MM-DD format
                            sale.product_expiry = f"{parts[1]}-{parts[0]}"
                except:
                    pass  # Keep original format if conversion fails
            
            # Get batch-specific rates if available
            batch_specific_rate = None
            if sale.product_batch_no:
                try:
                    batch_specific_rate = SaleRateMaster.objects.get(
                        productid=product, 
                        product_batch_no=sale.product_batch_no
                    )
                except SaleRateMaster.DoesNotExist:
                    batch_specific_rate = None
            
            # Get MRP from purchase record for this batch
            try:
                purchase_record = PurchaseMaster.objects.filter(
                    productid=product,
                    product_batch_no=sale.product_batch_no
                ).first()
                fallback_mrp = purchase_record.product_MRP if purchase_record else 0.0
            except:
                fallback_mrp = 0.0
            
            # Set appropriate rate based on customer type and selected rate type
            if form.cleaned_data.get('custom_rate') and sale.rate_applied == 'custom':
                # Use the custom rate provided
                sale.sale_rate = form.cleaned_data.get('custom_rate')
            elif sale.rate_applied == 'A':
                if batch_specific_rate:
                    sale.sale_rate = batch_specific_rate.rate_A
                else:
                    sale.sale_rate = fallback_mrp  # Fallback to MRP
            elif sale.rate_applied == 'B':
                if batch_specific_rate:
                    sale.sale_rate = batch_specific_rate.rate_B
                else:
                    sale.sale_rate = fallback_mrp  # Fallback to MRP
            elif sale.rate_applied == 'C':
                if batch_specific_rate:
                    sale.sale_rate = batch_specific_rate.rate_C
                else:
                    sale.sale_rate = fallback_mrp  # Fallback to MRP
            
            # Calculate base price for all units
            base_price = sale.sale_rate * sale.sale_quantity
            
            # Apply discount first
            if sale.sale_calculation_mode == 'flat':
                # Flat discount amount
                discounted_amount = base_price - sale.sale_discount
            else:
                # Percentage discount
                discounted_amount = base_price * (1 - (sale.sale_discount / 100))
            
            # Then add GST to the discounted amount
            sale.sale_total_amount = discounted_amount * (1 + (sale.sale_igst / 100))
            
            sale.save()
            
            messages.success(request, f"Sale for {sale.product_name} updated successfully!")
            return redirect('sales_invoice_detail', pk=invoice_id)
    else:
        form = SalesForm(instance=sale)
    
    context = {
        'form': form,
        'invoice': invoice,
        'sale': sale,
        'title': 'Edit Sale',
        'is_edit': True
    }
    return render(request, 'sales/sales_form.html', context)
@login_required
def delete_user(request, pk):
    if not request.user.user_type == 'admin':
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('user_list')
        
    user = get_object_or_404(Web_User, id=pk)
    
    if request.method == 'POST':
        username = user.username
        try:
            user.delete()
            messages.success(request, f"User '{username}' deleted successfully!")
        except Exception as e:
            messages.error(request, f"Cannot delete user. Error: {str(e)}")
        return redirect('user_list')
    
    context = {
        'user': user,
        'title': 'Delete User'
    }
    return render(request, 'user_confirm_delete.html', context)

@login_required
def edit_sales_invoice(request, pk):
    # Only allow POST requests to prevent accidental data deletion on page refresh
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    invoice = get_object_or_404(SalesInvoiceMaster, sales_invoice_no=pk)
    
    try:
        # Update invoice basic details (explicitly excluding primary key sales_invoice_no)
        # DO NOT update sales_invoice_no as it's the primary key and referenced by other tables
        invoice.sales_invoice_date = request.POST.get('sales_invoice_date')
        invoice.customerid_id = request.POST.get('customerid')
        
        # Process products data if provided
        products_data = request.POST.get('products_data')
        if products_data:
            try:
                products = json.loads(products_data)
                
                # Get existing products for this invoice
                existing_products = list(SalesMaster.objects.filter(sales_invoice_no=invoice))
                
                # Clear all existing products first
                for sale in existing_products:
                    sale.delete()
                
                # Add all products from the form (both existing and new)
                for product_data in products:
                    try:
                        product = ProductMaster.objects.get(productid=product_data['productid'])
                        
                        # Skip stock validation for edit mode - we're just updating existing invoice
                        # Stock was already validated when invoice was first created
                        
                        # Calculate total amount
                        base_price = float(product_data['sale_rate']) * float(product_data['quantity'])
                        discount = float(product_data.get('discount', 0))
                        igst = float(product_data.get('igst', 0))
                        
                        if product_data.get('calculation_mode', 'flat') == 'flat':
                            discounted_amount = base_price - discount
                        else:
                            discounted_amount = base_price * (1 - (discount / 100))
                        
                        total_amount = discounted_amount * (1 + (igst / 100))
                        
                        # Keep expiry date in MM-YYYY format for SalesMaster
                        expiry_date = product_data.get('expiry', '')
                        if expiry_date:
                            try:
                                # Ensure it's in MM-YYYY format
                                if len(expiry_date) == 10 and '-' in expiry_date:
                                    # Convert YYYY-MM-DD to MM-YYYY
                                    parts = expiry_date.split('-')
                                    if len(parts[0]) == 4:  # YYYY-MM-DD format
                                        expiry_date = f"{parts[1]}-{parts[0]}"
                                # If already in MM-YYYY format, keep as is
                            except:
                                pass  # Keep original format if conversion fails
                        
                        # Create sale entry
                        sale = SalesMaster(
                            sales_invoice_no=invoice,
                            customerid=invoice.customerid,
                            productid=product,
                            product_name=product.product_name,
                            product_company=product.product_company,
                            product_packing=product.product_packing,
                            product_batch_no=product_data.get('batch_no', ''),
                            product_expiry=expiry_date,
                            product_MRP=float(product_data.get('mrp', 0)),
                            sale_rate=float(product_data.get('sale_rate', 0)),
                            sale_quantity=float(product_data.get('quantity', 0)),
                            sale_scheme=float(product_data.get('scheme', 0)),
                            sale_discount=discount,
                            sale_calculation_mode=product_data.get('calculation_mode', 'flat'),
                            sale_igst=igst,
                            rate_applied=product_data.get('rate_applied', 'A'),
                            sale_total_amount=total_amount
                        )
                        sale.save()
                        
                    except ProductMaster.DoesNotExist:
                        continue
                
            except json.JSONDecodeError:
                pass  # If products_data is invalid, just update basic fields
        
        invoice.save()
        
        messages.success(request, f'Sales Invoice #{pk} updated successfully!')
        
        return JsonResponse({
            'success': True,
            'message': f'Sales Invoice #{pk} updated successfully!'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def delete_sales_invoice(request, pk):
    if not request.user.user_type.lower() in ['admin']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('sales_invoice_list')
        
    invoice = get_object_or_404(SalesInvoiceMaster, sales_invoice_no=pk)
    
    if request.method == 'POST':
        # Store invoice number before deletion with proper handling
        invoice_no = str(invoice.sales_invoice_no) if invoice.sales_invoice_no else str(pk)
        
        # Debug logging
        print(f"Deleting invoice: {invoice_no}")
        print(f"Invoice object: {invoice}")
        print(f"Invoice PK: {pk}")
        
        try:
            invoice.delete()
            messages.success(request, f"Sales Invoice #{invoice_no} deleted successfully!")
        except Exception as e:
            messages.error(request, f"Cannot delete invoice. Error: {str(e)}")
        return redirect('sales_invoice_list')
    
    context = {
        'invoice': invoice,
        'title': 'Delete Sales Invoice'
    }
    return render(request, 'sales/sales_invoice_confirm_delete.html', context)

@login_required
def delete_sale(request, invoice_id, sale_id):
    invoice = get_object_or_404(SalesInvoiceMaster, sales_invoice_no=invoice_id)
    sale = get_object_or_404(SalesMaster, id=sale_id)
    
    if sale.sales_invoice_no.sales_invoice_no != invoice.sales_invoice_no:
        messages.error(request, "This sale does not belong to the specified invoice.")
        return redirect('sales_invoice_detail', pk=invoice_id)
    
    if not request.user.user_type.lower() in ['admin']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('sales_invoice_detail', pk=invoice_id)
    
    if request.method == 'POST':
        product_name = sale.product_name
        try:
            sale.delete()
            messages.success(request, f"Sale for {product_name} deleted successfully!")
        except Exception as e:
            messages.error(request, f"Cannot delete sale. Error: {str(e)}")
        return redirect('sales_invoice_detail', pk=invoice_id)
    
    context = {
        'invoice': invoice,
        'sale': sale,
        'title': 'Delete Sale'
    }
    return render(request, 'sales/sale_confirm_delete.html', context)

@login_required
def add_invoice_with_products(request):
    # Import the function from combined_invoice_view
    from .combined_invoice_view import add_invoice_with_products as combined_view
    return combined_view(request)

@login_required
def add_sales_invoice_with_products(request):
    def convert_date_format(date_str):
        """Convert DDMM format to YYYY-MM-DD format"""
        from datetime import datetime
        
        if not date_str:
            return datetime.now().date()
        
        # Handle YYYY-MM-DD format (already correct)
        if len(date_str) == 10 and '-' in date_str:
            try:
                return datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Handle DDMM format
        if len(date_str) == 4 and date_str.isdigit():
            day = int(date_str[:2])
            month = int(date_str[2:4])
            year = datetime.now().year
            try:
                return datetime(year, month, day).date()
            except ValueError:
                return datetime.now().date()
        
        # Handle DD/MM format
        if '/' in date_str and len(date_str.split('/')) == 2:
            try:
                day, month = date_str.split('/')
                year = datetime.now().year
                return datetime(year, int(month), int(day)).date()
            except ValueError:
                return datetime.now().date()
        
        # Default to current date if format is unrecognized
        return datetime.now().date()
    
    if request.method == 'POST':
        try:
            # Debug: Print POST data
            print("POST Data:", request.POST)
            
            invoice_form = SalesInvoiceForm(request.POST)
            
            # Debug: Check form validity
            print("Form is valid:", invoice_form.is_valid())
            if not invoice_form.is_valid():
                print("Form errors:", invoice_form.errors)
                messages.error(request, f"Form validation failed: {invoice_form.errors}")
                
            if invoice_form.is_valid():
                # Create sales invoice
                invoice = invoice_form.save(commit=False)
                invoice.sales_invoice_no = generate_sales_invoice_number()
                invoice.sales_invoice_paid = 0
                
                # Convert date format if needed
                if hasattr(invoice, 'sales_invoice_date'):
                    if isinstance(invoice.sales_invoice_date, str):
                        invoice.sales_invoice_date = convert_date_format(invoice.sales_invoice_date)
                
                # Debug: Print invoice data before save
                print(f"Saving invoice: {invoice.sales_invoice_no}, Date: {invoice.sales_invoice_date}, Customer: {invoice.customerid}")
                
                invoice.save()
                print("Invoice saved successfully!")
                
                # Process products data
                products_data = request.POST.get('products_data')
                print(f"Products data received: {products_data}")
                
                sales_created_count = 0
                
                if products_data:
                    try:
                        products = json.loads(products_data)
                        print(f"Parsed {len(products)} products")
                        
                        sales_to_create = []
                        
                        # Validate all products first
                        for i, product_data in enumerate(products):
                            print(f"Processing product {i+1}: {product_data}")
                            
                            if not product_data.get('productid'):
                                print(f"Skipping product {i+1}: No product ID")
                                continue
                                
                            try:
                                product = ProductMaster.objects.get(productid=product_data['productid'])
                                print(f"Found product: {product.product_name}")
                            except ProductMaster.DoesNotExist:
                                error_msg = f"Product with ID {product_data['productid']} not found."
                                print(error_msg)
                                messages.error(request, error_msg)
                                continue
                            
                            # Check stock availability
                            batch_quantity, is_available = get_batch_stock_status(
                                product.productid, product_data['batch_no']
                            )
                            
                            sale_quantity = float(product_data['quantity'])
                            print(f"Stock check - Available: {batch_quantity}, Required: {sale_quantity}")
                            
                            if not is_available:
                                error_msg = f"Product {product.product_name} batch {product_data['batch_no']} is out of stock."
                                print(error_msg)
                                messages.error(request, error_msg)
                                continue
                            
                            if batch_quantity < sale_quantity:
                                error_msg = f"Insufficient stock for {product.product_name} batch {product_data['batch_no']}. Available: {batch_quantity}, Required: {sale_quantity}"
                                print(error_msg)
                                messages.error(request, error_msg)
                                continue
                            
                            # Calculate total amount
                            base_price = float(product_data['sale_rate']) * sale_quantity
                            discount = float(product_data.get('discount', 0))
                            igst = float(product_data.get('igst', 0))
                            
                            if product_data.get('calculation_mode', 'flat') == 'flat':
                                discounted_amount = base_price - discount
                            else:
                                discounted_amount = base_price * (1 - (discount / 100))
                            
                            total_amount = discounted_amount * (1 + (igst / 100))
                            
                            print(f"Calculated total: {total_amount}")
                            
                            # Convert expiry date to MM-YYYY format for SalesMaster
                            expiry_date = product_data.get('expiry', '')
                            if expiry_date:
                                try:
                                    # If it's in YYYY-MM-DD format, convert to MM-YYYY
                                    if len(expiry_date) == 10 and '-' in expiry_date:
                                        parts = expiry_date.split('-')
                                        if len(parts[0]) == 4:  # YYYY-MM-DD format
                                            expiry_formatted = f"{parts[1]}-{parts[0]}"
                                        else:
                                            expiry_formatted = expiry_date
                                    # If already in MM-YYYY format, keep as is
                                    elif len(expiry_date) == 7 and '-' in expiry_date:
                                        expiry_formatted = expiry_date
                                    else:
                                        expiry_formatted = expiry_date
                                except:
                                    expiry_formatted = expiry_date
                            else:
                                expiry_formatted = ''
                            
                            # Prepare sale object
                            sale_obj = SalesMaster(
                                sales_invoice_no=invoice,
                                customerid=invoice.customerid,
                                productid=product,
                                product_name=product.product_name,
                                product_company=product.product_company,
                                product_packing=product.product_packing,
                                product_batch_no=product_data['batch_no'],
                                product_expiry=expiry_formatted,
                                product_MRP=float(product_data['mrp']),
                                sale_rate=float(product_data['sale_rate']),
                                sale_quantity=sale_quantity,
                                sale_scheme=float(product_data.get('scheme', 0)),
                                sale_discount=discount,
                                sale_calculation_mode=product_data.get('calculation_mode', 'flat'),
                                sale_igst=igst,
                                rate_applied=product_data.get('rate_applied', 'A'),
                                sale_total_amount=total_amount
                            )
                            
                            sales_to_create.append(sale_obj)
                            print(f"Added sale object for {product.product_name}")
                        
                        # Bulk create all sales
                        if sales_to_create:
                            SalesMaster.objects.bulk_create(sales_to_create)
                            sales_created_count = len(sales_to_create)
                            print(f"Successfully created {sales_created_count} sales records")
                        else:
                            print("No valid products to create sales records")
                            
                    except json.JSONDecodeError as e:
                        error_msg = f"Invalid products data format: {str(e)}"
                        print(error_msg)
                        messages.error(request, error_msg)
                        return redirect('add_sales_invoice_with_products')
                    except Exception as e:
                        error_msg = f"Error processing products: {str(e)}"
                        print(error_msg)
                        messages.error(request, error_msg)
                        return redirect('add_sales_invoice_with_products')
                else:
                    print("No products data received")
                    messages.warning(request, "No products were added to the invoice.")
                
                success_msg = f"Sales Invoice #{invoice.sales_invoice_no} with {sales_created_count} products added successfully!"
                print(success_msg)
                messages.success(request, success_msg)
                # Redirect to invoice detail page after creating invoice
                return redirect('sales_invoice_detail', pk=invoice.sales_invoice_no)
            else:
                # Form validation failed
                print("Form validation failed")
                for field, errors in invoice_form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
                        
        except Exception as e:
            error_msg = f"Unexpected error: {str(e)}"
            print(error_msg)
            messages.error(request, error_msg)
            return redirect('add_sales_invoice_with_products')
    else:
        invoice_form = SalesInvoiceForm()
    
    # Get customers and products for dropdowns
    customers = CustomerMaster.objects.select_related().order_by('customer_name')
    products = ProductMaster.objects.only('productid', 'product_name', 'product_company').order_by('product_name')
    
    context = {
        'invoice_form': invoice_form,
        'customers': customers,
        'products': products,
        'preview_invoice_no': generate_sales_invoice_number(),
        'title': 'Add Sales Invoice with Products'
    }
    return render(request, 'sales/combined_sales_invoice_form.html', context)

@login_required
def add_sales_payment(request, invoice_id):
    invoice = get_object_or_404(SalesInvoiceMaster, sales_invoice_no=invoice_id)
    
    if request.method == 'POST':
        # Handle AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                # Validate required fields
                payment_date = request.POST.get('sales_payment_date')
                payment_amount_str = request.POST.get('sales_payment_amount')
                payment_mode = request.POST.get('sales_payment_mode')
                payment_ref_no = request.POST.get('sales_payment_ref_no', '')
                
                if not payment_date:
                    return JsonResponse({
                        'success': False,
                        'error': 'Payment date is required'
                    })
                
                if not payment_amount_str:
                    return JsonResponse({
                        'success': False,
                        'error': 'Payment amount is required'
                    })
                
                if not payment_mode:
                    return JsonResponse({
                        'success': False,
                        'error': 'Payment mode is required'
                    })
                
                try:
                    payment_amount = float(payment_amount_str)
                except (ValueError, TypeError):
                    return JsonResponse({
                        'success': False,
                        'error': 'Invalid payment amount'
                    })
                
                # Validate payment amount
                balance = float(invoice.sales_invoice_total) - float(invoice.sales_invoice_paid)
                if payment_amount > balance + 0.01:  # Add small tolerance for floating point
                    return JsonResponse({
                        'success': False,
                        'error': f'Payment amount cannot exceed balance due of ₹{balance:.2f}'
                    })
                
                if payment_amount <= 0:
                    return JsonResponse({
                        'success': False,
                        'error': 'Payment amount must be greater than 0'
                    })
                
                # Parse date
                from datetime import datetime
                try:
                    parsed_date = datetime.strptime(payment_date, '%Y-%m-%d').date()
                except ValueError:
                    return JsonResponse({
                        'success': False,
                        'error': 'Invalid date format'
                    })
                
                # Create payment record
                from .models import SalesInvoicePaid
                payment = SalesInvoicePaid.objects.create(
                    sales_ip_invoice_no=invoice,
                    sales_payment_date=parsed_date,
                    sales_payment_amount=payment_amount,
                    sales_payment_mode=payment_mode,
                    sales_payment_ref_no=payment_ref_no
                )
                
                # Update invoice paid amount
                invoice.sales_invoice_paid += payment_amount
                invoice.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Payment of ₹{payment_amount:.2f} added successfully!'
                })
                
            except Exception as e:
                import traceback
                print(f"Sales payment error: {str(e)}")
                print(traceback.format_exc())
                return JsonResponse({
                    'success': False,
                    'error': f'Server error: {str(e)}'
                })
        
        # Handle regular form submission
        form = SalesPaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.sales_ip_invoice_no = invoice
            
            if payment.sales_payment_amount > (invoice.sales_invoice_total - invoice.sales_invoice_paid):
                messages.error(request, "Payment amount cannot exceed the remaining balance.")
                return redirect('add_sales_payment', invoice_id=invoice_id)
            
            payment.save()
            invoice.sales_invoice_paid += payment.sales_payment_amount
            invoice.save()
            
            messages.success(request, f"Payment of {payment.sales_payment_amount} added successfully!")
            return redirect('sales_invoice_detail', pk=invoice_id)
    else:
        initial_data = {'sales_payment_date': timezone.now().date()}
        form = SalesPaymentForm(initial=initial_data)
    
    context = {
        'form': form,
        'invoice': invoice,
        'balance': invoice.sales_invoice_total - invoice.sales_invoice_paid,
        'title': 'Add Sales Payment'
    }
    return render(request, 'sales/payment_form.html', context)

@login_required
def edit_sales_payment(request, invoice_id, payment_id):
    invoice = get_object_or_404(SalesInvoiceMaster, sales_invoice_no=invoice_id)
    payment = get_object_or_404(SalesInvoicePaid, sales_payment_id=payment_id, sales_ip_invoice_no=invoice_id)
    
    original_amount = payment.sales_payment_amount
    
    if request.method == 'POST':
        form = SalesPaymentForm(request.POST, instance=payment)
        if form.is_valid():
            new_payment = form.save(commit=False)
            
            # No timezone handling needed for DateField
            
            difference = new_payment.sales_payment_amount - original_amount
            
            if invoice.sales_invoice_paid + difference > invoice.sales_invoice_total:
                messages.error(request, "Payment amount cannot exceed the invoice total.")
                return redirect('edit_sales_payment', invoice_id=invoice_id, payment_id=payment_id)
            
            new_payment.save()
            
            invoice.sales_invoice_paid += difference
            invoice.save()
            
            messages.success(request, f"Payment updated successfully!")
            return redirect('sales_invoice_detail', pk=invoice_id)
    else:
        form = SalesPaymentForm(instance=payment)
    
    context = {
        'form': form,
        'invoice': invoice,
        'payment': payment,
        'balance': invoice.sales_invoice_total - invoice.sales_invoice_paid + payment.sales_payment_amount,
        'is_edit': True,
        'title': 'Edit Sales Payment'
    }
    return render(request, 'sales/payment_form.html', context)

@login_required
def delete_sales_payment(request, invoice_id, payment_id):
    invoice = get_object_or_404(SalesInvoiceMaster, sales_invoice_no=invoice_id)
    payment = get_object_or_404(SalesInvoicePaid, sales_payment_id=payment_id, sales_ip_invoice_no=invoice_id)
    
    if request.method == 'POST':
        invoice.sales_invoice_paid -= payment.sales_payment_amount
        invoice.save()
        
        payment.delete()
        
        messages.success(request, "Payment deleted successfully!")
        return redirect('sales_invoice_detail', pk=invoice_id)
    
    context = {
        'payment': payment,
        'invoice': invoice,
        'title': 'Delete Sales Payment'
    }
    return render(request, 'sales/payment_confirm_delete.html', context)

# Purchase Return views
@login_required
def purchase_return_list(request):
    returns = ReturnInvoiceMaster.objects.all().order_by('-returninvoice_date')
    
    search_query = request.GET.get('search', '')
    if search_query:
        returns = returns.filter(
            Q(returninvoiceid__icontains=search_query) |
            Q(returnsupplierid__supplier_name__icontains=search_query)
        )
    
    # Filter by date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            returns = returns.filter(returninvoice_date__range=[start_date, end_date])
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
    
    paginator = Paginator(returns, 10)
    page_number = request.GET.get('page')
    returns_page = paginator.get_page(page_number)
    
    context = {
        'returns': returns_page,
        'search_query': search_query,
        'start_date': start_date if 'start_date' in locals() else '',
        'end_date': end_date if 'end_date' in locals() else '',
        'title': 'Purchase Returns'
    }
    return render(request, 'returns/purchase_return_list.html', context)

@login_required
def add_purchase_return(request):
    from datetime import datetime
    import json
    from django.db import transaction
    
    # Generate preview return ID
    today = datetime.now().date()
    count = ReturnInvoiceMaster.objects.filter(
        returninvoice_date=today
    ).count() + 1
    preview_id = f'PR-{today.strftime("%Y%m%d")}-{count:04d}'
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                form = PurchaseReturnInvoiceForm(request.POST)
                if form.is_valid():
                    # Create return invoice
                    return_invoice = form.save(commit=False)
                    return_invoice.returninvoiceid = preview_id
                    return_invoice.returninvoice_paid = 0
                    return_invoice.save()
                    
                    # Process products data
                    products_data = request.POST.get('products_data')
                    return_items_created = 0
                    
                    if products_data:
                        try:
                            products = json.loads(products_data)
                            
                            for product_data in products:
                                if not product_data.get('productid'):
                                    continue
                                    
                                try:
                                    product = ProductMaster.objects.get(productid=product_data['productid'])
                                except ProductMaster.DoesNotExist:
                                    messages.error(request, f"Product with ID {product_data['productid']} not found.")
                                    continue
                                
                                # Convert expiry date from MM-YYYY format to DateField
                                expiry_date = product_data.get('expiry', '')
                                if expiry_date:
                                    try:
                                        # Handle MM-YYYY format (from batch selection)
                                        if len(expiry_date) == 7 and expiry_date.count('-') == 1:
                                            month, year = expiry_date.split('-')
                                            expiry_formatted = datetime(int(year), int(month), 1).date()
                                        # Handle YYYY-MM-DD format
                                        elif len(expiry_date) == 10 and '-' in expiry_date:
                                            expiry_formatted = datetime.strptime(expiry_date, '%Y-%m-%d').date()
                                        else:
                                            expiry_formatted = datetime.strptime(expiry_date, '%Y-%m-%d').date()
                                    except (ValueError, TypeError):
                                        expiry_formatted = datetime.now().date()
                                else:
                                    expiry_formatted = datetime.now().date()
                                
                                # Calculate total amount
                                return_rate = float(product_data.get('return_rate', 0))
                                return_quantity = float(product_data.get('return_quantity', 0))
                                scheme = float(product_data.get('scheme', 0))
                                charges = float(product_data.get('charges', 0))
                                
                                subtotal = return_rate * return_quantity
                                after_scheme = subtotal - scheme
                                total_amount = after_scheme + charges
                                
                                # Create return item
                                return_item = ReturnPurchaseMaster.objects.create(
                                    returninvoiceid=return_invoice,
                                    returnproduct_supplierid=return_invoice.returnsupplierid,
                                    returnproductid=product,
                                    returnproduct_batch_no=product_data.get('batch_no', ''),
                                    returnproduct_expiry=expiry_formatted,
                                    returnproduct_MRP=float(product_data.get('mrp', 0)),
                                    returnproduct_purchase_rate=return_rate,
                                    returnproduct_quantity=return_quantity,
                                    returnproduct_scheme=scheme,
                                    returnproduct_charges=charges,
                                    returntotal_amount=total_amount,
                                    return_reason=product_data.get('reason', '')
                                )
                                
                                # STOCK UPDATE: Purchase return decreases stock
                                # Enhanced stock processing with better error handling
                                from .stock_manager import StockManager
                                try:
                                    stock_result = StockManager.process_purchase_return(return_item)
                                    
                                    if stock_result['success']:
                                        print(f"✅ Stock updated successfully: {stock_result['message']}")
                                        # Optional: Add success message for user feedback
                                        # messages.success(request, f"Stock updated: {stock_result['message']}")
                                    else:
                                        print(f"❌ Stock processing failed: {stock_result['message']}")
                                        
                                        # Handle different error types with appropriate user feedback
                                        if stock_result.get('error_type') == 'insufficient_stock':
                                            messages.error(request, f"Stock Error: {stock_result['message']}")
                                            # Could potentially prevent the return from being saved
                                        elif stock_result.get('error_type') == 'invalid_quantity':
                                            messages.error(request, f"Validation Error: {stock_result['message']}")
                                        else:
                                            messages.warning(request, f"Stock Warning: {stock_result['message']}")
                                            
                                except Exception as e:
                                    error_msg = f"Stock processing system error: {str(e)}"
                                    print(f"🔥 {error_msg}")
                                    messages.error(request, error_msg)
                                
                                return_items_created += 1
                            
                            # Update return invoice total
                            total_items = ReturnPurchaseMaster.objects.filter(
                                returninvoiceid=return_invoice
                            ).aggregate(Sum('returntotal_amount'))['returntotal_amount__sum'] or 0
                            
                            return_invoice.returninvoice_total = total_items + return_invoice.return_charges
                            return_invoice.save()
                            
                        except json.JSONDecodeError as e:
                            messages.error(request, f"Invalid products data format: {str(e)}")
                            return redirect('add_purchase_return')
                        except Exception as e:
                            messages.error(request, f"Error processing products: {str(e)}")
                            return redirect('add_purchase_return')
                    
                    success_msg = f"Purchase Return #{return_invoice.returninvoiceid} with {return_items_created} products created successfully! Stock updated."
                    messages.success(request, success_msg)
                    return redirect('purchase_return_detail', pk=return_invoice.returninvoiceid)
                else:
                    for field, errors in form.errors.items():
                        for error in errors:
                            messages.error(request, f"{field}: {error}")
                            
        except Exception as e:
            messages.error(request, f"Unexpected error: {str(e)}")
            return redirect('add_purchase_return')
    else:
        form = PurchaseReturnInvoiceForm()
    
    # Get suppliers and products for dropdowns
    suppliers = SupplierMaster.objects.all().order_by('supplier_name')
    products = ProductMaster.objects.all().order_by('product_name')
    
    context = {
        'form': form,
        'preview_id': preview_id,
        'suppliers': suppliers,
        'products': products,
        'title': 'Add Purchase Return with Products'
    }
    return render(request, 'returns/purchase_return_form.html', context)

@login_required
def purchase_return_detail(request, pk):
    return_invoice = get_object_or_404(ReturnInvoiceMaster, returninvoiceid=pk)
    return_items = ReturnPurchaseMaster.objects.filter(returninvoiceid=return_invoice)
    
    # Calculate totals
    items_total = return_items.aggregate(Sum('returntotal_amount'))['returntotal_amount__sum'] or 0
    
    # Get suppliers and products for edit modal
    suppliers = SupplierMaster.objects.all().order_by('supplier_name')
    products = ProductMaster.objects.all().order_by('product_name')
    
    context = {
        'return_invoice': return_invoice,
        'return_items': return_items,
        'items_total': items_total,
        'suppliers': suppliers,
        'products': products,
        'title': f'Purchase Return #{return_invoice.returninvoiceid}'
    }
    return render(request, 'returns/purchase_return_detail.html', context)

@login_required
def edit_purchase_return(request, pk):
    return_invoice = get_object_or_404(ReturnInvoiceMaster, returninvoiceid=pk)
    
    if request.method == 'POST':
        try:
            # Update basic return details
            return_invoice.returnsupplierid_id = request.POST.get('supplier_id')
            return_invoice.returninvoice_date = request.POST.get('return_date')
            return_invoice.return_charges = float(request.POST.get('return_charges', 0))
            
            # Process products data if provided
            products_data = request.POST.get('products_data')
            if products_data:
                try:
                    products = json.loads(products_data)
                    
                    # Clear existing return items
                    ReturnPurchaseMaster.objects.filter(returninvoiceid=return_invoice).delete()
                    
                    # Add updated products
                    total_items = 0
                    for product_data in products:
                        if not product_data.get('productid'):
                            continue
                            
                        try:
                            product = ProductMaster.objects.get(productid=product_data['productid'])
                        except ProductMaster.DoesNotExist:
                            continue
                        
                        # Calculate total amount
                        return_rate = float(product_data.get('return_rate', 0))
                        return_quantity = float(product_data.get('return_quantity', 0))
                        scheme = float(product_data.get('scheme', 0))
                        charges = float(product_data.get('charges', 0))
                        
                        subtotal = return_rate * return_quantity
                        after_scheme = subtotal - scheme
                        item_total = after_scheme + charges
                        
                        # Convert expiry date
                        expiry_date = product_data.get('expiry', '')
                        if expiry_date:
                            try:
                                if len(expiry_date) == 7 and expiry_date.count('-') == 1:
                                    month, year = expiry_date.split('-')
                                    expiry_formatted = datetime(int(year), int(month), 1).date()
                                else:
                                    expiry_formatted = datetime.strptime(expiry_date, '%Y-%m-%d').date()
                            except (ValueError, TypeError):
                                expiry_formatted = datetime.now().date()
                        else:
                            expiry_formatted = datetime.now().date()
                        
                        # Create return item
                        ReturnPurchaseMaster.objects.create(
                            returninvoiceid=return_invoice,
                            returnproduct_supplierid=return_invoice.returnsupplierid,
                            returnproductid=product,
                            returnproduct_batch_no=product_data.get('batch_no', ''),
                            returnproduct_expiry=expiry_formatted,
                            returnproduct_MRP=float(product_data.get('mrp', 0)),
                            returnproduct_purchase_rate=return_rate,
                            returnproduct_quantity=return_quantity,
                            returnproduct_scheme=scheme,
                            returnproduct_charges=charges,
                            returntotal_amount=item_total,
                            return_reason=product_data.get('reason', '')
                        )
                        
                        total_items += item_total
                    
                    # Update return invoice total
                    return_invoice.returninvoice_total = total_items + return_invoice.return_charges
                    
                except json.JSONDecodeError:
                    pass
            
            return_invoice.save()
            
            messages.success(request, f'Purchase Return #{return_invoice.returninvoiceid} updated successfully!')
            return redirect('purchase_return_detail', pk=pk)
            
        except Exception as e:
            messages.error(request, f'Error updating purchase return: {str(e)}')
            return redirect('purchase_return_detail', pk=pk)
    
    # For GET request, redirect to detail page
    return redirect('purchase_return_detail', pk=pk)

@login_required
def update_purchase_return_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        return_id = data.get('return_id')
        
        return_invoice = get_object_or_404(ReturnInvoiceMaster, returninvoiceid=return_id)
        
        # Update basic return details
        return_invoice.returnsupplierid_id = data.get('supplier_id')
        return_invoice.returninvoice_date = data.get('return_date')
        return_invoice.return_charges = float(data.get('return_charges', 0))
        
        # Clear existing return items
        ReturnPurchaseMaster.objects.filter(returninvoiceid=return_invoice).delete()
        
        # Add updated products
        total_items = 0
        products = data.get('products', [])
        
        for product_data in products:
            if not product_data.get('productid'):
                continue
                
            try:
                product = ProductMaster.objects.get(productid=product_data['productid'])
            except ProductMaster.DoesNotExist:
                continue
            
            # Calculate total amount
            return_rate = float(product_data.get('return_rate', 0))
            return_quantity = float(product_data.get('return_quantity', 0))
            scheme = float(product_data.get('scheme', 0))
            charges = float(product_data.get('charges', 0))
            
            subtotal = return_rate * return_quantity
            after_scheme = subtotal - scheme
            item_total = after_scheme + charges
            
            # Convert expiry date
            expiry_date = product_data.get('expiry', '')
            if expiry_date:
                try:
                    if len(expiry_date) == 7 and expiry_date.count('-') == 1:
                        month, year = expiry_date.split('-')
                        expiry_formatted = datetime(int(year), int(month), 1).date()
                    else:
                        expiry_formatted = datetime.strptime(expiry_date, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    expiry_formatted = datetime.now().date()
            else:
                expiry_formatted = datetime.now().date()
            
            # Create return item
            ReturnPurchaseMaster.objects.create(
                returninvoiceid=return_invoice,
                returnproduct_supplierid=return_invoice.returnsupplierid,
                returnproductid=product,
                returnproduct_batch_no=product_data.get('batch_no', ''),
                returnproduct_expiry=expiry_formatted,
                returnproduct_MRP=float(product_data.get('mrp', 0)),
                returnproduct_purchase_rate=return_rate,
                returnproduct_quantity=return_quantity,
                returnproduct_scheme=scheme,
                returnproduct_charges=charges,
                returntotal_amount=item_total,
                return_reason=product_data.get('reason', '')
            )
            
            total_items += item_total
        
        # Update return invoice total
        return_invoice.returninvoice_total = total_items + return_invoice.return_charges
        return_invoice.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Purchase Return #{return_invoice.returninvoiceid} updated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def delete_purchase_return(request, pk):
    if not request.user.user_type.lower() in ['admin']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('purchase_return_list')
    
    return_invoice = get_object_or_404(ReturnInvoiceMaster, returninvoiceid=pk)
    
    if request.method == 'POST':
        return_id = return_invoice.returninvoiceid  # Store ID before deletion
        try:
            return_invoice.delete()
            messages.success(request, f"Purchase Return #{return_id} deleted successfully!")
        except Exception as e:
            messages.error(request, f"Cannot delete return. Error: {str(e)}")
        return redirect('purchase_return_list')
    
    context = {
        'return_invoice': return_invoice,
        'title': 'Delete Purchase Return'
    }
    return render(request, 'returns/purchase_return_confirm_delete.html', context)

@login_required
def add_purchase_return_item(request, return_id):
    return_invoice = get_object_or_404(ReturnInvoiceMaster, returninvoiceid=return_id)
    
    if request.method == 'POST':
        form = PurchaseReturnForm(request.POST)
        if form.is_valid():
            return_item = form.save(commit=False)
            return_item.returninvoiceid = return_invoice
            return_item.returnproduct_supplierid = return_invoice.returnsupplierid
            
            # Calculate total amount
            return_item.returntotal_amount = return_item.returnproduct_purchase_rate * return_item.returnproduct_quantity
            return_item.save()
            
            # Update return invoice total
            total_items = ReturnPurchaseMaster.objects.filter(
                returninvoiceid=return_invoice
            ).aggregate(Sum('returntotal_amount'))['returntotal_amount__sum'] or 0
            
            return_invoice.returninvoice_total = total_items + return_invoice.return_charges
            return_invoice.save()
            
            messages.success(request, f"Return item for {return_item.returnproductid.product_name} added successfully!")
            return redirect('purchase_return_detail', pk=return_id)
    else:
        form = PurchaseReturnForm()
    
    context = {
        'form': form,
        'return_invoice': return_invoice,
        'title': 'Add Return Item'
    }
    return render(request, 'returns/purchase_return_item_form.html', context)

@login_required
def edit_purchase_return_item(request, return_id, item_id):
    return_invoice = get_object_or_404(ReturnInvoiceMaster, returninvoiceid=return_id)
    return_item = get_object_or_404(ReturnPurchaseMaster, returnpurchaseid=item_id)
    
    if return_item.returninvoiceid.returninvoiceid != return_invoice.returninvoiceid:
        messages.error(request, "This return item does not belong to the specified return.")
        return redirect('purchase_return_detail', pk=return_id)
    
    if request.method == 'POST':
        form = PurchaseReturnForm(request.POST, instance=return_item)
        if form.is_valid():
            return_item = form.save(commit=False)
            return_item.returntotal_amount = return_item.returnproduct_purchase_rate * return_item.returnproduct_quantity
            return_item.save()
            
            # Update return invoice total
            total_items = ReturnPurchaseMaster.objects.filter(
                returninvoiceid=return_invoice
            ).aggregate(Sum('returntotal_amount'))['returntotal_amount__sum'] or 0
            
            return_invoice.returninvoice_total = total_items + return_invoice.return_charges
            return_invoice.save()
            
            messages.success(request, f"Return item for {return_item.returnproductid.product_name} updated successfully!")
            return redirect('purchase_return_detail', pk=return_id)
    else:
        form = PurchaseReturnForm(instance=return_item)
    
    context = {
        'form': form,
        'return_invoice': return_invoice,
        'return_item': return_item,
        'title': 'Edit Return Item',
        'is_edit': True
    }
    return render(request, 'returns/purchase_return_item_edit_form.html', context)

@login_required
def delete_purchase_return_item(request, return_id, item_id):
    return_invoice = get_object_or_404(ReturnInvoiceMaster, returninvoiceid=return_id)
    return_item = get_object_or_404(ReturnPurchaseMaster, returnpurchaseid=item_id)
    
    if return_item.returninvoiceid.returninvoiceid != return_invoice.returninvoiceid:
        messages.error(request, "This return item does not belong to the specified return.")
        return redirect('purchase_return_detail', pk=return_id)
    
    if not request.user.user_type.lower() in ['admin']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('purchase_return_detail', pk=return_id)
    
    if request.method == 'POST':
        product_name = return_item.returnproductid.product_name if return_item.returnproductid else 'Unknown Product'
        try:
            return_item.delete()
            
            # Update return invoice total
            total_items = ReturnPurchaseMaster.objects.filter(
                returninvoiceid=return_invoice
            ).aggregate(Sum('returntotal_amount'))['returntotal_amount__sum'] or 0
            
            return_invoice.returninvoice_total = total_items + return_invoice.return_charges
            return_invoice.save()
            
            messages.success(request, f"Purchase Return item for {product_name} deleted successfully!")
        except Exception as e:
            messages.error(request, f"Cannot delete return item. Error: {str(e)}")
        return redirect('purchase_return_detail', pk=return_id)
    
    context = {
        'return_invoice': return_invoice,
        'return_item': return_item,
        'title': 'Delete Return Item'
    }
    return render(request, 'returns/purchase_return_item_confirm_delete.html', context)

@login_required
def sales_return_list(request):
    returns = ReturnSalesInvoiceMaster.objects.all().order_by('-return_sales_invoice_date')
    
    search_query = request.GET.get('search', '')
    if search_query:
        returns = returns.filter(
            Q(return_sales_invoice_no__icontains=search_query) |
            Q(return_sales_customerid__customer_name__icontains=search_query)
        )
    
    # Filter by date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            returns = returns.filter(return_sales_invoice_date__range=[start_date, end_date])
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
        
    paginator = Paginator(returns, 10)
    page_number = request.GET.get('page')
    returns_page = paginator.get_page(page_number)
    
    context = {
        'returns': returns_page,
        'search_query': search_query,
        'start_date': start_date if 'start_date' in locals() else '',
        'end_date': end_date if 'end_date' in locals() else '',
        'title': 'Sales Returns'
    }
    return render(request, 'returns/sales_return_list.html', context)

@login_required
def add_sales_return(request):
    from datetime import datetime
    import json
    from django.db import transaction
    
    def convert_date_format(date_str):
        """Convert DDMM format to YYYY-MM-DD format"""
        if not date_str:
            return datetime.now().date()
        
        # Handle YYYY-MM-DD format (already correct)
        if len(date_str) == 10 and '-' in date_str:
            try:
                return datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Handle DDMM format
        if len(date_str) == 4 and date_str.isdigit():
            day = int(date_str[:2])
            month = int(date_str[2:4])
            year = datetime.now().year
            try:
                return datetime(year, month, day).date()
            except ValueError:
                return datetime.now().date()
        
        # Handle DD/MM format
        if '/' in date_str and len(date_str.split('/')) == 2:
            try:
                day, month = date_str.split('/')
                year = datetime.now().year
                return datetime(year, int(month), int(day)).date()
            except ValueError:
                return datetime.now().date()
        
        # Default to current date if format is unrecognized
        return datetime.now().date()
    
    # Generate preview return ID
    today = datetime.now().date()
    count = ReturnSalesInvoiceMaster.objects.filter(
        return_sales_invoice_date=today
    ).count() + 1
    preview_id = f'SR-{today.strftime("%Y%m%d")}-{count:04d}'
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Convert date format
                return_date_str = request.POST.get('return_sales_invoice_date')
                return_date = convert_date_format(return_date_str)
                
                # Create return invoice
                return_invoice = ReturnSalesInvoiceMaster.objects.create(
                    return_sales_invoice_no=preview_id,
                    return_sales_invoice_date=return_date,
                    return_sales_customerid_id=request.POST.get('return_sales_customerid'),
                    return_sales_charges=float(request.POST.get('return_sales_charges', 0)),
                    return_sales_invoice_total=0,  # Will be calculated
                    return_sales_invoice_paid=0
                )
                
                # Process products data
                products_data = request.POST.get('products_data')
                return_items_created = 0
                total_amount = 0
                
                if products_data:
                    try:
                        products = json.loads(products_data)
                        
                        for product_data in products:
                            if not product_data.get('productid'):
                                continue
                                
                            try:
                                product = ProductMaster.objects.get(productid=product_data['productid'])
                            except ProductMaster.DoesNotExist:
                                messages.error(request, f"Product with ID {product_data['productid']} not found.")
                                continue
                            
                            # Calculate total amount
                            return_rate = float(product_data.get('return_rate', 0))
                            return_quantity = float(product_data.get('return_quantity', 0))
                            discount = float(product_data.get('discount', 0))
                            gst = float(product_data.get('gst', 0))
                            
                            base_price = return_rate * return_quantity
                            discounted_amount = base_price - discount
                            item_total = discounted_amount + (discounted_amount * gst / 100)
                            
                            # Convert expiry date format
                            expiry_date = product_data.get('expiry', '')
                            if expiry_date:
                                try:
                                    # If it's already in YYYY-MM-DD format, keep it
                                    if len(expiry_date) == 10 and '-' in expiry_date:
                                        expiry_formatted = expiry_date
                                    # If it's in MM-YYYY format, keep it as is (this is the expected format)
                                    elif len(expiry_date) == 7 and '-' in expiry_date:
                                        expiry_formatted = expiry_date
                                    else:
                                        # For other formats, try to convert but preserve MM-YYYY if possible
                                        converted_date = convert_date_format(expiry_date)
                                        expiry_formatted = converted_date.strftime('%Y-%m-%d')
                                except:
                                    expiry_formatted = expiry_date
                            else:
                                expiry_formatted = ''
                            
                            # Create return item
                            return_item = ReturnSalesMaster.objects.create(
                                return_sales_invoice_no=return_invoice,
                                return_customerid=return_invoice.return_sales_customerid,
                                return_productid=product,
                                return_product_name=product.product_name,
                                return_product_company=product.product_company,
                                return_product_packing=product.product_packing,
                                return_product_batch_no=product_data.get('batch_no', ''),
                                return_product_expiry=expiry_formatted,
                                return_product_MRP=float(product_data.get('mrp', 0)),
                                return_sale_rate=return_rate,
                                return_sale_quantity=return_quantity,
                                return_sale_discount=discount,
                                return_sale_calculation_mode='flat',
                                return_sale_igst=gst,
                                return_sale_total_amount=item_total
                            )
                            
                            # STOCK UPDATE: Sales return increases stock
                            # Enhanced stock processing with better error handling
                            from .stock_manager import StockManager
                            try:
                                stock_result = StockManager.process_sales_return(return_item)
                                
                                if stock_result['success']:
                                    print(f"✅ Stock updated successfully: {stock_result['message']}")
                                    # Optional: Add success message for user feedback
                                    # messages.success(request, f"Stock updated: {stock_result['message']}")
                                else:
                                    print(f"❌ Stock processing failed: {stock_result['message']}")
                                    
                                    # Handle different error types with appropriate user feedback
                                    if stock_result.get('error_type') == 'batch_not_found':
                                        messages.error(request, f"Batch Error: {stock_result['message']}")
                                    elif stock_result.get('error_type') == 'invalid_quantity':
                                        messages.error(request, f"Validation Error: {stock_result['message']}")
                                    else:
                                        messages.warning(request, f"Stock Warning: {stock_result['message']}")
                                        
                            except Exception as e:
                                error_msg = f"Stock processing system error: {str(e)}"
                                print(f"🔥 {error_msg}")
                                messages.error(request, error_msg)
                            
                            total_amount += item_total
                            return_items_created += 1
                        
                        # Update return invoice total
                        return_invoice.return_sales_invoice_total = total_amount + return_invoice.return_sales_charges
                        return_invoice.save()
                        
                    except json.JSONDecodeError as e:
                        messages.error(request, f"Invalid products data format: {str(e)}")
                        return redirect('add_sales_return')
                    except Exception as e:
                        messages.error(request, f"Error processing products: {str(e)}")
                        return redirect('add_sales_return')
                
                success_msg = f"Sales Return #{return_invoice.return_sales_invoice_no} with {return_items_created} products created successfully! Stock updated."
                messages.success(request, success_msg)
                return redirect('sales_return_detail', pk=return_invoice.return_sales_invoice_no)
            
        except Exception as e:
            messages.error(request, f"Unexpected error: {str(e)}")
            return redirect('add_sales_return')
    
    # Get customers and products for dropdowns
    customers = CustomerMaster.objects.all().order_by('customer_name')
    products = ProductMaster.objects.all().order_by('product_name')
    
    context = {
        'preview_id': preview_id,
        'customers': customers,
        'products': products,
        'title': 'Add Sales Return with Products'
    }
    return render(request, 'returns/sales_return_form.html', context)

@login_required
def sales_return_detail(request, pk):
    return_invoice = get_object_or_404(ReturnSalesInvoiceMaster, return_sales_invoice_no=pk)
    return_items = ReturnSalesMaster.objects.filter(return_sales_invoice_no=return_invoice)
    payments = ReturnSalesInvoicePaid.objects.filter(return_sales_ip_invoice_no=return_invoice)
    
    # Calculate totals
    items_total = return_items.aggregate(Sum('return_sale_total_amount'))['return_sale_total_amount__sum'] or 0
    
    # Get customers and products for edit modal
    customers = CustomerMaster.objects.all().order_by('customer_name')
    products = ProductMaster.objects.all().order_by('product_name')
    
    context = {
        'return_invoice': return_invoice,
        'return_items': return_items,
        'payments': payments,
        'items_total': items_total,
        'customers': customers,
        'products': products,
        'title': f'Sales Return #{return_invoice.return_sales_invoice_no}'
    }
    return render(request, 'returns/sales_return_detail.html', context)

@login_required
def delete_sales_return(request, pk):
    if not request.user.user_type.lower() in ['admin']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('sales_return_list')
    
    return_invoice = get_object_or_404(ReturnSalesInvoiceMaster, return_sales_invoice_no=pk)
    
    if request.method == 'POST':
        return_id = return_invoice.return_sales_invoice_no  # Store ID before deletion
        try:
            return_invoice.delete()
            messages.success(request, f"Sales Return #{return_id} deleted successfully!")
        except Exception as e:
            messages.error(request, f"Cannot delete return. Error: {str(e)}")
        return redirect('sales_return_list')
    
    context = {
        'return_invoice': return_invoice,
        'title': 'Delete Sales Return'
    }
    return render(request, 'returns/sales_return_confirm_delete.html', context)

@login_required
def add_sales_return_item(request, return_id):
    return_invoice = get_object_or_404(ReturnSalesInvoiceMaster, return_sales_invoice_no=return_id)
    
    if request.method == 'POST':
        form = SalesReturnForm(request.POST)
        if form.is_valid():
            return_item = form.save(commit=False)
            return_item.return_sales_invoice_no = return_invoice
            return_item.return_customerid = return_invoice.return_sales_customerid
            
            # Get product details
            product = return_item.return_productid
            return_item.return_product_name = product.product_name
            return_item.return_product_company = product.product_company
            return_item.return_product_packing = product.product_packing
            
            # Calculate total amount
            base_price = return_item.return_sale_rate * return_item.return_sale_quantity
            
            # Apply discount
            if return_item.return_sale_calculation_mode == 'flat':
                discounted_amount = base_price - return_item.return_sale_discount
            else:
                discounted_amount = base_price * (1 - (return_item.return_sale_discount / 100))
            
            # Add GST
            return_item.return_sale_total_amount = discounted_amount * (1 + (return_item.return_sale_igst / 100))
            return_item.save()
            
            # Update return invoice total
            total_items = ReturnSalesMaster.objects.filter(
                return_sales_invoice_no=return_invoice
            ).aggregate(Sum('return_sale_total_amount'))['return_sale_total_amount__sum'] or 0
            
            return_invoice.return_sales_invoice_total = total_items + return_invoice.return_sales_charges
            return_invoice.save()
            
            messages.success(request, f"Return item for {return_item.return_product_name} added successfully!")
            return redirect('sales_return_detail', pk=return_id)
    else:
        form = SalesReturnForm()
    
    context = {
        'form': form,
        'return_invoice': return_invoice,
        'title': 'Add Return Item'
    }
    return render(request, 'returns/sales_return_item_form.html', context)

@login_required
def edit_sales_return_item(request, return_id, item_id):
    return_invoice = get_object_or_404(ReturnSalesInvoiceMaster, return_sales_invoice_no=return_id)
    return_item = get_object_or_404(ReturnSalesMaster, return_sales_id=item_id)
    
    if return_item.return_sales_invoice_no.return_sales_invoice_no != return_invoice.return_sales_invoice_no:
        messages.error(request, "This return item does not belong to the specified return.")
        return redirect('sales_return_detail', pk=return_id)
    
    if request.method == 'POST':
        form = SalesReturnForm(request.POST, instance=return_item)
        if form.is_valid():
            return_item = form.save(commit=False)
            
            # Get product details
            product = return_item.return_productid
            return_item.return_product_name = product.product_name
            return_item.return_product_company = product.product_company
            return_item.return_product_packing = product.product_packing
            
            # Calculate total amount
            base_price = return_item.return_sale_rate * return_item.return_sale_quantity
            
            # Apply discount
            if return_item.return_sale_calculation_mode == 'flat':
                discounted_amount = base_price - return_item.return_sale_discount
            else:
                discounted_amount = base_price * (1 - (return_item.return_sale_discount / 100))
            
            # Add GST
            return_item.return_sale_total_amount = discounted_amount * (1 + (return_item.return_sale_igst / 100))
            return_item.save()
            
            # Update return invoice total
            total_items = ReturnSalesMaster.objects.filter(
                return_sales_invoice_no=return_invoice
            ).aggregate(Sum('return_sale_total_amount'))['return_sale_total_amount__sum'] or 0
            
            return_invoice.return_sales_invoice_total = total_items + return_invoice.return_sales_charges
            return_invoice.save()
            
            messages.success(request, f"Return item for {return_item.return_product_name} updated successfully!")
            return redirect('sales_return_detail', pk=return_id)
    else:
        form = SalesReturnForm(instance=return_item)
    
    context = {
        'form': form,
        'return_invoice': return_invoice,
        'return_item': return_item,
        'title': 'Edit Return Item',
        'is_edit': True
    }
    return render(request, 'returns/sales_return_item_edit_form.html', context)

@login_required
def delete_sales_return_item(request, return_id, item_id):
    return_invoice = get_object_or_404(ReturnSalesInvoiceMaster, return_sales_invoice_no=return_id)
    return_item = get_object_or_404(ReturnSalesMaster, return_sales_id=item_id)
    
    if return_item.return_sales_invoice_no.return_sales_invoice_no != return_invoice.return_sales_invoice_no:
        messages.error(request, "This return item does not belong to the specified return.")
        return redirect('sales_return_detail', pk=return_id)
    
    if not request.user.user_type.lower() in ['admin']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('sales_return_detail', pk=return_id)
    
    if request.method == 'POST':
        product_name = return_item.return_product_name or 'Unknown Product'
        try:
            return_item.delete()
            
            # Update return invoice total
            total_items = ReturnSalesMaster.objects.filter(
                return_sales_invoice_no=return_invoice
            ).aggregate(Sum('return_sale_total_amount'))['return_sale_total_amount__sum'] or 0
            
            return_invoice.return_sales_invoice_total = total_items + return_invoice.return_sales_charges
            return_invoice.save()
            
            messages.success(request, f"Sales Return item for {product_name} deleted successfully!")
        except Exception as e:
            messages.error(request, f"Cannot delete return item. Error: {str(e)}")
        return redirect('sales_return_detail', pk=return_id)
    
    context = {
        'return_invoice': return_invoice,
        'return_item': return_item,
        'title': 'Delete Return Item'
    }
    return render(request, 'returns/sales_return_item_confirm_delete.html', context)

@login_required
def get_sales_invoices_for_customer(request):
    customer_id = request.GET.get('customer_id')
    if not customer_id:
        return JsonResponse({'error': 'Customer ID is required'}, status=400)
    
    invoices = SalesInvoiceMaster.objects.filter(customerid=customer_id).order_by('-sales_invoice_date')
    invoice_list = [{'id': inv.sales_invoice_no, 'text': f"{inv.sales_invoice_no} - {inv.sales_invoice_date}"} for inv in invoices]
    
    return JsonResponse(invoice_list, safe=False)

@login_required
def get_sales_invoice_items(request):
    sales_invoice_no = request.GET.get('sales_invoice_no')
    if not sales_invoice_no:
        return JsonResponse({'error': 'Sales Invoice No is required'}, status=400)
        
    sales_items = SalesMaster.objects.filter(sales_invoice_no=sales_invoice_no)
    items_list = []
    for item in sales_items:
        items_list.append({
            'id': item.id,
            'product_name': item.product_name,
            'product_batch_no': item.product_batch_no,
            'sale_quantity': item.sale_quantity,
            'sale_rate': item.sale_rate,
            'sale_total_amount': item.sale_total_amount,
        })
        
    return JsonResponse(items_list, safe=False)

@login_required
def add_sales_return_payment(request, return_id):
    return_invoice = get_object_or_404(ReturnSalesInvoiceMaster, return_sales_invoice_no=return_id)
    
    if request.method == 'POST':
        form = SalesReturnPaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.return_sales_ip_invoice_no = return_invoice
            
            # Check if payment amount is valid
            if payment.return_sales_payment_amount > (return_invoice.return_sales_invoice_total - return_invoice.return_sales_invoice_paid):
                messages.error(request, "Payment amount cannot exceed the remaining balance.")
                return redirect('add_sales_return_payment', return_id=return_id)
            
            payment.save()
            
            # Update return invoice paid amount
            return_invoice.return_sales_invoice_paid += payment.return_sales_payment_amount
            return_invoice.save()
            
            messages.success(request, f"Payment of {payment.return_sales_payment_amount} added successfully!")
            return redirect('sales_return_detail', pk=return_id)
    else:
        form = SalesReturnPaymentForm()
    
    context = {
        'form': form,
        'return_invoice': return_invoice,
        'balance': return_invoice.return_sales_invoice_total - return_invoice.return_sales_invoice_paid,
        'title': 'Add Sales Return Payment'
    }
    return render(request, 'returns/sales_return_payment_form.html', context)

@login_required
def edit_sales_return_payment(request, return_id, payment_id):
    return_invoice = get_object_or_404(ReturnSalesInvoiceMaster, return_sales_invoice_no=return_id)
    payment = get_object_or_404(ReturnSalesInvoicePaid, return_sales_payment_id=payment_id, return_sales_ip_invoice_no=return_id)
    
    original_amount = payment.return_sales_payment_amount
    
    if request.method == 'POST':
        form = SalesReturnPaymentForm(request.POST, instance=payment)
        if form.is_valid():
            new_payment = form.save(commit=False)
            difference = new_payment.return_sales_payment_amount - original_amount
            
            if return_invoice.return_sales_invoice_paid + difference > return_invoice.return_sales_invoice_total:
                messages.error(request, "Payment amount cannot exceed the return total.")
                return redirect('edit_sales_return_payment', return_id=return_id, payment_id=payment_id)
            
            new_payment.save()
            
            return_invoice.return_sales_invoice_paid += difference
            return_invoice.save()
            
            messages.success(request, f"Payment updated successfully!")
            return redirect('sales_return_detail', pk=return_id)
    else:
        form = SalesReturnPaymentForm(instance=payment)
    
    context = {
        'form': form,
        'return_invoice': return_invoice,
        'payment': payment,
        'balance': return_invoice.return_sales_invoice_total - return_invoice.return_sales_invoice_paid + payment.return_sales_payment_amount,
        'is_edit': True,
        'title': 'Edit Sales Return Payment'
    }
    return render(request, 'returns/sales_return_payment_form.html', context)

@login_required
def update_sales_return_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        return_id = data.get('return_id')
        customer_id = data.get('customer_id')
        return_date = data.get('return_date')
        return_charges = float(data.get('return_charges', 0))
        products = data.get('products', [])
        
        # Use select_for_update to prevent database locking issues
        with transaction.atomic():
            return_invoice = ReturnSalesInvoiceMaster.objects.select_for_update().get(
                return_sales_invoice_no=return_id
            )
            
            # Update basic fields
            return_invoice.return_sales_customerid_id = customer_id
            return_invoice.return_sales_invoice_date = datetime.strptime(return_date, '%Y-%m-%d').date()
            return_invoice.return_sales_charges = return_charges
            
            # Delete existing items in a separate step
            existing_items = ReturnSalesMaster.objects.filter(return_sales_invoice_no=return_invoice)
            existing_items.delete()
            
            # Create new items
            total_amount = 0
            new_items = []
            
            for product_data in products:
                try:
                    product = ProductMaster.objects.get(productid=product_data['productid'])
                    
                    rate = float(product_data.get('return_rate', 0))
                    qty = float(product_data.get('return_quantity', 0))
                    discount = float(product_data.get('discount', 0))
                    gst = float(product_data.get('gst', 0))
                    
                    subtotal = rate * qty
                    after_discount = subtotal - discount
                    gst_amount = (after_discount * gst) / 100
                    item_total = after_discount + gst_amount
                    
                    new_item = ReturnSalesMaster(
                        return_sales_invoice_no=return_invoice,
                        return_customerid_id=customer_id,
                        return_productid=product,
                        return_product_name=product.product_name,
                        return_product_company=product.product_company,
                        return_product_packing=product.product_packing,
                        return_product_batch_no=product_data.get('batch_no', ''),
                        return_product_expiry=product_data.get('expiry', ''),
                        return_product_MRP=float(product_data.get('mrp', 0)),
                        return_sale_rate=rate,
                        return_sale_quantity=qty,
                        return_sale_discount=discount,
                        return_sale_calculation_mode='flat',
                        return_sale_igst=gst,
                        return_sale_total_amount=item_total,
                        return_reason=product_data.get('reason', '')
                    )
                    
                    new_items.append(new_item)
                    total_amount += item_total
                    
                except ProductMaster.DoesNotExist:
                    continue
                except Exception as item_error:
                    print(f"Error processing item: {item_error}")
                    continue
            
            # Bulk create all items at once
            if new_items:
                ReturnSalesMaster.objects.bulk_create(new_items)
            
            # Update total and save
            return_invoice.return_sales_invoice_total = total_amount + return_charges
            return_invoice.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Sales return {return_id} updated successfully!'
            })
            
    except ReturnSalesInvoiceMaster.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Sales return not found'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        print(f"Error in update_sales_return_api: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Database error: {str(e)}'
        }, status=500)

@login_required
def delete_sales_return_item_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    if not request.user.user_type.lower() in ['admin']:
        return JsonResponse({'success': False, 'error': "You don't have permission to perform this action."}, status=403)
    
    try:
        data = json.loads(request.body)
        return_id = data.get('return_id')
        item_id = data.get('item_id')
        
        if not return_id or not item_id:
            return JsonResponse({'success': False, 'error': 'Return ID and Item ID are required'}, status=400)
        
        with transaction.atomic():
            return_invoice = get_object_or_404(ReturnSalesInvoiceMaster, return_sales_invoice_no=return_id)
            return_item = get_object_or_404(ReturnSalesMaster, return_sales_id=item_id)
            
            # Verify item belongs to the return
            if return_item.return_sales_invoice_no.return_sales_invoice_no != return_invoice.return_sales_invoice_no:
                return JsonResponse({'success': False, 'error': 'Item does not belong to this return'}, status=400)
            
            # Delete the item
            return_item.delete()
            
            # Recalculate return total
            total_items = ReturnSalesMaster.objects.filter(
                return_sales_invoice_no=return_invoice
            ).aggregate(Sum('return_sale_total_amount'))['return_sale_total_amount__sum'] or 0
            
            return_invoice.return_sales_invoice_total = total_items + return_invoice.return_sales_charges
            return_invoice.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Item deleted successfully'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def delete_sales_return_payment(request, return_id, payment_id):
    return_invoice = get_object_or_404(ReturnSalesInvoiceMaster, return_sales_invoice_no=return_id)
    payment = get_object_or_404(ReturnSalesInvoicePaid, return_sales_payment_id=payment_id, return_sales_ip_invoice_no=return_id)
    
    if request.method == 'POST':
        return_invoice.return_sales_invoice_paid -= payment.return_sales_payment_amount
        return_invoice.save()
        
        payment.delete()
        
        messages.success(request, "Payment deleted successfully!")
        return redirect('sales_return_detail', pk=return_id)
    
    context = {
        'payment': payment,
        'return_invoice': return_invoice,
        'title': 'Delete Sales Return Payment'
    }
    return render(request, 'returns/sales_return_payment_confirm_delete.html', context)

@login_required
def inventory_list(request):
    from django.db.models import Avg, Case, When, FloatField, Sum
    from django.http import JsonResponse

    # Get search query and offset
    search_query = request.GET.get('search', '').strip()
    offset = int(request.GET.get('offset', 0))
    limit = 50  # Reduced for better performance

    # Base query for products
    products_query = ProductMaster.objects.all().order_by('product_name')
    
    # Enhanced search filter - startswith only
    if search_query:
        products_query = products_query.filter(
            Q(product_name__istartswith=search_query) |
            Q(product_company__istartswith=search_query) |
            Q(product_salt__istartswith=search_query)
        )
    
    # Get total count for "More" button logic
    total_products = products_query.count()
    
    # Get products with offset and limit
    products = products_query[offset:offset + limit]
    
    # Process results with detailed batch information
    inventory_data = []
    for product in products:
        try:
            from .utils import get_inventory_batches_info
            
            # Get stock status using the enhanced function
            stock_info = get_stock_status(product.productid)
            current_stock = stock_info.get('current_stock', 0)
            
            # Get all batches information for this product with rates and MRP
            batches_info = get_inventory_batches_info(product.productid)
            
            # Calculate average MRP from batches
            if batches_info:
                total_mrp = sum(batch['mrp'] for batch in batches_info if batch['mrp'] > 0)
                avg_mrp = total_mrp / len([b for b in batches_info if b['mrp'] > 0]) if any(b['mrp'] > 0 for b in batches_info) else 0
                
                # Get primary batch info for backward compatibility
                first_batch = batches_info[0]
                batch_no = first_batch['batch_no']
                expiry_date = first_batch['expiry']
                batch_rates = first_batch['rates']
            else:
                avg_mrp = 0
                batch_no = None
                expiry_date = None
                batch_rates = {'rate_A': 0, 'rate_B': 0, 'rate_C': 0}
            
            # Calculate stock value
            stock_value = current_stock * avg_mrp
            
            inventory_data.append({
                'product': product,
                'current_stock': current_stock,
                'avg_mrp': avg_mrp,
                'stock_value': stock_value,
                'batch_no': batch_no,
                'expiry_date': expiry_date,
                'batch_rates': batch_rates,
                'batches_info': batches_info,  # Add all batches info with rates and MRP
                'status': 'Out of Stock' if current_stock <= 0 else 'Low Stock' if current_stock < 10 else 'In Stock'
            })
            
        except Exception as e:
            print(f"Error processing inventory for {product.product_name}: {e}")
            # Fallback data with empty batch info
            inventory_data.append({
                'product': product,
                'current_stock': 0,
                'avg_mrp': 0,
                'stock_value': 0,
                'batch_no': None,
                'expiry_date': None,
                'batch_rates': {'rate_A': 0, 'rate_B': 0, 'rate_C': 0},
                'batches_info': [],
                'status': 'Error'
            })
    
    # Check if AJAX request for loading more products
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Return JSON response for AJAX requests
        from django.template.loader import render_to_string
        html_content = render_to_string('inventory/inventory_list_partial.html', {
            'inventory_data': inventory_data
        }, request=request)
        
        # Calculate stats for the new batch
        batch_total_value = sum(item['stock_value'] for item in inventory_data)
        batch_out_of_stock = sum(1 for item in inventory_data if item['current_stock'] <= 0)
        batch_low_stock = sum(1 for item in inventory_data if 0 < item['current_stock'] < 10)
        
        return JsonResponse({
            'success': True,
            'html': html_content,
            'has_more': (offset + limit) < total_products,
            'next_offset': offset + limit,
            'loaded_count': len(inventory_data),
            'total_count': total_products,
            'batch_stats': {
                'total_value': batch_total_value,
                'out_of_stock': batch_out_of_stock,
                'low_stock': batch_low_stock
            }
        })
    
    # Quick summary stats from current batch only (for performance)
    page_total_value = sum(item['stock_value'] for item in inventory_data)
    page_out_of_stock = sum(1 for item in inventory_data if item['current_stock'] <= 0)
    page_low_stock = sum(1 for item in inventory_data if 0 < item['current_stock'] < 10)
    
    context = {
        'inventory_data': inventory_data,
        'total_products': total_products,
        'page_total_value': page_total_value,
        'page_out_of_stock': page_out_of_stock,
        'page_low_stock': page_low_stock,
        'search_query': search_query,
        'has_more': (offset + limit) < total_products,
        'next_offset': offset + limit,
        'title': 'Inventory - All Products'
    }
    return render(request, 'inventory/inventory_list.html', context)

@login_required
def batch_inventory_report(request):
    from django.core.paginator import Paginator
    from django.db.models import Sum, Avg, Case, When, DecimalField
    
    # Search functionality
    search_query = request.GET.get('search', '')
    
    # Ultra-optimized query - single query with all calculations
    batches_query = PurchaseMaster.objects.select_related('productid').values(
        'productid__productid',
        'productid__product_name',
        'productid__product_company', 
        'productid__product_packing',
        'product_batch_no',
        'product_expiry'
    ).annotate(
        batch_purchased=Sum('product_quantity'),
        avg_mrp=Avg('product_MRP'),
        batch_sold=Sum('productid__salesmaster__sale_quantity', 
                      filter=Q(productid__salesmaster__product_batch_no=F('product_batch_no')))
    ).annotate(
        current_stock=Case(
            When(batch_sold__isnull=True, then=F('batch_purchased')),
            default=F('batch_purchased') - F('batch_sold')
        ),
        stock_value=Case(
            When(batch_sold__isnull=True, then=F('batch_purchased') * F('avg_mrp')),
            default=(F('batch_purchased') - F('batch_sold')) * F('avg_mrp'),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    ).filter(
        current_stock__gt=0
    ).order_by('productid__product_name', 'product_batch_no')
    
    if search_query:
        batches_query = batches_query.filter(
            Q(productid__product_name__icontains=search_query) |
            Q(productid__product_company__icontains=search_query) |
            Q(product_batch_no__icontains=search_query)
        )
    
    # Pagination with reasonable page size
    paginator = Paginator(batches_query, 100)
    page_number = request.GET.get('page')
    batches_page = paginator.get_page(page_number)
    
    # Process results - all calculations done at DB level
    inventory_data = []
    for batch in batches_page:
        inventory_data.append({
            'product_id': batch['productid__productid'],
            'product_name': batch['productid__product_name'],
            'product_company': batch['productid__product_company'],
            'product_packing': batch['productid__product_packing'],
            'batch_no': batch['product_batch_no'],
            'expiry': batch['product_expiry'],
            'mrp': batch['avg_mrp'] or 0,
            'stock': batch['current_stock'] or 0,
            'value': batch['stock_value'] or 0
        })
    
    # Page-level total value
    page_total_value = sum(item['value'] for item in inventory_data)
    
    context = {
        'inventory_data': inventory_data,
        'batches_page': batches_page,
        'page_total_value': page_total_value,
        'search_query': search_query,
        'title': 'Batch-wise Inventory Report'
    }
    return render(request, 'reports/batch_inventory_report.html', context)

@login_required
def dateexpiry_inventory_report(request):
    from datetime import datetime, timedelta
    from django.db.models import Sum, F, Case, When, Value, CharField, DateField, FloatField
    from django.db.models.functions import Coalesce
    from collections import defaultdict

    # Get filter parameters
    search_query = request.GET.get('search', '')
    expiry_from = request.GET.get('expiry_from', '')
    expiry_to = request.GET.get('expiry_to', '')

    # Base query for all purchase batches with stock > 0
    batch_query = PurchaseMaster.objects.select_related('productid').annotate(
        total_sold=Coalesce(
            Sum('productid__salesmaster__sale_quantity',
                filter=Q(productid__salesmaster__product_batch_no=F('product_batch_no'))),
            Value(0.0, output_field=FloatField())
        ),
        current_stock=F('product_quantity') - F('total_sold')
    ).filter(
        current_stock__gt=0  # Only include batches with stock
    ).values(
        'productid__product_name',
        'productid__product_company',
        'productid__product_packing',
        'product_batch_no',
        'product_expiry',
        'product_actual_rate',
        'product_MRP',
        'current_stock'
    )
    
    # Debug: Print first few items to see data structure


    # Apply filters at the database level
    if search_query:
        batch_query = batch_query.filter(
            Q(productid__product_name__icontains=search_query) |
            Q(productid__product_company__icontains=search_query)
        )

    if expiry_from:
        try:
            from_date = datetime.strptime(expiry_from, '%Y-%m-%d').date()
            batch_query = batch_query.filter(product_expiry__gte=from_date)
        except (ValueError, TypeError):
            pass  # Ignore invalid date format

    if expiry_to:
        try:
            to_date = datetime.strptime(expiry_to, '%Y-%m-%d').date()
            batch_query = batch_query.filter(product_expiry__lte=to_date)
        except (ValueError, TypeError):
            pass  # Ignore invalid date format

    # Group and process data in Python
    grouped_data = defaultdict(list)
    today = datetime.now().date()
    total_value = 0

    for item in batch_query:
        expiry_date = item['product_expiry']
        current_stock = item['current_stock']
        
        # Skip if no stock
        if current_stock <= 0:
            continue
            
        # Parse expiry date and convert to MM-YYYY format
        parsed_expiry_date = None
        expiry_mmyyyy = None
        
        if expiry_date:
            if isinstance(expiry_date, str):
                # Try different date formats
                for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%m-%Y', '%d%m%Y']:
                    try:
                        if fmt == '%m-%Y':
                            # Already in MM-YYYY format
                            temp_date = datetime.strptime(expiry_date, fmt)
                            parsed_expiry_date = temp_date.replace(day=1).date()  # Use first day for sorting
                            expiry_mmyyyy = expiry_date
                        else:
                            parsed_expiry_date = datetime.strptime(expiry_date, fmt).date()
                            expiry_mmyyyy = parsed_expiry_date.strftime('%m-%Y')
                        break
                    except (ValueError, TypeError):
                        continue
            elif hasattr(expiry_date, 'strftime'):
                parsed_expiry_date = expiry_date.date() if hasattr(expiry_date, 'date') else expiry_date
                expiry_mmyyyy = parsed_expiry_date.strftime('%m-%Y')
            else:
                parsed_expiry_date = expiry_date
                if parsed_expiry_date:
                    expiry_mmyyyy = parsed_expiry_date.strftime('%m-%Y')
        
        # Use MM-YYYY format for grouping
        if not expiry_mmyyyy:
            group_key = 'No Expiry Date'
            days_to_expiry = 999999  # Large number for sorting
            expiry_display = 'No Expiry Date'
        else:
            group_key = expiry_mmyyyy
            # Calculate days to expiry using last day of the month
            import calendar
            month, year = map(int, expiry_mmyyyy.split('-'))  # MM-YYYY format
            last_day = calendar.monthrange(year, month)[1]
            month_end_date = datetime(year, month, last_day).date()
            days_to_expiry = (month_end_date - today).days
            expiry_display = expiry_mmyyyy
            
        stock_value = current_stock * (item['product_actual_rate'] or 0)
        total_value += stock_value

        grouped_data[group_key].append({
            'product_name': item['productid__product_name'],
            'product_company': item['productid__product_company'],
            'product_packing': item['productid__product_packing'],
            'batch_no': item['product_batch_no'],
            'quantity': current_stock,
            'purchase_rate': item['product_actual_rate'] or 0,
            'mrp': item['product_MRP'] or 0,
            'value': stock_value,
            'days_to_expiry': days_to_expiry,
            'expiry_date': parsed_expiry_date,
            'expiry_mmyyyy': expiry_mmyyyy,
            'expiry_display': expiry_display,
        })

    # Create sorted groups for the template
    expiry_groups = []
    for group_key, products_list in grouped_data.items():
        group_total_value = sum(p['value'] for p in products_list)
        
        if group_key == 'No Expiry Date':
            days_to_expiry = 999999
            expiry_display = 'No Expiry Date'
            expiry_date = None
        else:
            # group_key is MM-YYYY string
            month, year = map(int, group_key.split('-'))  # MM-YYYY format
            import calendar
            last_day = calendar.monthrange(year, month)[1]
            month_end_date = datetime(year, month, last_day).date()
            days_to_expiry = (month_end_date - today).days
            expiry_display = group_key  # MM-YYYY format
            expiry_date = month_end_date
        
        expiry_groups.append({
            'expiry_date': expiry_date,
            'expiry_display': expiry_display,
            'days_to_expiry': days_to_expiry,
            'products': products_list,
            'total_value': group_total_value,
        })
    
    # Sort by expiry date (earliest first, no expiry last)
    expiry_groups.sort(key=lambda x: x['days_to_expiry'])

    context = {
        'expiry_data': expiry_groups,
        'total_value': total_value,
        'search_query': search_query,
        'expiry_from': expiry_from,
        'expiry_to': expiry_to,
        'title': 'Date-wise Inventory Report'
    }
    return render(request, 'reports/dateexpiry_inventory_report.html', context)

@login_required
def sales_report(request):
    from datetime import datetime
    from .sales_analytics import SalesAnalytics
    
    # Get date range from request
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    
    # Parse dates with defaults - show all available data by default
    today = datetime.now().date()
    try:
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            # Get the earliest sales invoice date or default to 6 months ago
            earliest_invoice = SalesInvoiceMaster.objects.order_by('sales_invoice_date').first()
            if earliest_invoice:
                start_date = earliest_invoice.sales_invoice_date
            else:
                start_date = today.replace(month=max(1, today.month-5), day=1)
        
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
    except ValueError:
        # Fallback to 6 months ago
        start_date = today.replace(month=max(1, today.month-5), day=1)
        end_date = today
    
    # Get comprehensive analytics
    analytics = SalesAnalytics(start_date, end_date)
    report_data = analytics.get_comprehensive_report()
    
    # Prepare context with all analytics data
    context = {
        'title': 'Enhanced Sales Analytics Report',
        'start_date': start_date,
        'end_date': end_date,
        'sales_invoices': report_data['invoices'],
        'total_sales': report_data['core_metrics']['total_sales'],
        'total_received': report_data['core_metrics']['total_received'],
        'total_pending': report_data['core_metrics']['total_pending'],
        'product_sales': report_data['product_analytics'],
        'customer_sales': report_data['customer_analytics'],
        'invoice_analysis': {
            'total_invoices': report_data['core_metrics']['total_invoices'],
            'paid_invoices': report_data['invoice_analysis']['paid_invoices'],
            'partial_paid': report_data['invoice_analysis']['partial_paid'],
            'unpaid_invoices': report_data['invoice_analysis']['unpaid_invoices'],
            'largest_invoice': report_data['invoice_analysis']['largest_invoice'],
            'smallest_invoice': report_data['invoice_analysis']['smallest_invoice'],
            'avg_invoice_value': report_data['core_metrics']['avg_invoice_value']
        },
        'category_sales': report_data['category_analytics'],
        'payment_analysis': {
            'collection_rate': report_data['core_metrics']['collection_rate'],
            'pending_rate': report_data['core_metrics']['pending_rate'],
            'avg_payment_per_invoice': report_data['core_metrics']['avg_invoice_value']
        },
        'realtime_stats': report_data['realtime_stats'],
        'daily_sales': report_data['daily_trend'],
        'monthly_sales': report_data['monthly_trend'],
        'top_products': report_data['top_performers']['top_products'],
        'top_customers': report_data['top_performers']['top_customers'],
        'sales_distribution': {
            'product_distribution': report_data['product_analytics'][:20],
            'customer_distribution': report_data['customer_analytics'][:20],
            'category_distribution': report_data['category_analytics'],
            'daily_trend': report_data['daily_trend'],
            'monthly_trend': report_data['monthly_trend']
        }
    }
    
    return render(request, 'reports/enhanced_sales_analytics.html', context)

@login_required
def purchase_report(request):
    from datetime import datetime
    from .purchase_analytics import PurchaseAnalytics
    
    # Get date range from request
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    
    # Parse dates with defaults - show all available data by default
    today = datetime.now().date()
    try:
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            # Get the earliest purchase invoice date or default to 6 months ago
            earliest_invoice = InvoiceMaster.objects.order_by('invoice_date').first()
            if earliest_invoice:
                start_date = earliest_invoice.invoice_date
            else:
                start_date = today.replace(month=max(1, today.month-5), day=1)
        
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
    except ValueError:
        # Fallback to 6 months ago
        start_date = today.replace(month=max(1, today.month-5), day=1)
        end_date = today
    
    # Get comprehensive analytics
    analytics = PurchaseAnalytics(start_date, end_date)
    report_data = analytics.get_comprehensive_report()
    
    # Prepare context with all analytics data
    context = {
        'title': 'Enhanced Purchase Analytics Report',
        'start_date': start_date,
        'end_date': end_date,
        'purchase_invoices': report_data['invoices'],
        'total_purchases': report_data['core_metrics']['total_purchases'],
        'total_paid': report_data['core_metrics']['total_paid'],
        'total_pending': report_data['core_metrics']['total_pending'],
        'product_purchases': report_data['product_analytics'],
        'supplier_purchases': report_data['supplier_analytics'],
        'invoice_analysis': {
            'total_invoices': report_data['core_metrics']['total_invoices'],
            'paid_invoices': report_data['invoice_analysis']['paid_invoices'],
            'partial_paid': report_data['invoice_analysis']['partial_paid'],
            'unpaid_invoices': report_data['invoice_analysis']['unpaid_invoices'],
            'largest_invoice': report_data['invoice_analysis']['largest_invoice'],
            'smallest_invoice': report_data['invoice_analysis']['smallest_invoice'],
            'avg_invoice_value': report_data['core_metrics']['avg_invoice_value']
        },
        'category_purchases': report_data['category_analytics'],
        'payment_analysis': {
            'payment_rate': report_data['core_metrics']['payment_rate'],
            'pending_rate': report_data['core_metrics']['pending_rate'],
            'payment_modes': report_data['payment_analysis']['payment_modes'],
            'avg_payment_days': report_data['payment_analysis']['avg_payment_days']
        },
        'realtime_stats': report_data['realtime_stats'],
        'daily_purchases': report_data['daily_trend'],
        'monthly_purchases': report_data['monthly_trend'],
        'top_products': report_data['top_performers']['top_products'],
        'top_suppliers': report_data['top_performers']['top_suppliers'],
        'purchase_distribution': {
            'product_distribution': report_data['product_analytics'][:20],
            'supplier_distribution': report_data['supplier_analytics'][:20],
            'category_distribution': report_data['category_analytics'],
            'daily_trend': report_data['daily_trend'],
            'monthly_trend': report_data['monthly_trend']
        }
    }
    
    return render(request, 'reports/enhanced_purchase_analytics.html', context)

@login_required
def financial_report(request):
    from datetime import datetime, timedelta
    from django.db.models import Sum, F

    # Get date range from request - no defaults, user must select
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    
    start_date = None
    end_date = None
    sales = 0
    purchases = 0
    
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            
            # Calculate financial data
            sales_invoices = SalesInvoiceMaster.objects.filter(
                sales_invoice_date__range=[start_date, end_date]
            )
            purchase_invoices = InvoiceMaster.objects.filter(
                invoice_date__range=[start_date, end_date]
            )

            # Calculate totals
            sales = SalesMaster.objects.filter(
                sales_invoice_no__in=sales_invoices
            ).aggregate(total=Sum('sale_total_amount'))['total'] or 0

            purchases = purchase_invoices.aggregate(total=Sum('invoice_total'))['total'] or 0
        except (ValueError, TypeError):
            pass
    
    # Placeholder for returns
    sales_returns = 0
    purchase_returns = 0
    
    # Calculate net figures
    net_sales = sales - sales_returns
    net_purchases = purchases - purchase_returns
    gross_profit = net_sales - net_purchases

    # Outstanding amounts (total, not date-filtered)
    total_receivables = 0
    total_payables = 0
    
    # Calculate receivables properly
    for invoice in SalesInvoiceMaster.objects.all():
        balance = invoice.sales_invoice_total - invoice.sales_invoice_paid
        if balance > 0:
            total_receivables += balance
    
    # Calculate payables
    total_payables = InvoiceMaster.objects.aggregate(
        total=Sum(F('invoice_total') - F('invoice_paid'))
    )['total'] or 0

    # Monthly sales for chart (last 12 months)
    monthly_sales = []
    current_date = datetime.now()
    for i in range(12):
        if i == 0:
            month_start = current_date.replace(day=1)
        else:
            if month_start.month == 1:
                month_start = month_start.replace(year=month_start.year-1, month=12, day=1)
            else:
                month_start = month_start.replace(month=month_start.month-1, day=1)
        
        if month_start.month == 12:
            month_end = month_start.replace(year=month_start.year+1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = month_start.replace(month=month_start.month+1, day=1) - timedelta(days=1)
        
        month_sales = SalesMaster.objects.filter(
            sales_invoice_no__sales_invoice_date__range=[month_start, month_end]
        ).aggregate(total=Sum('sale_total_amount'))['total'] or 0
        
        monthly_sales.insert(0, {
            'month': month_start,
            'total': month_sales
        })

    context = {
        'title': 'Financial Report',
        'start_date': start_date,
        'end_date': end_date,
        'sales': sales,
        'purchases': purchases,
        'sales_returns': sales_returns,
        'purchase_returns': purchase_returns,
        'net_sales': net_sales,
        'net_purchases': net_purchases,
        'gross_profit': gross_profit,
        'total_receivables': total_receivables,
        'total_payables': total_payables,
        'monthly_sales': monthly_sales,
        'outstanding_receivables': [],
        'outstanding_payables': [],
    }
    return render(request, 'reports/financial_report.html', context)

@login_required
def get_product_info(request):
    product_id = request.GET.get('product_id')
    batch_no = request.GET.get('batch_no')
    
    if not product_id:
        return JsonResponse({'error': 'Product ID is required'}, status=400)
    
    try:
        product = ProductMaster.objects.get(productid=product_id)
        
        # Get product HSN percent for GST
        product_hsn_percent = product.product_hsn_percent or '0'
        
        response_data = {
            'product_name': product.product_name,
            'product_company': product.product_company,
            'product_packing': product.product_packing,
            'product_hsn_percent': product_hsn_percent,
            'rate_A': 0,
            'rate_B': 0,
            'rate_C': 0,
            'product_expiry': '',
            'batch_stock_available': False,
            'batch_stock_quantity': 0
        }
        
        # If batch number is provided, get batch-specific information
        if batch_no:
            # Get batch-specific rates
            try:
                batch_rate = SaleRateMaster.objects.get(
                    productid=product,
                    product_batch_no=batch_no
                )
                response_data.update({
                    'rate_A': float(batch_rate.rate_A or 0),
                    'rate_B': float(batch_rate.rate_B or 0),
                    'rate_C': float(batch_rate.rate_C or 0)
                })
            except SaleRateMaster.DoesNotExist:
                pass
            
            # Get batch stock and expiry information
            try:
                purchase_record = PurchaseMaster.objects.filter(
                    productid=product,
                    product_batch_no=batch_no
                ).first()
                
                if purchase_record:
                    response_data['product_expiry'] = purchase_record.product_expiry
                    
                    # Check stock availability
                    batch_quantity, is_available = get_batch_stock_status(product_id, batch_no)
                    response_data.update({
                        'batch_stock_available': is_available,
                        'batch_stock_quantity': batch_quantity
                    })
            except Exception as e:
                pass
        
        return JsonResponse(response_data)
        
    except ProductMaster.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def get_batch_rates(request):
    product_id = request.GET.get('product_id')
    batch_no = request.GET.get('batch_no')
    
    if not product_id or not batch_no:
        return JsonResponse({
            'success': False,
            'error': 'Product ID and batch number are required'
        })
    
    try:
        # Get batch-specific rates from SaleRateMaster
        try:
            sale_rate = SaleRateMaster.objects.get(
                productid=product_id,
                product_batch_no=batch_no
            )
            rates = {
                'rate_A': float(sale_rate.rate_A or 0),
                'rate_B': float(sale_rate.rate_B or 0),
                'rate_C': float(sale_rate.rate_C or 0)
            }
        except SaleRateMaster.DoesNotExist:
            rates = {
                'rate_A': 0,
                'rate_B': 0,
                'rate_C': 0
            }
        
        return JsonResponse({
            'success': True,
            'rates': rates
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def get_product_by_barcode(request):
    barcode = request.GET.get('barcode')
    
    if not barcode:
        return JsonResponse({'error': 'Barcode is required'}, status=400)
    
    try:
        product = ProductMaster.objects.get(product_barcode=barcode)
        
        # Get the first available batch for this product
        purchase_record = PurchaseMaster.objects.filter(
            productid=product
        ).first()
        
        if not purchase_record:
            return JsonResponse({
                'error': f'No purchase records found for product {product.product_name}'
            }, status=404)
        
        # Get batch-specific rates if available
        batch_rates = {'rate_a': 0, 'rate_b': 0, 'rate_c': 0}
        try:
            sale_rate = SaleRateMaster.objects.get(
                productid=product,
                product_batch_no=purchase_record.product_batch_no
            )
            batch_rates = {
                'rate_a': float(sale_rate.rate_A or 0),
                'rate_b': float(sale_rate.rate_B or 0),
                'rate_c': float(sale_rate.rate_C or 0)
            }
        except SaleRateMaster.DoesNotExist:
            # Use MRP as fallback
            batch_rates = {
                'rate_a': float(purchase_record.product_MRP or 0),
                'rate_b': float(purchase_record.product_MRP or 0),
                'rate_c': float(purchase_record.product_MRP or 0)
            }
        
        return JsonResponse({
            'success': True,
            'product_id': product.productid,
            'product_name': product.product_name,
            'product_company': product.product_company,
            'batch_no': purchase_record.product_batch_no,
            'expiry': purchase_record.product_expiry,
            'mrp': float(purchase_record.product_MRP or 0),
            **batch_rates
        })
        
    except ProductMaster.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'Product with barcode {barcode} not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def get_product_batch_selector(request):
    product_id = request.GET.get('product_id')
    
    if not product_id:
        return JsonResponse({'error': 'Product ID is required'}, status=400)
    
    try:
        product = ProductMaster.objects.get(productid=product_id)
        
        # Get all unique batches for this product with stock information
        batches = PurchaseMaster.objects.filter(
            productid=product
        ).values(
            'product_batch_no',
            'product_expiry', 
            'product_MRP',
            'product_actual_rate'
        ).distinct()
        
        # Convert expiry date to MM-YYYY format
        def convert_expiry_to_mmyyyy(expiry_input):
            if not expiry_input:
                return ''
            
            expiry_str = str(expiry_input).strip()
            
            # Handle YYYY-MM-DD format (convert to MM-YYYY)
            if len(expiry_str) == 10 and expiry_str.count('-') == 2:
                parts = expiry_str.split('-')
                if len(parts) == 3 and len(parts[0]) == 4:
                    return f"{parts[1]}-{parts[0]}"
            
            # Handle MM-YYYY format (already correct)
            if len(expiry_str) == 7 and expiry_str.count('-') == 1:
                parts = expiry_str.split('-')
                if len(parts) == 2 and len(parts[0]) == 2 and len(parts[1]) == 4:
                    return expiry_str
            
            # Handle MMYY format (convert to MM-YYYY)
            if len(expiry_str) == 4 and expiry_str.isdigit():
                month = expiry_str[:2]
                year = '20' + expiry_str[2:4]
                return f"{month}-{year}"
            
            return expiry_str
        
        batch_list = []
        for batch in batches:
            batch_no = batch['product_batch_no']
            
            # Use the corrected stock calculation function
            current_stock, is_available = get_batch_stock_status(product_id, batch_no)
            
            # Only include batches with stock > 0
            if current_stock > 0:
                # Convert expiry to MM-YYYY format
                expiry_mmyyyy = convert_expiry_to_mmyyyy(batch['product_expiry'])
                
                # Get batch-specific rates
                rates = {'rate_A': 0, 'rate_B': 0, 'rate_C': 0}
                try:
                    sale_rate = SaleRateMaster.objects.get(
                        productid=product,
                        product_batch_no=batch_no
                    )
                    rates = {
                        'rate_A': float(sale_rate.rate_A or 0),
                        'rate_B': float(sale_rate.rate_B or 0),
                        'rate_C': float(sale_rate.rate_C or 0)
                    }
                except SaleRateMaster.DoesNotExist:
                    pass
                
                batch_list.append({
                    'batch_no': batch_no,
                    'expiry': expiry_mmyyyy,
                    'stock': current_stock,
                    'mrp': float(batch['product_MRP'] or 0),
                    'purchase_rate': float(batch['product_actual_rate'] or batch['product_MRP'] or 0),
                    'rate_a': rates['rate_A'],
                    'rate_b': rates['rate_B'],
                    'rate_c': rates['rate_C'],
                    'is_available': True
                })
        
        # Sort by expiry date (earliest first)
        batch_list.sort(key=lambda x: x['expiry'] if x['expiry'] else '9999-12')
        
        return JsonResponse({
            'success': True,
            'batches': batch_list
        })
        
    except ProductMaster.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'Product with ID {product_id} not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def get_product_batches(request):
    product_id = request.GET.get('product_id')
    
    if not product_id:
        return JsonResponse({'error': 'Product ID is required'}, status=400)
    
    try:
        batches = PurchaseMaster.objects.filter(
            productid=product_id
        ).values_list('product_batch_no', flat=True).distinct()
        
        return JsonResponse({
            'success': True,
            'batches': list(batches)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def get_batch_details(request):
    product_id = request.GET.get('product_id')
    batch_no = request.GET.get('batch_no')
    
    if not product_id or not batch_no:
        return JsonResponse({'error': 'Product ID and Batch No are required'}, status=400)
    
    try:
        batch = PurchaseMaster.objects.filter(
            productid=product_id,
            product_batch_no=batch_no
        ).first()
        
        if not batch:
            return JsonResponse({
                'success': False,
                'error': 'Batch not found'
            }, status=404)
        
        # Use the StockManager for accurate stock calculation
        from .stock_manager import StockManager
        batch_stock_info = StockManager._get_batch_stock(product_id, batch_no)
        current_stock = batch_stock_info['batch_stock']
        
        # Convert expiry date to MM-YYYY format
        def convert_expiry_to_mmyyyy(expiry_input):
            if not expiry_input:
                return ''
            
            expiry_str = str(expiry_input).strip()
            
            # Handle YYYY-MM-DD format (convert to MM-YYYY)
            if len(expiry_str) == 10 and expiry_str.count('-') == 2:
                parts = expiry_str.split('-')
                if len(parts) == 3 and len(parts[0]) == 4:
                    return f"{parts[1]}-{parts[0]}"
            
            # Handle MM-YYYY format (already correct)
            if len(expiry_str) == 7 and expiry_str.count('-') == 1:
                parts = expiry_str.split('-')
                if len(parts) == 2 and len(parts[0]) == 2 and len(parts[1]) == 4:
                    return expiry_str
            
            # Handle MMYY format (convert to MM-YYYY)
            if len(expiry_str) == 4 and expiry_str.isdigit():
                month = expiry_str[:2]
                year = '20' + expiry_str[2:4]
                return f"{month}-{year}"
            
            return expiry_str
        
        # Convert expiry to MM-YYYY format
        expiry_mmyyyy = convert_expiry_to_mmyyyy(batch.product_expiry)
        
        # Get batch-specific rates
        rates = {'rate_A': 0, 'rate_B': 0, 'rate_C': 0}
        try:
            sale_rate = SaleRateMaster.objects.get(
                productid=product_id,
                product_batch_no=batch_no
            )
            rates = {
                'rate_A': float(sale_rate.rate_A or 0),
                'rate_B': float(sale_rate.rate_B or 0),
                'rate_C': float(sale_rate.rate_C or 0)
            }
        except SaleRateMaster.DoesNotExist:
            pass
        
        return JsonResponse({
            'success': True,
            'batch_no': batch.product_batch_no,
            'expiry': expiry_mmyyyy,
            'mrp': float(batch.product_MRP or 0),
            'purchase_rate': float(batch.product_actual_rate or 0),
            'available_stock': current_stock,
            'rates': rates
        })
        
    except Exception as e:
        print(f"Error in get_batch_details: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def search_products_api(request):
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({
            'success': True,
            'products': []
        })
    
    try:
        products = ProductMaster.objects.filter(
            Q(product_name__icontains=query) |
            Q(product_company__icontains=query) |
            Q(product_barcode__icontains=query)
        ).values(
            'productid',
            'product_name',
            'product_company',
            'product_packing'
        )[:10]
        
        return JsonResponse({
            'success': True,
            'products': list(products)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def get_customer_rate_info(request):
    product_id = request.GET.get('product_id')
    batch_no = request.GET.get('batch_no')
    customer_type = request.GET.get('customer_type')
    
    if not all([product_id, batch_no, customer_type]):
        return JsonResponse({
            'success': False,
            'error': 'Product ID, batch number, and customer type are required'
        }, status=400)
    
    try:
        # Get product details for context
        try:
            product = ProductMaster.objects.get(productid=product_id)
            product_name = product.product_name
        except ProductMaster.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': f'Product with ID {product_id} not found'
            }, status=404)
        
        # Get batch-specific rates
        batch_rate = None
        try:
            batch_rate = SaleRateMaster.objects.get(
                productid=product_id,
                product_batch_no=batch_no
            )
            print(f"Found batch-specific rates for {product_name} batch {batch_no}: A={batch_rate.rate_A}, B={batch_rate.rate_B}, C={batch_rate.rate_C}")
        except SaleRateMaster.DoesNotExist:
            print(f"No batch-specific rates found for {product_name} batch {batch_no}")
        
        # Get MRP as fallback
        purchase_record = PurchaseMaster.objects.filter(
            productid=product_id,
            product_batch_no=batch_no
        ).first()
        
        mrp = float(purchase_record.product_MRP) if purchase_record and purchase_record.product_MRP else 0.0
        purchase_rate = float(purchase_record.product_actual_rate) if purchase_record and purchase_record.product_actual_rate else mrp
        
        # Determine rates with proper fallback hierarchy
        if batch_rate:
            rate_A = float(batch_rate.rate_A) if batch_rate.rate_A and batch_rate.rate_A > 0 else mrp
            rate_B = float(batch_rate.rate_B) if batch_rate.rate_B and batch_rate.rate_B > 0 else rate_A
            rate_C = float(batch_rate.rate_C) if batch_rate.rate_C and batch_rate.rate_C > 0 else rate_B
        else:
            # No batch rates available, use MRP for all
            rate_A = mrp
            rate_B = mrp
            rate_C = mrp
        
        # Log the rate matching process
        print(f"Rate matching for customer type '{customer_type}':")
        print(f"  Product: {product_name}")
        print(f"  Batch: {batch_no}")
        print(f"  MRP: ₹{mrp}")
        print(f"  Rate A: ₹{rate_A}")
        print(f"  Rate B: ₹{rate_B}")
        print(f"  Rate C: ₹{rate_C}")
        
        response_data = {
            'success': True,
            'rate_A': rate_A,
            'rate_B': rate_B,
            'rate_C': rate_C,
            'mrp': mrp,
            'purchase_rate': purchase_rate,
            'product_name': product_name,
            'batch_no': batch_no,
            'customer_type': customer_type,
            'has_batch_rates': batch_rate is not None,
            'rate_source': 'batch_specific' if batch_rate else 'mrp_fallback'
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        print(f"Error in get_customer_rate_info: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Server error: {str(e)}'
        }, status=500)

@login_required
def export_inventory_csv(request):
    return JsonResponse({'status': 'gtt'})

# Finance - Payments
@login_required
def payment_list(request):
    payments = PaymentMaster.objects.all().order_by('-payment_date')
    
    search_query = request.GET.get('search', '')
    if search_query:
        payments = payments.filter(
            Q(payment_ref_no__icontains=search_query) |
            Q(payment_mode__icontains=search_query)
        )
    
    paginator = Paginator(payments, 10)
    page_number = request.GET.get('page')
    payments = paginator.get_page(page_number)
    
    context = {
        'payments': payments,
        'search_query': search_query,
        'title': 'Payment List'
    }
    return render(request, 'finance/payment_list.html', context)

@login_required
def add_payment(request):
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            payment = form.save()
            messages.success(request, f"Payment of ₹{payment.payment_amount} added successfully!")
            return redirect('payment_list')
    else:
        form = PaymentForm()
    
    context = {
        'form': form,
        'title': 'Add Payment'
    }
    return render(request, 'finance/payment_form.html', context)

@login_required
def edit_payment(request, payment_id):
    payment = get_object_or_404(PaymentMaster, payment_id=payment_id)
    
    if request.method == 'POST':
        form = PaymentForm(request.POST, instance=payment)
        if form.is_valid():
            form.save()
            messages.success(request, "Payment updated successfully!")
            return redirect('payment_list')
    else:
        form = PaymentForm(instance=payment)
    
    context = {
        'form': form,
        'payment': payment,
        'title': 'Edit Payment',
        'is_edit': True
    }
    return render(request, 'finance/payment_form.html', context)

@login_required
def delete_payment(request, pk):
    payment = get_object_or_404(PaymentMaster, payment_id=pk)
    
    if request.method == 'POST':
        try:
            payment.delete()
            messages.success(request, "Payment deleted successfully!")
        except Exception as e:
            messages.error(request, f"Cannot delete payment. Error: {str(e)}")
        return redirect('payment_list')
    
    context = {
        'payment': payment,
        'title': 'Delete Payment'
    }
    return render(request, 'finance/payment_confirm_delete.html', context)

@login_required
def export_payments_pdf(request):
    from django.http import HttpResponse
    from datetime import datetime
    
    def convert_date_format(date_str):
        """Convert DDMM format to YYYY-MM-DD format"""
        if not date_str:
            return None
        
        # Handle YYYY-MM-DD format (already correct)
        if len(date_str) == 10 and '-' in date_str:
            try:
                return datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Handle DDMM format
        if len(date_str) == 4 and date_str.isdigit():
            day = int(date_str[:2])
            month = int(date_str[2:4])
            year = datetime.now().year
            try:
                return datetime(year, month, day).date()
            except ValueError:
                return None
        
        # Handle DD/MM format
        if '/' in date_str and len(date_str.split('/')) == 2:
            try:
                day, month = date_str.split('/')
                year = datetime.now().year
                return datetime(year, int(month), int(day)).date()
            except ValueError:
                return None
        
        return None
    
    # Get date filters from request
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    
    # Convert dates
    start_date = convert_date_format(start_date_str)
    end_date = convert_date_format(end_date_str)
    
    # Filter payments by date if provided
    payments = PaymentMaster.objects.all()
    if start_date and end_date:
        payments = payments.filter(payment_date__range=[start_date, end_date])
    elif start_date:
        payments = payments.filter(payment_date__gte=start_date)
    elif end_date:
        payments = payments.filter(payment_date__lte=end_date)
    payments = payments[:100]
    
    response = HttpResponse(content_type='text/html')
    response['Content-Disposition'] = 'inline; filename="payments_report.html"'
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Payments Report</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background-color: #f2f2f2; font-weight: bold; }}
            h1 {{ color: #333; text-align: center; }}
            @media print {{ body {{ margin: 0; }} }}
        </style>
    </head>
    <body>
        <h1>Payments Report</h1>
        <table>
            <thead>
                <tr>
                    <th>Date</th>
                    <th>Amount</th>
                    <th>Mode</th>
                    <th>Reference</th>
                </tr>
            </thead>
            <tbody>
    """
    
    for payment in payments:
        html_content += f"""
                <tr>
                    <td>{payment.payment_date}</td>
                    <td>₹{payment.payment_amount:.2f}</td>
                    <td>{payment.payment_mode}</td>
                    <td>{payment.payment_ref_no or 'N/A'}</td>
                </tr>
        """
    
    html_content += """
            </tbody>
        </table>
        <script>window.print();</script>
    </body>
    </html>
    """
    
    response.write(html_content)
    return response

@login_required
def export_payments_excel(request):
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="payments_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Date', 'Amount', 'Mode', 'Reference'])
    
    payments = PaymentMaster.objects.all()[:100]
    for payment in payments:
        writer.writerow([
            payment.payment_date,
            payment.payment_amount,
            payment.payment_mode,
            payment.payment_ref_no or 'N/A'
        ])
    
    return response

# Finance - Receipts
@login_required
def receipt_list(request):
    receipts = ReceiptMaster.objects.all().order_by('-receipt_date')
    
    search_query = request.GET.get('search', '')
    if search_query:
        receipts = receipts.filter(
            Q(receipt_ref_no__icontains=search_query) |
            Q(receipt_mode__icontains=search_query)
        )
    
    paginator = Paginator(receipts, 10)
    page_number = request.GET.get('page')
    receipts = paginator.get_page(page_number)
    
    context = {
        'receipts': receipts,
        'search_query': search_query,
        'title': 'Receipt List'
    }
    return render(request, 'finance/receipt_list.html', context)

@login_required
def add_receipt(request):
    if request.method == 'POST':
        form = ReceiptForm(request.POST)
        if form.is_valid():
            receipt = form.save()
            messages.success(request, f"Receipt of ₹{receipt.receipt_amount} added successfully!")
            return redirect('receipt_list')
    else:
        form = ReceiptForm()
    
    context = {
        'form': form,
        'title': 'Add Receipt'
    }
    return render(request, 'finance/receipt_form.html', context)

@login_required
def edit_receipt(request, pk):
    receipt = get_object_or_404(ReceiptMaster, receipt_id=pk)
    
    if request.method == 'POST':
        form = ReceiptForm(request.POST, instance=receipt)
        if form.is_valid():
            form.save()
            messages.success(request, "Receipt updated successfully!")
            return redirect('receipt_list')
    else:
        form = ReceiptForm(instance=receipt)
    
    context = {
        'form': form,
        'receipt': receipt,
        'title': 'Edit Receipt',
        'is_edit': True
    }
    return render(request, 'finance/receipt_form.html', context)

@login_required
def delete_receipt(request, pk):
    receipt = get_object_or_404(ReceiptMaster, receipt_id=pk)
    
    if request.method == 'POST':
        try:
            receipt.delete()
            messages.success(request, "Receipt deleted successfully!")
        except Exception as e:
            messages.error(request, f"Cannot delete receipt. Error: {str(e)}")
        return redirect('receipt_list')
    
    context = {
        'receipt': receipt,
        'title': 'Delete Receipt'
    }
    return render(request, 'finance/receipt_confirm_delete.html', context)

@login_required
def export_receipts_pdf(request):
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/html')
    response['Content-Disposition'] = 'inline; filename="receipts_report.html"'
    
    receipts = ReceiptMaster.objects.all()[:100]
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Receipts Report</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background-color: #f2f2f2; font-weight: bold; }}
            h1 {{ color: #333; text-align: center; }}
            @media print {{ body {{ margin: 0; }} }}
        </style>
    </head>
    <body>
        <h1>Receipts Report</h1>
        <table>
            <thead>
                <tr>
                    <th>Date</th>
                    <th>Amount</th>
                    <th>Mode</th>
                    <th>Reference</th>
                </tr>
            </thead>
            <tbody>
    """
    
    for receipt in receipts:
        html_content += f"""
                <tr>
                    <td>{receipt.receipt_date}</td>
                    <td>₹{receipt.receipt_amount:.2f}</td>
                    <td>{receipt.receipt_mode}</td>
                    <td>{receipt.receipt_ref_no or 'N/A'}</td>
                </tr>
        """
    
    html_content += """
            </tbody>
        </table>
        <script>window.print();</script>
    </body>
    </html>
    """
    
    response.write(html_content)
    return response

@login_required
def export_receipts_excel(request):
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="receipts_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Date', 'Amount', 'Mode', 'Reference'])
    
    receipts = ReceiptMaster.objects.all()[:100]
    for receipt in receipts:
        writer.writerow([
            receipt.receipt_date,
            receipt.receipt_amount,
            receipt.receipt_mode,
            receipt.receipt_ref_no or 'N/A'
        ])
    
    return response

# Sale Rate Management
@login_required
def sale_rate_list(request):
    rates = SaleRateMaster.objects.select_related('productid').all().order_by('productid__product_name')
    
    search_query = request.GET.get('search', '')
    if search_query:
        rates = rates.filter(
            Q(productid__product_name__icontains=search_query) |
            Q(product_batch_no__icontains=search_query)
        )
    
    paginator = Paginator(rates, 20)
    page_number = request.GET.get('page')
    rates = paginator.get_page(page_number)
    
    context = {
        'rates': rates,
        'search_query': search_query,
        'title': 'Sale Rate List'
    }
    return render(request, 'rates/sale_rate_list.html', context)

@login_required
def add_sale_rate(request):
    if request.method == 'POST':
        form = SaleRateForm(request.POST)
        if form.is_valid():
            rate = form.save()
            messages.success(request, f"Sale rate for {rate.productid.product_name} added successfully!")
            return redirect('sale_rate_list')
    else:
        form = SaleRateForm()
    
    context = {
        'form': form,
        'title': 'Add Sale Rate'
    }
    return render(request, 'rates/sale_rate_form.html', context)

@login_required
def update_sale_rate(request, pk):
    rate = get_object_or_404(SaleRateMaster, id=pk)
    
    if request.method == 'POST':
        form = SaleRateForm(request.POST, instance=rate)
        if form.is_valid():
            form.save()
            messages.success(request, f"Sale rate for {rate.productid.product_name} updated successfully!")
            return redirect('sale_rate_list')
    else:
        form = SaleRateForm(instance=rate)
    
    context = {
        'form': form,
        'rate': rate,
        'title': 'Update Sale Rate',
        'is_edit': True
    }
    return render(request, 'rates/sale_rate_form.html', context)

@login_required
def delete_sale_rate(request, pk):
    rate = get_object_or_404(SaleRateMaster, id=pk)
    
    if request.method == 'POST':
        product_name = rate.productid.product_name
        try:
            rate.delete()
            messages.success(request, f"Sale rate for {product_name} deleted successfully!")
        except Exception as e:
            messages.error(request, f"Cannot delete sale rate. Error: {str(e)}")
        return redirect('sale_rate_list')
    
    context = {
        'rate': rate,
        'title': 'Delete Sale Rate'
    }
    return render(request, 'rates/sale_rate_confirm_delete.html', context)

@login_required
def get_sales_analytics_api(request):
    from datetime import datetime, timedelta
    from django.db.models import Sum, Count
    
    # Get date range (default to last 30 days)
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=30)
    
    # Get sales data
    sales_data = SalesMaster.objects.filter(
        sales_invoice_no__sales_invoice_date__range=[start_date, end_date]
    ).aggregate(
        total_sales=Sum('sale_total_amount'),
        total_quantity=Sum('sale_quantity'),
        total_invoices=Count('sales_invoice_no', distinct=True)
    )
    
    return JsonResponse({
        'success': True,
        'data': {
            'total_sales': float(sales_data['total_sales'] or 0),
            'total_quantity': float(sales_data['total_quantity'] or 0),
            'total_invoices': sales_data['total_invoices'] or 0,
            'period': f"{start_date} to {end_date}"
        }
    })

@login_required
def export_inventory_pdf(request):
    from django.http import HttpResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    import io
    from datetime import datetime
    from django.db.models import Sum, Avg, Count

    try:
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Create styles
        styles = getSampleStyleSheet()
        story = []

        # Get inventory data
        products = ProductMaster.objects.all().order_by('product_name')[:100]  # Limit for performance
        inventory_data = []
        
        for product in products:
            try:
                # Get current stock
                stock_info = get_stock_status(product.productid)
                
                # Get latest batch info
                latest_batch = PurchaseMaster.objects.filter(
                    productid=product.productid
                ).order_by('-purchase_entry_date').first()
                
                # Calculate stock value
                current_stock = stock_info.get('current_stock', 0)
                avg_mrp = stock_info.get('avg_mrp', 0)
                stock_value = current_stock * avg_mrp
                
                inventory_data.append({
                    'product_name': product.product_name or 'N/A',
                    'company': product.product_company or 'N/A',
                    'category': product.product_category or 'N/A',
                    'batch_no': latest_batch.product_batch_no if latest_batch else 'N/A',
                    'stock': current_stock,
                    'mrp': avg_mrp,
                    'value': stock_value
                })
            except Exception as e:
                print(f"Error processing product {product.product_name}: {e}")
                continue

        # Title
        title_style = styles['Heading1']
        title_style.alignment = 1  # Center
        title = Paragraph("Inventory Report", title_style)
        story.append(title)
        
        # Date
        date_style = styles['Normal']
        date_style.alignment = 1
        date_text = Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y at %H:%M')}", date_style)
        story.append(date_text)
        story.append(Spacer(1, 0.3*inch))

        # Summary Statistics
        total_products = len(inventory_data)
        total_value = sum(item['value'] for item in inventory_data)
        out_of_stock = sum(1 for item in inventory_data if item['stock'] <= 0)
        low_stock = sum(1 for item in inventory_data if 0 < item['stock'] <= 10)
        
        summary_data = [
            ['Total Products', 'Total Inventory Value', 'Out of Stock', 'Low Stock'],
            [
                str(total_products), 
                f"₹{total_value:,.2f}", 
                str(out_of_stock), 
                str(low_stock)
            ]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch, 1.5*inch, 1.5*inch])
        summary_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))

        # Inventory Table
        story.append(Paragraph(f"Inventory Items ({total_products} products)", styles['Heading2']))
        
        table_data = [['Product Name', 'Company', 'Category', 'Batch No', 'Stock', 'MRP', 'Value']]
        
        for item in inventory_data:
            table_data.append([
                item['product_name'][:20] + '...' if len(item['product_name']) > 20 else item['product_name'],
                item['company'][:15] + '...' if len(item['company']) > 15 else item['company'],
                item['category'][:10] + '...' if len(item['category']) > 10 else item['category'],
                item['batch_no'][:8] + '...' if len(item['batch_no']) > 8 else item['batch_no'],
                str(int(item['stock'])),
                f"₹{item['mrp']:.0f}",
                f"₹{item['value']:.0f}"
            ])

        # Create table with proper styling
        inventory_table = Table(table_data, repeatRows=1)
        inventory_table.setStyle(TableStyle([
            # Header style
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Data rows style
            ('FONT', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (4, 1), (6, -1), 'RIGHT'),  # Align numeric columns right
            
            # Grid and alternating colors
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        story.append(inventory_table)

        # Build PDF
        doc.build(story)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="inventory_report.pdf"'
        return response
        
    except Exception as e:
        # Return error response
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)



@login_required
def export_inventory_excel(request):
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="inventory_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Product Name', 'Company', 'Batch No', 'MRP', 'Value'])
    
    products = ProductMaster.objects.all()[:100]
    for product in products:
        stock_info = get_stock_status(product.productid)
        batch = PurchaseMaster.objects.filter(productid=product.productid).first()
        writer.writerow([
            product.product_name,
            product.product_company,
            batch.product_batch_no if batch else 'N/A',
            stock_info.get('avg_mrp', 0),
            stock_info['current_stock'] * stock_info.get('avg_mrp', 0)
        ])
    
    return response

@login_required
def export_products_pdf(request):
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/html')
    response['Content-Disposition'] = 'inline; filename="products_report.html"'
    
    products = ProductMaster.objects.all()[:100]
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Products Report</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background-color: #f2f2f2; font-weight: bold; }}
            h1 {{ color: #333; text-align: center; }}
            @media print {{ body {{ margin: 0; }} }}
        </style>
    </head>
    <body>
        <h1>Products Report</h1>
        <table>
            <thead>
                <tr>
                    <th>ID</th>
                    <th>Product Name</th>
                    <th>Company</th>
                    <th>Batch No</th>
                    <th>Category</th>
                </tr>
            </thead>
            <tbody>
    """
    
    for product in products:
        batch = PurchaseMaster.objects.filter(productid=product.productid).first()
        html_content += f"""
                <tr>
                    <td>{product.productid}</td>
                    <td>{product.product_name}</td>
                    <td>{product.product_company}</td>
                    <td>{batch.product_batch_no if batch else 'N/A'}</td>
                    <td>{product.product_category or 'N/A'}</td>
                </tr>
        """
    
    html_content += """
            </tbody>
        </table>
        <script>window.print();</script>
    </body>
    </html>
    """
    
    response.write(html_content)
    return response

@login_required
def export_products_excel(request):
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Product Name', 'Company', 'Batch No', 'Category'])
    
    products = ProductMaster.objects.all()[:100]
    for product in products:
        batch = PurchaseMaster.objects.filter(productid=product.productid).first()
        writer.writerow([
            product.productid,
            product.product_name,
            product.product_company,
            batch.product_batch_no if batch else 'N/A',
            product.product_category or 'N/A'
        ])
    
    return response

@login_required
def export_sales_pdf(request):
    from django.http import HttpResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from datetime import datetime, date
    from .sales_analytics import SalesAnalytics
    import io
    import os

    # Get date range
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    
    # Parse dates - default to current month if no dates provided
    today = date.today()
    try:
        if start_date_str:
            if len(start_date_str) == 4:  # DDMM format
                day = int(start_date_str[:2])
                month = int(start_date_str[2:4])
                year = today.year
                start_date = date(year, month, day)
            else:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = today.replace(day=1)
            
        if end_date_str:
            if len(end_date_str) == 4:  # DDMM format
                day = int(end_date_str[:2])
                month = int(end_date_str[2:4])
                year = today.year
                end_date = date(year, month, day)
            else:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = today
    except (ValueError, TypeError):
        start_date = today.replace(day=1)
        end_date = today

    # Use SalesAnalytics for accurate calculations
    analytics = SalesAnalytics(start_date, end_date)
    report_data = analytics.get_comprehensive_report()
    
    # Extract data from analytics
    core_metrics = report_data['core_metrics']
    invoice_analysis = report_data['invoice_analysis']
    realtime_stats = report_data['realtime_stats']
    
    total_sales = core_metrics['total_sales']
    total_received = core_metrics['total_received']
    total_pending = core_metrics['total_pending']
    total_invoices = core_metrics['total_invoices']
    collection_rate = core_metrics['collection_rate']
    
    paid_invoices = invoice_analysis['paid_invoices']
    partial_paid = invoice_analysis['partial_paid']
    unpaid_invoices = invoice_analysis['unpaid_invoices']
    
    unique_customers = realtime_stats['unique_customers']
    total_products_sold = realtime_stats['total_products_sold']
    
    # Get sales invoices and top products
    sales_invoices = report_data['invoices'][:20]  # Limit for PDF
    product_sales = report_data['product_analytics'][:10]
    customer_sales = report_data['customer_analytics'][:10]

    # Create PDF buffer
    buffer = io.BytesIO()
    
    # Register custom fonts with pdfmetrics
    try:
        # Try to register Arial font if available
        pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
        pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))
        title_font = 'Arial-Bold'
        body_font = 'Arial'
    except:
        # Fallback to built-in fonts
        title_font = 'Helvetica-Bold'
        body_font = 'Helvetica'
    
    # Create document with custom styling
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.7*inch,
        bottomMargin=0.5*inch
    )
    
    # Create custom styles
    styles = getSampleStyleSheet()
    
    # Enhanced title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=30,
        alignment=1,  # Center
        textColor=colors.HexColor('#1e40af'),
        fontName=title_font
    )
    
    # Enhanced subtitle style
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=20,
        alignment=1,  # Center
        textColor=colors.HexColor('#64748b'),
        fontName=body_font
    )
    
    # Section header style
    section_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        spaceBefore=20,
        textColor=colors.HexColor('#374151'),
        fontName=title_font
    )
    
    # Build PDF content
    story = []
    
    # Title and header
    story.append(Paragraph("📊 Enhanced Sales Analytics Report", title_style))
    story.append(Paragraph(
        f"Period: {start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}<br/>"
        f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}", 
        subtitle_style
    ))
    
    # Summary statistics table
    summary_data = [
        ['Metric', 'Value'],
        ['Total Sales', f"₹{total_sales:,.0f}"],
        ['Amount Received', f"₹{total_received:,.0f}"],
        ['Amount Pending', f"₹{total_pending:,.0f}"],
        ['Total Invoices', str(total_invoices)],
        ['Customers Served', str(unique_customers)],
        ['Products Sold', f"{total_products_sold:,.0f}"],
        ['Collection Rate', f"{collection_rate:.1f}%"]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#374151')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), title_font),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTNAME', (0, 1), (-1, -1), body_font),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Invoice status breakdown
    status_data = [
        ['Invoice Status', 'Count'],
        ['Paid Invoices', str(paid_invoices)],
        ['Partial Paid', str(partial_paid)],
        ['Unpaid Invoices', str(unpaid_invoices)]
    ]
    
    status_table = Table(status_data, colWidths=[3*inch, 2*inch])
    status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dcfce7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#166534')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), title_font),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTNAME', (0, 1), (-1, -1), body_font),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0'))
    ]))
    
    story.append(status_table)
    story.append(Spacer(1, 20))
    
    # Sales invoices section
    if sales_invoices:
        story.append(Paragraph(f"Sales Invoices (Top {len(sales_invoices)})", section_style))
        
        invoice_data = [['Invoice No', 'Date', 'Customer', 'Total', 'Paid', 'Balance', 'Status']]
        
        for invoice in sales_invoices:
            invoice_total = invoice['sales_invoice_total']
            balance = invoice_total - invoice['sales_invoice_paid']
            
            if invoice['sales_invoice_paid'] >= invoice_total:
                status = 'Paid'
            elif invoice['sales_invoice_paid'] > 0:
                status = 'Partial'
            else:
                status = 'Unpaid'
            
            invoice_data.append([
                str(invoice['sales_invoice_no']),
                invoice['sales_invoice_date'].strftime('%d-%m-%Y'),
                (invoice['customer_name'] or 'N/A')[:15],
                f"₹{invoice_total:,.0f}",
                f"₹{invoice['sales_invoice_paid']:,.0f}",
                f"₹{balance:,.0f}",
                status
            ])
        
        invoice_table = Table(invoice_data, repeatRows=1)
        invoice_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), title_font),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTNAME', (0, 1), (-1, -1), body_font),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
        ]))
        
        story.append(invoice_table)
        story.append(Spacer(1, 20))
    
    # Top products section
    if product_sales:
        story.append(Paragraph("Top Products by Sales", section_style))
        
        product_data = [['Product', 'Company', 'Quantity', 'Amount', 'Avg Rate']]
        
        for product in product_sales:
            product_data.append([
                (product['productid__product_name'] or 'N/A')[:20],
                (product['productid__product_company'] or 'N/A')[:15],
                f"{product['total_quantity']:,.0f}",
                f"₹{product['total_amount']:,.0f}",
                f"₹{(product['avg_rate'] or 0):,.0f}"
            ])
        
        product_table = Table(product_data, repeatRows=1)
        product_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), title_font),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTNAME', (0, 1), (-1, -1), body_font),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdf4')])
        ]))
        
        story.append(product_table)
    
    # Build PDF
    doc.build(story)
    
    # Return PDF response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="enhanced_sales_analytics.pdf"'
    return response

# @login_required
# def export_sales_pdf(request):
#     from django.http import HttpResponse
#     from datetime import datetime
#     from .sales_analytics import SalesAnalytics
    
#     # Get date range
#     start_date_str = request.GET.get('start_date', '')
#     end_date_str = request.GET.get('end_date', '')
    
#     today = datetime.now().date()
#     try:
#         start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else today.replace(day=1)
#         end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
#     except ValueError:
#         start_date = today.replace(day=1)
#         end_date = today
    
#     # Get analytics data
#     analytics = SalesAnalytics(start_date, end_date)
#     report_data = analytics.get_comprehensive_report()
    
#     response = HttpResponse(content_type='text/html')
#     response['Content-Disposition'] = 'inline; filename="enhanced_sales_analytics.html"'
    
#     html_content = f"""
#     <!DOCTYPE html>
#     <html>
#     <head>
#         <title>Enhanced Sales Analytics Report</title>
#         <style>
#             body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 20px; background: #f8fafc; }}
#             .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
#             .header {{ text-align: center; margin-bottom: 40px; padding: 20px; background: linear-gradient(135deg, #2563eb, #3b82f6); color: white; border-radius: 8px; }}
#             .kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 30px 0; }}
#             .kpi-card {{ background: #f8fafc; padding: 20px; border-radius: 8px; text-align: center; border-left: 4px solid #2563eb; }}
#             .kpi-value {{ font-size: 2rem; font-weight: bold; color: #1e293b; margin: 10px 0; }}
#             .kpi-label {{ color: #64748b; font-size: 0.9rem; text-transform: uppercase; }}
#             table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
#             th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #e2e8f0; }}
#             th {{ background: #f1f5f9; font-weight: 600; color: #374151; }}
#             .section-title {{ font-size: 1.5rem; font-weight: 600; color: #1e293b; margin: 30px 0 15px 0; border-bottom: 2px solid #e2e8f0; padding-bottom: 10px; }}
#             .status-paid {{ background: #dcfce7; color: #166534; padding: 4px 8px; border-radius: 4px; font-size: 0.8rem; }}
#             .status-pending {{ background: #fef3c7; color: #92400e; padding: 4px 8px; border-radius: 4px; font-size: 0.8rem; }}
#             @media print {{ body {{ background: white; }} .container {{ box-shadow: none; }} }}
#         </style>
#     </head>
#     <body>
#         <div class="container">
#             <div class="header">
#                 <h1>Enhanced Sales Analytics Report</h1>
#                 <p>Period: {start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}</p>
#                 <p>Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}</p>
#             </div>
            
#             <div class="kpi-grid">
#                 <div class="kpi-card">
#                     <div class="kpi-value">₹{report_data['core_metrics']['total_sales']:,.0f}</div>
#                     <div class="kpi-label">Total Sales</div>
#                 </div>
#                 <div class="kpi-card">
#                     <div class="kpi-value">₹{report_data['core_metrics']['total_received']:,.0f}</div>
#                     <div class="kpi-label">Amount Received</div>
#                 </div>
#                 <div class="kpi-card">
#                     <div class="kpi-value">₹{report_data['core_metrics']['total_pending']:,.0f}</div>
#                     <div class="kpi-label">Amount Pending</div>
#                 </div>
#                 <div class="kpi-card">
#                     <div class="kpi-value">{report_data['core_metrics']['total_invoices']}</div>
#                     <div class="kpi-label">Total Invoices</div>
#                 </div>
#                 <div class="kpi-card">
#                     <div class="kpi-value">{report_data['realtime_stats']['unique_customers']}</div>
#                     <div class="kpi-label">Customers Served</div>
#                 </div>
#                 <div class="kpi-card">
#                     <div class="kpi-value">{report_data['realtime_stats']['total_products_sold']:,.0f}</div>
#                     <div class="kpi-label">Products Sold</div>
#                 </div>
#             </div>
            
#             <h2 class="section-title">Sales Invoices ({len(analytics.invoices)} invoices)</h2>
#             <table>
#                 <thead>
#                     <tr>
#                         <th>Invoice No</th>
#                         <th>Date</th>
#                         <th>Customer</th>
#                         <th>Type</th>
#                         <th>Total</th>
#                         <th>Paid</th>
#                         <th>Balance</th>
#                         <th>Status</th>
#                     </tr>
#                 </thead>
#                 <tbody>
#     """
    
#     for invoice in analytics.invoices:
#         balance = invoice.sales_invoice_total - invoice.sales_invoice_paid
#         if balance <= 0:
#             status = '<span class="status-paid">Paid</span>'
#         else:
#             status = '<span class="status-pending">Pending</span>'
        
#         html_content += f"""
#                     <tr>
#                         <td>{invoice.sales_invoice_no}</td>
#                         <td>{invoice.sales_invoice_date.strftime('%d-%m-%Y')}</td>
#                         <td>{invoice.customerid.customer_name}</td>
#                         <td>{invoice.customerid.customer_type}</td>
#                         <td>₹{invoice.sales_invoice_total:,.2f}</td>
#                         <td>₹{invoice.sales_invoice_paid:,.2f}</td>
#                         <td>₹{balance:,.2f}</td>
#                         <td>{status}</td>
#                     </tr>
#         """
    
#     html_content += f"""
#                 </tbody>
#             </table>
            
#             <h2 class="section-title">Top Products</h2>
#             <table>
#                 <thead>
#                     <tr>
#                         <th>Product</th>
#                         <th>Company</th>
#                         <th>Quantity</th>
#                         <th>Amount</th>
#                         <th>Avg Rate</th>
#                     </tr>
#                 </thead>
#                 <tbody>
#     """
    
#     for product in report_data['top_performers']['top_products'][:10]:
#         html_content += f"""
#                     <tr>
#                         <td>{product['productid__product_name']}</td>
#                         <td>{product['productid__product_company']}</td>
#                         <td>{product['total_quantity']:,.0f}</td>
#                         <td>₹{product['total_amount']:,.2f}</td>
#                         <td>₹{product['avg_rate']:,.2f}</td>
#                     </tr>
#         """
    
#     html_content += """
#                 </tbody>
#             </table>
#         </div>
#         <script>window.print();</script>
#     </body>
#     </html>
#     """
    
#     response.write(html_content)
#     return response

# @login_required
# def export_sales_excel(request):
#     import csv
#     from django.http import HttpResponse
#     from datetime import datetime
#     from .sales_analytics import SalesAnalytics
    
#     # Get date range
#     start_date_str = request.GET.get('start_date', '')
#     end_date_str = request.GET.get('end_date', '')
    
#     today = datetime.now().date()
#     try:
#         start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else today.replace(day=1)
#         end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else today
#     except ValueError:
#         start_date = today.replace(day=1)
#         end_date = today
    
#     # Get analytics data
#     analytics = SalesAnalytics(start_date, end_date)
#     report_data = analytics.get_comprehensive_report()
    
#     response = HttpResponse(content_type='text/csv')
#     response['Content-Disposition'] = 'attachment; filename="enhanced_sales_analytics.csv"'
    
#     writer = csv.writer(response)
    
#     # Header
#     writer.writerow(['Enhanced Sales Analytics Report'])
#     writer.writerow([f'Period: {start_date.strftime("%d %B %Y")} to {end_date.strftime("%d %B %Y")}'])
#     writer.writerow([f'Generated: {datetime.now().strftime("%d %B %Y at %H:%M")}'])
#     writer.writerow([])
    
#     # KPI Summary
#     writer.writerow(['KEY PERFORMANCE INDICATORS'])
#     writer.writerow(['Metric', 'Value'])
#     writer.writerow(['Total Sales', f"₹{report_data['core_metrics']['total_sales']:,.2f}"])
#     writer.writerow(['Amount Received', f"₹{report_data['core_metrics']['total_received']:,.2f}"])
#     writer.writerow(['Amount Pending', f"₹{report_data['core_metrics']['total_pending']:,.2f}"])
#     writer.writerow(['Total Invoices', report_data['core_metrics']['total_invoices']])
#     writer.writerow(['Collection Rate', f"{report_data['core_metrics']['collection_rate']:.1f}%"])
#     writer.writerow(['Unique Customers', report_data['realtime_stats']['unique_customers']])
#     writer.writerow(['Products Sold', f"{report_data['realtime_stats']['total_products_sold']:,.0f}"])
#     writer.writerow(['Total Discount Given', f"₹{report_data['realtime_stats']['total_discount_given']:,.2f}"])
#     writer.writerow([])
    
#     # Sales Invoices
#     writer.writerow(['SALES INVOICES'])
#     writer.writerow(['Invoice No', 'Date', 'Customer', 'Type', 'Total', 'Paid', 'Balance', 'Status'])
    
#     for invoice in analytics.invoices:
#         balance = invoice.sales_invoice_total - invoice.sales_invoice_paid
#         status = "Paid" if balance <= 0 else "Pending"
        
#         writer.writerow([
#             invoice.sales_invoice_no,
#             invoice.sales_invoice_date.strftime('%d-%m-%Y'),
#             invoice.customerid.customer_name,
#             invoice.customerid.customer_type,
#             f"₹{invoice.sales_invoice_total:,.2f}",
#             f"₹{invoice.sales_invoice_paid:,.2f}",
#             f"₹{balance:,.2f}",
#             status
#         ])
    
#     writer.writerow([])
    
#     # Top Products
#     writer.writerow(['TOP PRODUCTS'])
#     writer.writerow(['Product', 'Company', 'Quantity', 'Amount', 'Avg Rate', 'Invoices'])
    
#     for product in report_data['top_performers']['top_products'][:15]:
#         writer.writerow([
#             product['productid__product_name'],
#             product['productid__product_company'],
#             f"{product['total_quantity']:,.0f}",
#             f"₹{product['total_amount']:,.2f}",
#             f"₹{product['avg_rate']:,.2f}",
#             product['invoice_count']
#         ])
    
#     writer.writerow([])
    
#     # Top Customers
#     writer.writerow(['TOP CUSTOMERS'])
#     writer.writerow(['Customer', 'Type', 'Amount', 'Invoices', 'Avg Invoice', 'Last Purchase'])
    
#     for customer in report_data['top_performers']['top_customers'][:15]:
#         writer.writerow([
#             customer['sales_invoice_no__customerid__customer_name'],
#             customer['sales_invoice_no__customerid__customer_type'],
#             f"₹{customer['total_amount']:,.2f}",
#             customer['invoice_count'],
#             f"₹{customer['avg_invoice_value']:,.2f}",
#             customer['last_purchase_date'].strftime('%d-%m-%Y') if customer['last_purchase_date'] else 'N/A'
#         ])
    
#     return response


# def export_sales_excel(request):
#     from django.http import HttpResponse
#     from datetime import datetime, date
#     from django.db.models import Sum, Count, Avg, Max  # Added Max import here
#     from .models import SalesInvoiceMaster, SalesMaster, CustomerMaster
#     from openpyxl import Workbook
#     from openpyxl.styles import Font, Alignment, PatternFill
#     from openpyxl.utils import get_column_letter
#     import calendar

#     # Get date range
#     start_date_str = request.GET.get('start_date', '')
#     end_date_str = request.GET.get('end_date', '')
    
#     # Parse dates
#     today = date.today()
#     try:
#         if start_date_str:
#             if len(start_date_str) == 4:  # DDMM format
#                 day = int(start_date_str[:2])
#                 month = int(start_date_str[2:4])
#                 year = today.year
#                 start_date = date(year, month, day)
#             else:
#                 start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
#         else:
#             start_date = today.replace(day=1)
            
#         if end_date_str:
#             if len(end_date_str) == 4:  # DDMM format
#                 day = int(end_date_str[:2])
#                 month = int(end_date_str[2:4])
#                 year = today.year
#                 end_date = date(year, month, day)
#             else:
#                 end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
#         else:
#             end_date = today
#     except (ValueError, TypeError):
#         today = date.today()
#         start_date = today.replace(day=1)
#         end_date = today

#     # Get data directly from database
#     sales_invoices = SalesInvoiceMaster.objects.filter(
#         sales_invoice_date__range=[start_date, end_date]
#     ).select_related('customerid')
    
#     # Calculate metrics
#     total_sales = 0
#     total_received = 0
#     for invoice in sales_invoices:
#         invoice_total = SalesMaster.objects.filter(
#             sales_invoice_no=invoice
#         ).aggregate(total=Sum('sale_total_amount'))['total'] or 0
#         total_sales += invoice_total
#         total_received += invoice.sales_invoice_paid
    
#     total_pending = total_sales - total_received
#     total_invoices = sales_invoices.count()
    
#     # Get product sales data
#     product_sales = SalesMaster.objects.filter(
#         sales_invoice_no__in=sales_invoices
#     ).values(
#         'productid__product_name', 
#         'productid__product_company'
#     ).annotate(
#         total_quantity=Sum('sale_quantity'),
#         total_amount=Sum('sale_total_amount'),
#         avg_rate=Avg('sale_rate'),
#         invoice_count=Count('sales_invoice_no', distinct=True)
#     ).order_by('-total_amount')
    
#     # Get customer sales data - NOW USING Max WHICH IS IMPORTED
#     customer_sales = SalesInvoiceMaster.objects.filter(
#         sales_invoice_date__range=[start_date, end_date]
#     ).values(
#         'customerid__customer_name',
#         'customerid__customer_type'
#     ).annotate(
#         total_amount=Sum('salesmaster__sale_total_amount'),
#         invoice_count=Count('sales_invoice_no'),
#         avg_invoice_value=Avg('salesmaster__sale_total_amount'),
#         last_purchase_date=Max('sales_invoice_date')  # This now works
#     ).order_by('-total_amount')

#     # Get unique customers count
#     unique_customers = sales_invoices.values('customerid').distinct().count()
    
#     # Get total products sold
#     total_products_sold = SalesMaster.objects.filter(
#         sales_invoice_no__in=sales_invoices
#     ).aggregate(total=Sum('sale_quantity'))['total'] or 0

#     # Create Excel workbook
#     wb = Workbook()
    
#     # Remove default sheet
#     wb.remove(wb.active)
    
#     # Define styles
#     header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
#     normal_font = Font(name='Arial', size=10)
#     title_font = Font(name='Arial', size=14, bold=True)
#     header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
#     currency_format = '"₹"#,##0.00'
    
#     # Summary Sheet
#     ws_summary = wb.create_sheet("Summary")
    
#     # Title
#     ws_summary.merge_cells('A1:H1')
#     ws_summary['A1'] = "Enhanced Sales Analytics Report"
#     ws_summary['A1'].font = title_font
#     ws_summary['A1'].alignment = Alignment(horizontal='center')
    
#     ws_summary['A2'] = f"Period: {start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}"
#     ws_summary['A3'] = f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}"
    
#     # KPI Section
#     ws_summary['A5'] = "KEY PERFORMANCE INDICATORS"
#     ws_summary['A5'].font = Font(name='Arial', size=12, bold=True)
    
#     kpi_data = [
#         ['Metric', 'Value'],
#         ['Total Sales', total_sales],
#         ['Amount Received', total_received],
#         ['Amount Pending', total_pending],
#         ['Total Invoices', total_invoices],
#         ['Collection Rate', f"{(total_received/total_sales*100) if total_sales > 0 else 0:.1f}%"],
#         ['Unique Customers', unique_customers],
#         ['Products Sold', total_products_sold],
#     ]
    
#     for row_num, row_data in enumerate(kpi_data, 6):
#         for col_num, cell_value in enumerate(row_data, 1):
#             cell = ws_summary.cell(row=row_num, column=col_num)
#             cell.value = cell_value
#             if row_num == 6:  # Header row
#                 cell.font = header_font
#                 cell.fill = header_fill
#                 cell.alignment = Alignment(horizontal='center')
#             else:
#                 cell.font = normal_font
#             if col_num == 2 and row_num > 6 and row_num < 10:  # Currency values (rows 7-9)
#                 if row_num in [7, 8, 9]:  # Sales, Received, Pending
#                     cell.number_format = currency_format

#     # Sales Invoices Sheet
#     ws_invoices = wb.create_sheet("Sales Invoices")
    
#     invoice_headers = ['Invoice No', 'Date', 'Customer', 'Type', 'Total Amount', 'Paid Amount', 'Balance', 'Status']
#     for col_num, header in enumerate(invoice_headers, 1):
#         cell = ws_invoices.cell(row=1, column=col_num)
#         cell.value = header
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = Alignment(horizontal='center')
    
#     for row_num, invoice in enumerate(sales_invoices, 2):
#         invoice_total = SalesMaster.objects.filter(sales_invoice_no=invoice).aggregate(total=Sum('sale_total_amount'))['total'] or 0
#         balance = invoice_total - invoice.sales_invoice_paid
#         status = "Paid" if invoice.sales_invoice_paid >= invoice_total else "Partial" if invoice.sales_invoice_paid > 0 else "Unpaid"
        
#         invoice_data = [
#             invoice.sales_invoice_no,
#             invoice.sales_invoice_date,
#             invoice.customerid.customer_name,
#             invoice.customerid.customer_type,
#             invoice_total,
#             invoice.sales_invoice_paid,
#             balance,
#             status
#         ]
        
#         for col_num, cell_value in enumerate(invoice_data, 1):
#             cell = ws_invoices.cell(row=row_num, column=col_num)
#             cell.value = cell_value
#             cell.font = normal_font
#             if col_num in [5, 6, 7]:  # Currency columns
#                 cell.number_format = currency_format

#     # Top Products Sheet
#     ws_products = wb.create_sheet("Top Products")
    
#     product_headers = ['Product', 'Company', 'Quantity', 'Amount', 'Avg Rate', 'Invoices']
#     for col_num, header in enumerate(product_headers, 1):
#         cell = ws_products.cell(row=1, column=col_num)
#         cell.value = header
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = Alignment(horizontal='center')
    
#     for row_num, product in enumerate(product_sales[:20], 2):
#         product_data = [
#             product['productid__product_name'] or 'N/A',
#             product['productid__product_company'] or 'N/A',
#             product['total_quantity'],
#             product['total_amount'],
#             product['avg_rate'] or 0,
#             product['invoice_count']
#         ]
        
#         for col_num, cell_value in enumerate(product_data, 1):
#             cell = ws_products.cell(row=row_num, column=col_num)
#             cell.value = cell_value
#             cell.font = normal_font
#             if col_num in [4, 5]:  # Currency columns
#                 cell.number_format = currency_format

#     # Top Customers Sheet
#     ws_customers = wb.create_sheet("Top Customers")
    
#     customer_headers = ['Customer', 'Type', 'Amount', 'Invoices', 'Avg Invoice', 'Last Purchase']
#     for col_num, header in enumerate(customer_headers, 1):
#         cell = ws_customers.cell(row=1, column=col_num)
#         cell.value = header
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = Alignment(horizontal='center')
    
#     for row_num, customer in enumerate(customer_sales[:20], 2):
#         customer_data = [
#             customer['customerid__customer_name'] or 'N/A',
#             customer['customerid__customer_type'] or 'N/A',
#             customer['total_amount'] or 0,
#             customer['invoice_count'],
#             customer['avg_invoice_value'] or 0,
#             customer['last_purchase_date'].strftime('%d-%m-%Y') if customer['last_purchase_date'] else 'N/A'
#         ]
        
#         for col_num, cell_value in enumerate(customer_data, 1):
#             cell = ws_customers.cell(row=row_num, column=col_num)
#             cell.value = cell_value
#             cell.font = normal_font
#             if col_num in [3, 5]:  # Currency columns
#                 cell.number_format = currency_format

#     # Auto-adjust column widths for all sheets
#     for ws in wb.worksheets:
#         for column in ws.columns:
#             max_length = 0
#             column_letter = get_column_letter(column[0].column)
#             for cell in column:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass
#             adjusted_width = min((max_length + 2), 50)
#             ws.column_dimensions[column_letter].width = adjusted_width

#     # Create response
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'attachment; filename="sales_analytics_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
#     wb.save(response)
#     return response

@login_required
def export_sales_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Get data from database
    sales_data = Sales.objects.all()
    if start_date and end_date:
        sales_data = sales_data.filter(sale_date__range=[start_date, end_date])
    
    # Convert to DataFrame
    df = pd.DataFrame(list(sales_data.values()))
    
    # Create Excel response
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sales Report', index=False)
    
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'
    return response

@login_required
def export_purchases_pdf(request):
    from django.http import HttpResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from datetime import datetime, date
    from .purchase_analytics import PurchaseAnalytics
    import io
    
    # Get date range
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    
    # Parse dates - default to current month if no dates provided
    today = date.today()
    try:
        if start_date_str:
            if len(start_date_str) == 4:  # DDMM format
                day = int(start_date_str[:2])
                month = int(start_date_str[2:4])
                year = today.year
                start_date = date(year, month, day)
            else:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = today.replace(day=1)
            
        if end_date_str:
            if len(end_date_str) == 4:  # DDMM format
                day = int(end_date_str[:2])
                month = int(end_date_str[2:4])
                year = today.year
                end_date = date(year, month, day)
            else:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = today
    except (ValueError, TypeError):
        start_date = today.replace(day=1)
        end_date = today

    # Use PurchaseAnalytics for accurate calculations
    analytics = PurchaseAnalytics(start_date, end_date)
    report_data = analytics.get_comprehensive_report()
    
    # Extract data from analytics
    core_metrics = report_data['core_metrics']
    invoice_analysis = report_data['invoice_analysis']
    realtime_stats = report_data['realtime_stats']
    
    total_purchases = core_metrics['total_purchases']
    total_paid = core_metrics['total_paid']
    total_pending = core_metrics['total_pending']
    total_invoices = core_metrics['total_invoices']
    payment_rate = core_metrics['payment_rate']
    
    paid_invoices = invoice_analysis['paid_invoices']
    partial_paid = invoice_analysis['partial_paid']
    unpaid_invoices = invoice_analysis['unpaid_invoices']
    
    unique_suppliers = realtime_stats['unique_suppliers']
    total_products_purchased = realtime_stats['total_products_purchased']
    
    # Get purchase invoices and top data
    purchase_invoices = report_data['invoices'][:20]  # Limit for PDF
    product_purchases = report_data['product_analytics'][:10]
    supplier_purchases = report_data['supplier_analytics'][:10]

    # Create PDF buffer
    buffer = io.BytesIO()
    
    # Register custom fonts with pdfmetrics
    try:
        pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
        pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))
        title_font = 'Arial-Bold'
        body_font = 'Arial'
    except:
        title_font = 'Helvetica-Bold'
        body_font = 'Helvetica'
    
    # Create document with custom styling
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.7*inch,
        bottomMargin=0.5*inch
    )
    
    # Create custom styles
    styles = getSampleStyleSheet()
    
    # Enhanced title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=30,
        alignment=1,  # Center
        textColor=colors.HexColor('#059669'),
        fontName=title_font
    )
    
    # Enhanced subtitle style
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=20,
        alignment=1,  # Center
        textColor=colors.HexColor('#64748b'),
        fontName=body_font
    )
    
    # Section header style
    section_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        spaceBefore=20,
        textColor=colors.HexColor('#374151'),
        fontName=title_font
    )
    
    # Build PDF content
    story = []
    
    # Title and header
    story.append(Paragraph("🛒 Enhanced Purchase Analytics Report", title_style))
    story.append(Paragraph(
        f"Period: {start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}<br/>"
        f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}", 
        subtitle_style
    ))
    
    # Summary statistics table
    summary_data = [
        ['Metric', 'Value'],
        ['Total Purchases', f"₹{total_purchases:,.0f}"],
        ['Amount Paid', f"₹{total_paid:,.0f}"],
        ['Amount Pending', f"₹{total_pending:,.0f}"],
        ['Total Invoices', str(total_invoices)],
        ['Suppliers Engaged', str(unique_suppliers)],
        ['Products Purchased', f"{total_products_purchased:,.0f}"],
        ['Payment Rate', f"{payment_rate:.1f}%"]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0fdf4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#166534')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), title_font),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTNAME', (0, 1), (-1, -1), body_font),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Invoice status breakdown
    status_data = [
        ['Invoice Status', 'Count'],
        ['Paid Invoices', str(paid_invoices)],
        ['Partial Paid', str(partial_paid)],
        ['Unpaid Invoices', str(unpaid_invoices)]
    ]
    
    status_table = Table(status_data, colWidths=[3*inch, 2*inch])
    status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fef3c7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#92400e')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), title_font),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTNAME', (0, 1), (-1, -1), body_font),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0'))
    ]))
    
    story.append(status_table)
    story.append(Spacer(1, 20))
    
    # Purchase invoices section
    if purchase_invoices:
        story.append(Paragraph(f"Purchase Invoices (Top {len(purchase_invoices)})", section_style))
        
        invoice_data = [['Invoice No', 'Date', 'Supplier', 'Total', 'Paid', 'Balance', 'Status']]
        
        for invoice in purchase_invoices:
            invoice_total = invoice['invoice_total']
            balance = invoice_total - invoice['invoice_paid']
            
            if invoice['invoice_paid'] >= invoice_total:
                status = 'Paid'
            elif invoice['invoice_paid'] > 0:
                status = 'Partial'
            else:
                status = 'Unpaid'
            
            invoice_data.append([
                str(invoice['invoice_no']),
                invoice['invoice_date'].strftime('%d-%m-%Y'),
                (invoice['supplier_name'] or 'N/A')[:15],
                f"₹{invoice_total:,.0f}",
                f"₹{invoice['invoice_paid']:,.0f}",
                f"₹{balance:,.0f}",
                status
            ])
        
        invoice_table = Table(invoice_data, repeatRows=1)
        invoice_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), title_font),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTNAME', (0, 1), (-1, -1), body_font),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdf4')])
        ]))
        
        story.append(invoice_table)
        story.append(Spacer(1, 20))
    
    # Top products section
    if product_purchases:
        story.append(Paragraph("Top Products by Purchase Value", section_style))
        
        product_data = [['Product', 'Company', 'Quantity', 'Amount', 'Avg Rate']]
        
        for product in product_purchases:
            product_data.append([
                (product['productid__product_name'] or 'N/A')[:20],
                (product['productid__product_company'] or 'N/A')[:15],
                f"{product['total_quantity']:,.0f}",
                f"₹{product['total_amount']:,.0f}",
                f"₹{(product['avg_rate'] or 0):,.0f}"
            ])
        
        product_table = Table(product_data, repeatRows=1)
        product_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), title_font),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTNAME', (0, 1), (-1, -1), body_font),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#faf5ff')])
        ]))
        
        story.append(product_table)
        story.append(Spacer(1, 15))
    
    # Top suppliers section
    if supplier_purchases:
        story.append(Paragraph("Top Suppliers by Purchase Value", section_style))
        
        supplier_data = [['Supplier', 'Type', 'Amount', 'Invoices', 'Payment Rate']]
        
        for supplier in supplier_purchases:
            supplier_data.append([
                (supplier['supplierid__supplier_name'] or 'N/A')[:20],
                (supplier['supplierid__supplier_type'] or 'N/A')[:10],
                f"₹{supplier['total_amount']:,.0f}",
                str(supplier['invoice_count']),
                f"{(supplier['payment_rate'] or 0):,.1f}%"
            ])
        
        supplier_table = Table(supplier_data, repeatRows=1)
        supplier_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), title_font),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTNAME', (0, 1), (-1, -1), body_font),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fef2f2')])
        ]))
        
        story.append(supplier_table)
    
    # Build PDF
    doc.build(story)
    
    # Return PDF response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="enhanced_purchase_analytics.pdf"'
    return response

@login_required
def export_purchases_excel(request):
    from django.http import HttpResponse
    from datetime import datetime, date
    from .purchase_analytics import PurchaseAnalytics
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    # Get date range
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    
    # Parse dates
    today = date.today()
    try:
        if start_date_str:
            if len(start_date_str) == 4:  # DDMM format
                day = int(start_date_str[:2])
                month = int(start_date_str[2:4])
                year = today.year
                start_date = date(year, month, day)
            else:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = today.replace(day=1)
            
        if end_date_str:
            if len(end_date_str) == 4:  # DDMM format
                day = int(end_date_str[:2])
                month = int(end_date_str[2:4])
                year = today.year
                end_date = date(year, month, day)
            else:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = today
    except (ValueError, TypeError):
        today = date.today()
        start_date = today.replace(day=1)
        end_date = today

    # Get comprehensive analytics
    analytics = PurchaseAnalytics(start_date, end_date)
    report_data = analytics.get_comprehensive_report()
    
    # Extract data
    core_metrics = report_data['core_metrics']
    purchase_invoices = report_data['invoices']
    product_purchases = report_data['product_analytics']
    supplier_purchases = report_data['supplier_analytics']
    realtime_stats = report_data['realtime_stats']

    # Create Excel workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    normal_font = Font(name='Arial', size=10)
    title_font = Font(name='Arial', size=14, bold=True)
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    currency_format = '"₹"#,##0.00'
    
    # Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    
    # Title
    ws_summary.merge_cells('A1:H1')
    ws_summary['A1'] = "Enhanced Purchase Analytics Report"
    ws_summary['A1'].font = title_font
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    
    ws_summary['A2'] = f"Period: {start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}"
    ws_summary['A3'] = f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}"
    
    # KPI Section
    ws_summary['A5'] = "KEY PERFORMANCE INDICATORS"
    ws_summary['A5'].font = Font(name='Arial', size=12, bold=True)
    
    kpi_data = [
        ['Metric', 'Value'],
        ['Total Purchases', core_metrics['total_purchases']],
        ['Amount Paid', core_metrics['total_paid']],
        ['Amount Pending', core_metrics['total_pending']],
        ['Total Invoices', core_metrics['total_invoices']],
        ['Payment Rate', f"{core_metrics['payment_rate']:.1f}%"],
        ['Unique Suppliers', realtime_stats['unique_suppliers']],
        ['Products Purchased', realtime_stats['total_products_purchased']],
    ]
    
    for row_num, row_data in enumerate(kpi_data, 6):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_num, column=col_num)
            cell.value = cell_value
            if row_num == 6:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.font = normal_font
            if col_num == 2 and row_num > 6 and row_num < 10:  # Currency values (rows 7-9)
                if row_num in [7, 8, 9]:  # Purchases, Paid, Pending
                    cell.number_format = currency_format

    # Purchase Invoices Sheet
    ws_invoices = wb.create_sheet("Purchase Invoices")
    
    invoice_headers = ['Invoice No', 'Date', 'Supplier', 'Type', 'Total Amount', 'Paid Amount', 'Balance', 'Status']
    for col_num, header in enumerate(invoice_headers, 1):
        cell = ws_invoices.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for row_num, invoice in enumerate(purchase_invoices, 2):
        invoice_total = invoice['invoice_total']
        balance = invoice_total - invoice['invoice_paid']
        status = "Paid" if invoice['invoice_paid'] >= invoice_total else "Partial" if invoice['invoice_paid'] > 0 else "Unpaid"
        
        invoice_data = [
            invoice['invoice_no'],
            invoice['invoice_date'],
            invoice['supplier_name'],
            invoice['supplier_type'],
            invoice_total,
            invoice['invoice_paid'],
            balance,
            status
        ]
        
        for col_num, cell_value in enumerate(invoice_data, 1):
            cell = ws_invoices.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.font = normal_font
            if col_num in [5, 6, 7]:  # Currency columns
                cell.number_format = currency_format

    # Top Products Sheet
    ws_products = wb.create_sheet("Top Products")
    
    product_headers = ['Product', 'Company', 'Category', 'Quantity', 'Amount', 'Avg Rate', 'Invoices']
    for col_num, header in enumerate(product_headers, 1):
        cell = ws_products.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for row_num, product in enumerate(product_purchases[:20], 2):
        product_data = [
            product['productid__product_name'] or 'N/A',
            product['productid__product_company'] or 'N/A',
            product['productid__product_category'] or 'N/A',
            product['total_quantity'],
            product['total_amount'],
            product['avg_rate'] or 0,
            product['invoice_count']
        ]
        
        for col_num, cell_value in enumerate(product_data, 1):
            cell = ws_products.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.font = normal_font
            if col_num in [5, 6]:  # Currency columns
                cell.number_format = currency_format

    # Top Suppliers Sheet
    ws_suppliers = wb.create_sheet("Top Suppliers")
    
    supplier_headers = ['Supplier', 'Type', 'Amount', 'Invoices', 'Payment Rate', 'Last Purchase']
    for col_num, header in enumerate(supplier_headers, 1):
        cell = ws_suppliers.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for row_num, supplier in enumerate(supplier_purchases[:20], 2):
        supplier_data = [
            supplier['supplierid__supplier_name'] or 'N/A',
            supplier['supplierid__supplier_type'] or 'N/A',
            supplier['total_amount'] or 0,
            supplier['invoice_count'],
            f"{supplier['payment_rate']:.1f}%" if supplier['payment_rate'] else '0.0%',
            supplier['last_purchase_date'].strftime('%d-%m-%Y') if supplier['last_purchase_date'] else 'N/A'
        ]
        
        for col_num, cell_value in enumerate(supplier_data, 1):
            cell = ws_suppliers.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.font = normal_font
            if col_num == 3:  # Currency column
                cell.number_format = currency_format

    # Auto-adjust column widths for all sheets
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="purchase_analytics_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def export_financial_pdf(request):
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/html')
    response['Content-Disposition'] = 'inline; filename="financial_report.html"'
    
    # Get basic financial data
    total_receivables = 0
    for invoice in SalesInvoiceMaster.objects.all():
        balance = invoice.sales_invoice_total - invoice.sales_invoice_paid
        if balance > 0:
            total_receivables += balance
    
    total_payables = InvoiceMaster.objects.aggregate(
        total=Sum(F('invoice_total') - F('invoice_paid'))
    )['total'] or 0
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Financial Report</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background-color: #f2f2f2; font-weight: bold; }}
            h1 {{ color: #333; text-align: center; }}
            @media print {{ body {{ margin: 0; }} }}
        </style>
    </head>
    <body>
        <h1>Financial Report</h1>
        <table>
            <thead>
                <tr>
                    <th>Metric</th>
                    <th>Amount</th>
                </tr>
            </thead>
            <tbody>
                <tr>
                    <td>Total Receivables</td>
                    <td>&#8377;{total_receivables:.2f}</td>
                </tr>
                <tr>
                    <td>Total Payables</td>
                    <td>&#8377;{total_payables:.2f}</td>
                </tr>
                <tr>
                    <td>Net Position</td>
                    <td>&#8377;{total_receivables - total_payables:.2f}</td>
                </tr>
            </tbody>
        </table>
        <script>window.print();</script>
    </body>
    </html>
    """
    
    response.write(html_content)
    return response

@login_required
def export_financial_excel(request):
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="financial_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Metric', 'Amount'])
    
    # Get basic financial data
    total_receivables = 0
    for invoice in SalesInvoiceMaster.objects.all():
        balance = invoice.sales_invoice_total - invoice.sales_invoice_paid
        if balance > 0:
            total_receivables += balance
    
    total_payables = InvoiceMaster.objects.aggregate(
        total=Sum(F('invoice_total') - F('invoice_paid'))
    )['total'] or 0
    
    writer.writerow(['Total Receivables', total_receivables])
    writer.writerow(['Total Payables', total_payables])
    writer.writerow(['Net Position', total_receivables - total_payables])
    
    return response

# Finance views
@login_required
def payment_list(request):
    payments = PaymentMaster.objects.all().order_by('-payment_date')
    context = {
        'payments': payments,
        'title': 'Payments'
    }
    return render(request, 'finance/payment_list.html', context)

@login_required
def add_payment(request):
    if request.method == 'POST':
        payment = PaymentMaster.objects.create(
            payment_date=request.POST.get('payment_date'),
            payment_amount=request.POST.get('payment_amount'),
            payment_method=request.POST.get('payment_method'),
            payment_description=request.POST.get('payment_description'),
            payment_reference=request.POST.get('payment_reference')
        )
        messages.success(request, 'Payment added successfully!')
        return redirect('payment_list')
    
    context = {'title': 'Add Payment'}
    return render(request, 'finance/payment_form.html', context)

@login_required
def receipt_list(request):
    receipts = ReceiptMaster.objects.all().order_by('-receipt_date')
    context = {
        'receipts': receipts,
        'title': 'Receipts'
    }
    return render(request, 'finance/receipt_list.html', context)

@login_required
def add_receipt(request):
    if request.method == 'POST':
        receipt = ReceiptMaster.objects.create(
            receipt_date=request.POST.get('receipt_date'),
            receipt_amount=request.POST.get('receipt_amount'),
            receipt_method=request.POST.get('receipt_method'),
            receipt_description=request.POST.get('receipt_description'),
            receipt_reference=request.POST.get('receipt_reference')
        )
        messages.success(request, 'Receipt added successfully!')
        return redirect('receipt_list')
    
    context = {'title': 'Add Receipt'}
    return render(request, 'finance/receipt_form.html', context)

@login_required
@login_required
def sale_rate_list(request):
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Get search query
    search_query = request.GET.get('search', '')
    
    # Get all sale rates directly from SaleRateMaster
    sale_rates = SaleRateMaster.objects.select_related('productid').all()
    
    # Apply search filter
    if search_query:
        sale_rates = sale_rates.filter(
            Q(productid__product_name__icontains=search_query) |
            Q(productid__product_company__icontains=search_query) |
            Q(product_batch_no__icontains=search_query)
        )
    
    # Order by product name and batch
    sale_rates = sale_rates.order_by('productid__product_name', 'product_batch_no')
    
    # Pagination
    paginator = Paginator(sale_rates, 50)
    page_number = request.GET.get('page')
    sale_rates_page = paginator.get_page(page_number)
    
    context = {
        'sale_rates': sale_rates_page,
        'search_query': search_query,
        'title': 'Batch Rates'
    }
    return render(request, 'rates/sale_rate_list.html', context)

@login_required
def add_sale_rate(request):
    if request.method == 'POST':
        form = SaleRateForm(request.POST)
        if form.is_valid():
            try:
                # Check if rate already exists for this product and batch
                existing_rate = SaleRateMaster.objects.filter(
                    productid=form.cleaned_data['productid'],
                    product_batch_no=form.cleaned_data['product_batch_no']
                ).first()
                
                if existing_rate:
                    # Update existing rate
                    existing_rate.rate_A = form.cleaned_data['rate_A']
                    existing_rate.rate_B = form.cleaned_data['rate_B']
                    existing_rate.rate_C = form.cleaned_data['rate_C']
                    existing_rate.save()
                    messages.success(request, 'Sale rate updated successfully!')
                else:
                    # Create new rate
                    form.save()
                    messages.success(request, 'Sale rate added successfully!')
                
                return redirect('sale_rate_list')
            except Exception as e:
                messages.error(request, f'Error saving sale rate: {str(e)}')
        else:
            # Form validation errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = SaleRateForm()
    
    context = {
        'form': form,
        'title': 'Add Sale Rate'
    }
    return render(request, 'rates/sale_rate_form.html', context)

@login_required
def update_sale_rate(request, pk):
    rate = get_object_or_404(SaleRateMaster, id=pk)
    
    if request.method == 'POST':
        form = SaleRateForm(request.POST, instance=rate)
        if form.is_valid():
            form.save()
            messages.success(request, 'Sale rate updated successfully!')
            return redirect('sale_rate_list')
    else:
        form = SaleRateForm(instance=rate)
    
    context = {
        'form': form,
        'rate': rate,
        'title': 'Update Sale Rate'
    }
    return render(request, 'rates/sale_rate_form.html', context)

@login_required
def delete_sale_rate(request, pk):
    rate = get_object_or_404(SaleRateMaster, id=pk)
    
    if request.method == 'POST':
        rate.delete()
        messages.success(request, 'Sale rate deleted successfully!')
        return redirect('sale_rate_list')
    
    context = {
        'rate': rate,
        'title': 'Delete Sale Rate'
    }
    return render(request, 'rates/sale_rate_confirm_delete.html', context)

@login_required
def get_sales_analytics_api(request):
    """API endpoint for sales analytics data"""
    from datetime import datetime
    from .sales_analytics import get_sales_analytics
    
    try:
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        start_date = None
        end_date = None
        
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        analytics_data = get_sales_analytics(start_date, end_date)
        
        return JsonResponse({
            'success': True,
            'data': analytics_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def get_product_batches(request):
    product_id = request.GET.get('product_id')
    
    if not product_id:
        return JsonResponse({'error': 'Product ID is required'}, status=400)
    
    try:
        # Get all unique batches for this product from purchase records
        batches = PurchaseMaster.objects.filter(
            productid=product_id
        ).values(
            'product_batch_no',
            'product_expiry',
            'product_MRP'
        ).distinct()
        
        batch_list = []
        for batch in batches:
            # Get stock for this batch
            batch_quantity, is_available = get_batch_stock_status(
                product_id, batch['product_batch_no']
            )
            
            batch_list.append({
                'batch_no': batch['product_batch_no'],
                'expiry': batch['product_expiry'],
                'stock': batch_quantity,
                'mrp': float(batch['product_MRP'] or 0),
                'is_available': is_available
            })
        
        return JsonResponse({
            'success': True,
            'batches': batch_list
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def get_batch_details(request):
    product_id = request.GET.get('product_id')
    batch_no = request.GET.get('batch_no')
    
    if not product_id or not batch_no:
        return JsonResponse({'error': 'Missing parameters'}, status=400)
    
    try:
        # Get batch details from purchase records
        purchase = PurchaseMaster.objects.filter(
            productid=product_id,
            product_batch_no=batch_no
        ).first()
        
        if purchase:
            # Get available stock
            batch_quantity, is_available = get_batch_stock_status(product_id, batch_no)
            
            # Get sale rates if available
            try:
                sale_rate = SaleRateMaster.objects.get(
                    productid=product_id,
                    product_batch_no=batch_no
                )
                rates = {
                    'rate_A': float(sale_rate.rate_A or 0),
                    'rate_B': float(sale_rate.rate_B or 0),
                    'rate_C': float(sale_rate.rate_C or 0)
                }
            except SaleRateMaster.DoesNotExist:
                rates = {'rate_A': 0, 'rate_B': 0, 'rate_C': 0}
            
            # Handle expiry date formatting
            expiry_str = ''
            if purchase.product_expiry:
                try:
                    if hasattr(purchase.product_expiry, 'strftime'):
                        expiry_str = purchase.product_expiry.strftime('%m-%Y')
                    else:
                        expiry_str = str(purchase.product_expiry)
                except:
                    expiry_str = str(purchase.product_expiry)
            
            return JsonResponse({
                'mrp': float(purchase.product_MRP or 0),
                'expiry': expiry_str,
                'available_stock': batch_quantity,
                'is_available': is_available,
                'rates': rates
            })
        else:
            return JsonResponse({'error': 'Batch not found'}, status=404)
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def get_product_batch_selector(request):
    """API endpoint for Alt+W batch selection dialog"""
    product_id = request.GET.get('product_id')
    
    if not product_id:
        return JsonResponse({'error': 'Product ID is required'}, status=400)
    
    try:
        # Get all batches for this product with details
        batches = PurchaseMaster.objects.filter(
            productid=product_id
        ).values(
            'product_batch_no',
            'product_expiry', 
            'product_MRP'
        ).distinct().order_by('product_batch_no')
        
        batch_options = []
        for batch in batches:
            # Get stock for each batch
            batch_quantity, is_available = get_batch_stock_status(
                product_id, batch['product_batch_no']
            )
            
            # Get sale rates if available
            try:
                sale_rate = SaleRateMaster.objects.get(
                    productid=product_id,
                    product_batch_no=batch['product_batch_no']
                )
                rates = {
                    'rate_A': float(sale_rate.rate_A or 0),
                    'rate_B': float(sale_rate.rate_B or 0), 
                    'rate_C': float(sale_rate.rate_C or 0)
                }
            except SaleRateMaster.DoesNotExist:
                rates = {'rate_A': 0, 'rate_B': 0, 'rate_C': 0}
            
            # Handle expiry date formatting
            expiry_display = 'N/A'
            if batch['product_expiry']:
                try:
                    if hasattr(batch['product_expiry'], 'strftime'):
                        expiry_display = batch['product_expiry'].strftime('%d/%m/%Y')
                    else:
                        expiry_display = str(batch['product_expiry'])
                except:
                    expiry_display = str(batch['product_expiry'])
            
            batch_options.append({
                'batch_no': batch['product_batch_no'],
                'expiry': expiry_display,
                'mrp': float(batch['product_MRP'] or 0),
                'stock': batch_quantity,
                'is_available': is_available,
                'rates': rates
            })
        
        return JsonResponse({
            'success': True,
            'batches': batch_options
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def search_products_api(request):
    """API endpoint for product search functionality"""
    search_term = request.GET.get('search', '').strip()
    
    if not search_term or len(search_term) < 2:
        return JsonResponse({'success': False, 'error': 'Search term too short'}, status=400)
    
    try:
        # Search products by name or company
        products = ProductMaster.objects.filter(
            Q(product_name__icontains=search_term) |
            Q(product_company__icontains=search_term)
        ).order_by('product_name')[:20]  # Limit to 20 results
        
        product_list = []
        for product in products:
            product_list.append({
                'id': product.productid,
                'name': product.product_name,
                'company': product.product_company,
                'category': product.product_category or 'N/A'
            })
        
        return JsonResponse({
            'success': True,
            'products': product_list
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def get_customer_rate_info(request):
    """API endpoint to get customer rate information"""
    customer_id = request.GET.get('customer_id')
    
    if not customer_id:
        return JsonResponse({'success': False, 'error': 'Customer ID required'}, status=400)
    
    try:
        customer = CustomerMaster.objects.get(customerid=customer_id)
        
        # Map customer type to rate type
        rate_mapping = {
            'A': 'Rate A',
            'B': 'Rate B', 
            'C': 'Rate C',
            'Type-A': 'Rate A',
            'Type-B': 'Rate B',
            'Type-C': 'Rate C'
        }
        
        customer_type = customer.customer_type or 'A'
        rate_type = rate_mapping.get(customer_type, 'Rate A')
        
        return JsonResponse({
            'success': True,
            'customer_type': customer_type,
            'rate_type': rate_type
        })
        
    except CustomerMaster.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Customer not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# Missing finance view functions
@login_required
def edit_payment(request, pk):
    payment = get_object_or_404(PaymentMaster, payment_id=pk)
    
    if request.method == 'POST':
        payment.payment_date = request.POST.get('payment_date')
        payment.payment_amount = request.POST.get('payment_amount')
        payment.payment_method = request.POST.get('payment_method')
        payment.payment_description = request.POST.get('payment_description')
        payment.payment_reference = request.POST.get('payment_reference')
        payment.save()
        messages.success(request, 'Payment updated successfully!')
        return redirect('payment_list')
    
    context = {
        'payment': payment,
        'title': 'Edit Payment'
    }
    return render(request, 'finance/payment_form.html', context)

@login_required
def delete_payment(request, pk):
    payment = get_object_or_404(PaymentMaster, payment_id=pk)
    
    if request.method == 'POST':
        payment.delete()
        messages.success(request, 'Payment deleted successfully!')
        return redirect('payment_list')
    
    context = {
        'payment': payment,
        'title': 'Delete Payment'
    }
    return render(request, 'finance/payment_confirm_delete.html', context)

@login_required
def edit_receipt(request, pk):
    receipt = get_object_or_404(ReceiptMaster, receipt_id=pk)
    
    if request.method == 'POST':
        receipt.receipt_date = request.POST.get('receipt_date')
        receipt.receipt_amount = request.POST.get('receipt_amount')
        receipt.receipt_method = request.POST.get('receipt_method')
        receipt.receipt_description = request.POST.get('receipt_description')
        receipt.receipt_reference = request.POST.get('receipt_reference')
        receipt.save()
        messages.success(request, 'Receipt updated successfully!')
        return redirect('receipt_list')
    
    context = {
        'receipt': receipt,
        'title': 'Edit Receipt'
    }
    return render(request, 'finance/receipt_form.html', context)

@login_required
def delete_receipt(request, pk):
    receipt = get_object_or_404(ReceiptMaster, receipt_id=pk)
    
    if request.method == 'POST':
        receipt.delete()
        messages.success(request, 'Receipt deleted successfully!')
        return redirect('receipt_list')
    
    context = {
        'receipt': receipt,
        'title': 'Delete Receipt'
    }
    return render(request, 'finance/receipt_confirm_delete.html', context)

@login_required
def get_sales_analytics_api(request):
    """API endpoint for real-time sales analytics data"""
    from datetime import datetime
    from .sales_analytics import get_sales_analytics
    
    # Get date range
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None
    except ValueError:
        start_date = end_date = None
    
    # Get comprehensive analytics
    analytics_data = get_sales_analytics(start_date, end_date)
    
    return JsonResponse({
        'success': True,
        'data': analytics_data
    })



# Export functions for finance
@login_required
def export_payments_pdf(request):
    return JsonResponse({'status': 'ok'})

@login_required
def export_payments_excel(request):
    return JsonResponse({'status': 'ok'})

@login_required
def export_payments_print(request):
    return JsonResponse({'status': 'ok'})

@login_required
def export_receipts_pdf(request):
    return JsonResponse({'status': 'ok'})

@login_required
def export_receipts_excel(request):
    return JsonResponse({'status': 'ok'})

@login_required
def export_receipts_print(request):
    return JsonResponse({'status': 'ok'})

@login_required
def get_customer_rate_info(request):
    customer_id = request.GET.get('customer_id')
    
    if not customer_id:
        return JsonResponse({'error': 'Customer ID is required'}, status=400)
    
    try:
        customer = CustomerMaster.objects.get(customerid=customer_id)
        
        # Extract rate type from customer type
        customer_type = customer.customer_type
        if 'A' in customer_type or customer_type == 'TYPE-A':
            rate_type = 'A'
        elif 'B' in customer_type or customer_type == 'TYPE-B':
            rate_type = 'B'
        elif 'C' in customer_type or customer_type == 'TYPE-C':
            rate_type = 'C'
        else:
            rate_type = 'A'  # Default
        
        return JsonResponse({
            'success': True,
            'customer_type': customer_type,
            'rate_type': rate_type
        })
        
    except CustomerMaster.DoesNotExist:
        return JsonResponse({'error': 'Customer not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def search_products_api(request):
    """API endpoint for product search"""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'products': []})
    
    try:
        products = ProductMaster.objects.filter(
            Q(product_name__icontains=query) |
            Q(product_company__icontains=query)
        ).order_by('product_name')[:10]
        
        product_list = []
        for product in products:
            product_list.append({
                'id': product.productid,
                'name': product.product_name,
                'company': product.product_company,
                'packing': product.product_packing
            })
        
        return JsonResponse({'products': product_list})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def get_sales_analytics_api(request):
    """API endpoint for sales analytics"""
    try:
        # Basic analytics data
        total_invoices = SalesInvoiceMaster.objects.count()
        total_sales = SalesMaster.objects.aggregate(Sum('sale_total_amount'))['sale_total_amount__sum'] or 0
        
        return JsonResponse({
            'success': True,
            'total_invoices': total_invoices,
            'total_sales': total_sales
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Sale Rate Management Views
@login_required
@login_required
def add_sale_rate(request):
    if request.method == 'POST':
        form = SaleRateForm(request.POST)
        if form.is_valid():
            sale_rate = form.save()
            messages.success(request, f"Sale rate for {sale_rate.productid.product_name} - Batch {sale_rate.product_batch_no} added successfully!")
            return redirect('sale_rate_list')
    else:
        form = SaleRateForm()
    
    context = {
        'form': form,
        'title': 'Add Sale Rate'
    }
    return render(request, 'rates/sale_rate_form.html', context)

@login_required
def update_sale_rate(request, pk):
    sale_rate = get_object_or_404(SaleRateMaster, id=pk)
    
    if request.method == 'POST':
        form = SaleRateForm(request.POST, instance=sale_rate)
        if form.is_valid():
            form.save()
            messages.success(request, f"Sale rate for {sale_rate.productid.product_name} - Batch {sale_rate.product_batch_no} updated successfully!")
            return redirect('sale_rate_list')
    else:
        form = SaleRateForm(instance=sale_rate)
    
    context = {
        'form': form,
        'sale_rate': sale_rate,
        'title': 'Update Sale Rate'
    }
    return render(request, 'rates/sale_rate_form.html', context)

@login_required
def delete_sale_rate(request, pk):
    if not request.user.user_type.lower() in ['admin']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('sale_rate_list')
        
    sale_rate = get_object_or_404(SaleRateMaster, id=pk)
    
    if request.method == 'POST':
        product_name = sale_rate.productid.product_name
        batch_no = sale_rate.product_batch_no
        try:
            sale_rate.delete()
            messages.success(request, f"Sale rate for {product_name} - Batch {batch_no} deleted successfully!")
        except Exception as e:
            messages.error(request, f"Cannot delete sale rate. Error: {str(e)}")
        return redirect('sale_rate_list')
    
    context = {
        'sale_rate': sale_rate,
        'title': 'Delete Sale Rate'
    }
    return render(request, 'rates/sale_rate_confirm_delete.html', context)

# Missing API endpoints
@login_required
def get_sales_analytics_api(request):
    return JsonResponse({'status': 'gtt'})


@login_required
def export_purchases_excel(request):
    """
    Export purchases data to Excel using OpenPyXL
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from datetime import datetime, date
    from .models import PurchaseMaster, InvoiceMaster, SupplierMaster, ProductMaster
    import io

    def parse_date(date_str):
        """Parse date from various formats including DDMM"""
        if not date_str:
            return None
        
        if not date_str.strip():
            return None
            
        # Handle DDMM format (4 digits)
        if len(date_str) == 4 and date_str.isdigit():
            try:
                day = int(date_str[:2])
                month = int(date_str[2:4])
                year = datetime.now().year
                return date(year, month, day)
            except ValueError:
                return None
        
        # Handle DD/MM format
        if len(date_str) == 5 and '/' in date_str:
            try:
                day, month = date_str.split('/')
                year = datetime.now().year
                return date(year, int(month), int(day))
            except ValueError:
                return None
        
        # Handle YYYY-MM-DD format
        if len(date_str) == 10 and '-' in date_str:
            try:
                return datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return None
        
        # Handle DD-MM-YYYY format
        if len(date_str) == 10 and '-' in date_str:
            try:
                return datetime.strptime(date_str, '%d-%m-%Y').date()
            except ValueError:
                return None
        
        return None

    try:
        # Get filter parameters
        start_date_str = request.GET.get('start_date', '')
        end_date_str = request.GET.get('end_date', '')
        supplier_id = request.GET.get('supplier_id', '')
        invoice_no = request.GET.get('invoice_no', '')
        
        # Parse dates with DDMM support
        today = date.today()
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
        
        # Set default dates if parsing failed or not provided
        if not start_date:
            start_date = today.replace(day=1)  # First day of current month
            
        if not end_date:
            end_date = today

        # Get purchases data
        purchases_data = PurchaseMaster.objects.filter(
            product_invoiceid__invoice_date__range=[start_date, end_date]
        ).select_related(
            'product_invoiceid',
            'product_invoiceid__supplierid',
            'productid'
        )

        # Apply filters
        if supplier_id:
            purchases_data = purchases_data.filter(product_invoiceid__supplierid=supplier_id)
            
        if invoice_no:
            purchases_data = purchases_data.filter(product_invoice_no__icontains=invoice_no)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Purchase Report"

        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        normal_font = Font(name='Arial', size=10)
        title_font = Font(name='Arial', size=14, bold=True)
        header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")  # Green header
        currency_format = '#,##0.00'
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title and headers
        ws.merge_cells('A1:J1')
        ws['A1'] = "PURCHASE ORDER REPORT"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:J2')
        ws['A2'] = f"Report Period: {start_date.strftime('%d-%m-%Y')} to {end_date.strftime('%d-%m-%Y')}"
        ws['A2'].font = Font(bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:J3')
        ws['A3'] = f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        ws['A3'].alignment = Alignment(horizontal='center')
        
        if supplier_id:
            supplier = SupplierMaster.objects.filter(supplierid=supplier_id).first()
            if supplier:
                ws.merge_cells('A4:J4')
                ws['A4'] = f"Supplier: {supplier.supplier_name}"
                ws['A4'].alignment = Alignment(horizontal='center')

        # Column headers
        headers = [
            'S.No', 'Invoice No', 'Date', 'Supplier', 'Product', 
            'Batch No', 'Expiry', 'Qty', 'Purchase Rate', 'Total Amount'
        ]
        
        header_row = 6
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Data rows
        row_num = header_row + 1
        total_amount = 0
        total_quantity = 0
        serial_number = 1
        
        for purchase in purchases_data:
            # Serial number
            ws.cell(row=row_num, column=1, value=serial_number)
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
            
            # Invoice details
            ws.cell(row=row_num, column=2, value=purchase.product_invoiceid.invoice_no)
            ws.cell(row=row_num, column=3, value=purchase.product_invoiceid.invoice_date.strftime('%d-%m-%Y'))
            
            # Supplier
            ws.cell(row=row_num, column=4, value=purchase.product_invoiceid.supplierid.supplier_name)
            
            # Product details
            ws.cell(row=row_num, column=5, value=purchase.product_name)
            ws.cell(row=row_num, column=6, value=purchase.product_batch_no)
            
            # Expiry date
            expiry_display = 'N/A'
            if purchase.product_expiry:
                if isinstance(purchase.product_expiry, str):
                    expiry_display = purchase.product_expiry
                else:
                    expiry_display = purchase.product_expiry.strftime('%m-%Y')
            ws.cell(row=row_num, column=7, value=expiry_display)
            ws.cell(row=row_num, column=7).alignment = Alignment(horizontal='center')
            
            # Numeric values
            ws.cell(row=row_num, column=8, value=float(purchase.product_quantity))
            ws.cell(row=row_num, column=8).alignment = Alignment(horizontal='center')
            
            ws.cell(row=row_num, column=9, value=float(purchase.product_actual_rate))
            ws.cell(row=row_num, column=9).number_format = currency_format
            
            ws.cell(row=row_num, column=10, value=float(purchase.total_amount))
            ws.cell(row=row_num, column=10).number_format = currency_format
            
            # Add borders to all cells
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).border = thin_border
                if col not in [1, 7, 8]:  # Center align serial, expiry, and quantity
                    ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='left', vertical='center')
            
            total_amount += float(purchase.total_amount)
            total_quantity += float(purchase.product_quantity)
            serial_number += 1
            row_num += 1

        # Summary section
        if row_num > header_row + 1:
            # Empty row
            row_num += 1
            
            # Total row for amount
            ws.merge_cells(f'A{row_num}:I{row_num}')
            total_cell = ws.cell(row=row_num, column=1, value="TOTAL PURCHASE AMOUNT:")
            total_cell.font = Font(bold=True, size=12)
            total_cell.alignment = Alignment(horizontal='right')
            
            amount_cell = ws.cell(row=row_num, column=10, value=total_amount)
            amount_cell.font = Font(bold=True, size=12)
            amount_cell.number_format = currency_format
            
            # Add border and color to total row
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).border = thin_border
                ws.cell(row=row_num, column=col).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

            # Quantity summary row
            row_num += 1
            ws.merge_cells(f'A{row_num}:I{row_num}')
            qty_cell = ws.cell(row=row_num, column=1, value="TOTAL QUANTITY:")
            qty_cell.font = Font(bold=True)
            qty_cell.alignment = Alignment(horizontal='right')
            
            qty_value_cell = ws.cell(row=row_num, column=10, value=total_quantity)
            qty_value_cell.font = Font(bold=True)
            qty_value_cell.alignment = Alignment(horizontal='center')

            # Statistics row
            row_num += 1
            ws.merge_cells(f'A{row_num}:I{row_num}')
            stats_cell = ws.cell(row=row_num, column=1, value=f"Total Purchase Items: {serial_number - 1}")
            stats_cell.font = Font(bold=True)
            stats_cell.alignment = Alignment(horizontal='right')

        else:
            # No data message
            ws.merge_cells(f'A{row_num}:J{row_num}')
            ws.cell(row=row_num, column=1, value="No purchase data found for the selected period.")
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=1).font = Font(italic=True, color="FF0000")

        # Auto-adjust column widths
        column_widths = {
            'A': 8,   # S.No
            'B': 15,  # Invoice No
            'C': 12,  # Date
            'D': 25,  # Supplier
            'E': 35,  # Product
            'F': 15,  # Batch No
            'G': 10,  # Expiry
            'H': 8,   # Qty
            'I': 15,  # Purchase Rate
            'J': 15   # Total Amount
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Freeze header row
        ws.freeze_panes = ws['A7']

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"purchase_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        return response

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Purchase Excel generation error: {str(e)}")
        print(f"Error details: {error_details}")
        
        from django.http import JsonResponse
        return JsonResponse({
            'error': f'Failed to generate Excel: {str(e)}',
            'details': 'Check the date format. Use DDMM, DD/MM, or YYYY-MM-DD format'
        }, status=500)

@login_required
def export_financial_pdf(request):
    return JsonResponse({'status': 'ok'})

@login_required
def export_financial_excel(request):
    """
    Export comprehensive financial data to Excel with all specified metrics
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from datetime import datetime, date, timedelta
    from django.db.models import Sum, Count, Avg, F, Q
    from .models import (
        SalesInvoiceMaster, InvoiceMaster, SalesMaster, PurchaseMaster,
        CustomerMaster, SupplierMaster, ReturnSalesInvoiceMaster, ReturnInvoiceMaster,
        ReturnSalesMaster, ReturnPurchaseMaster
    )
    import io

    def parse_date(date_str):
        """Parse date from various formats including DDMM"""
        if not date_str:
            return None
        if not date_str.strip():
            return None
            
        # Handle DDMM format (4 digits)
        if len(date_str) == 4 and date_str.isdigit():
            try:
                day = int(date_str[:2])
                month = int(date_str[2:4])
                year = datetime.now().year
                return date(year, month, day)
            except ValueError:
                return None
        
        # Handle DD/MM format
        if len(date_str) == 5 and '/' in date_str:
            try:
                day, month = date_str.split('/')
                year = datetime.now().year
                return date(year, int(month), int(day))
            except ValueError:
                return None
        
        # Handle YYYY-MM-DD format
        if len(date_str) == 10 and '-' in date_str:
            try:
                return datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return None
        
        return None

    try:
        # Get filter parameters
        start_date_str = request.GET.get('start_date', '')
        end_date_str = request.GET.get('end_date', '')
        
        # Parse dates with DDMM support
        today = date.today()
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
        
        # Set default dates if parsing failed or not provided
        if not start_date:
            start_date = today.replace(day=1)  # First day of current month
        if not end_date:
            end_date = today

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Report"

        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        normal_font = Font(name='Arial', size=10)
        title_font = Font(name='Arial', size=14, bold=True)
        highlight_font = Font(name='Arial', size=11, bold=True, color='2E8B57')
        loss_font = Font(name='Arial', size=11, bold=True, color='FF0000')
        currency_format = '#,##0.00'
        
        # Color schemes
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        sales_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
        purchase_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        profit_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        receivables_fill = PatternFill(start_color="FFF0E6", end_color="FFF0E6", fill_type="solid")
        payables_fill = PatternFill(start_color="F0E6FF", end_color="F0E6FF", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title Section
        ws.merge_cells('A1:F1')
        ws['A1'] = "PHARMACY FINANCIAL REPORT"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Date Range: {start_date.strftime('%d-%m-%Y')} to {end_date.strftime('%d-%m-%Y')}"
        ws['A2'].font = Font(bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:F3')
        ws['A3'] = f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        ws['A3'].alignment = Alignment(horizontal='center')

        # SECTION 1: Calculate Financial Metrics
        
        # Gross Sales
        sales_invoices = SalesInvoiceMaster.objects.filter(
            sales_invoice_date__range=[start_date, end_date]
        )
        gross_sales = SalesMaster.objects.filter(
            sales_invoice_no__in=sales_invoices
        ).aggregate(total=Sum('sale_total_amount'))['total'] or 0

        # Gross Purchases
        purchase_invoices = InvoiceMaster.objects.filter(
            invoice_date__range=[start_date, end_date]
        )
        gross_purchases = purchase_invoices.aggregate(total=Sum('invoice_total'))['total'] or 0

        # Sales Returns
        sales_returns = ReturnSalesMaster.objects.filter(
            return_sales_invoice_no__return_sales_invoice_date__range=[start_date, end_date]
        ).aggregate(total=Sum('return_sale_total_amount'))['total'] or 0

        # Purchase Returns
        purchase_returns = ReturnPurchaseMaster.objects.filter(
            returninvoiceid__returninvoice_date__range=[start_date, end_date]
        ).aggregate(total=Sum('returntotal_amount'))['total'] or 0

        # Net Calculations
        net_sales = gross_sales - sales_returns
        net_purchases = gross_purchases - purchase_returns
        gross_profit = net_sales - net_purchases

        # SECTION 2: Financial Summary Table
        current_row = 5
        
        # Table Headers
        headers = ['FINANCIAL METRIC', 'AMOUNT (₹)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        current_row += 1

        # Financial Data
        financial_data = [
            ('Gross Sales', gross_sales, sales_fill),
            ('Gross Purchases', gross_purchases, purchase_fill),
            ('Sales Returns', sales_returns, sales_fill),
            ('Purchase Returns', purchase_returns, purchase_fill),
            ('Net Sales (After Returns)', net_sales, sales_fill),
            ('Net Purchases (After Returns)', net_purchases, purchase_fill),
            ('Gross Profit', gross_profit, profit_fill),
        ]

        for metric, amount, fill_color in financial_data:
            ws.cell(row=current_row, column=1, value=metric).font = normal_font
            ws.cell(row=current_row, column=1).border = thin_border
            ws.cell(row=current_row, column=1).fill = fill_color
            
            amount_cell = ws.cell(row=current_row, column=2, value=float(amount))
            amount_cell.number_format = currency_format
            amount_cell.border = thin_border
            amount_cell.fill = fill_color
            
            # Highlight profit/loss
            if metric == 'Gross Profit':
                amount_cell.font = highlight_font if amount >= 0 else loss_font
            else:
                amount_cell.font = normal_font
                
            current_row += 1

        current_row += 2  # Add spacing

        # SECTION 3: Monthly Sales Trend (Past 12 Months)
        ws.merge_cells(f'A{current_row}:F{current_row}')
        title_cell = ws.cell(row=current_row, column=1, value="Monthly Sales Trend (Past 12 Months)")
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 1

        # Calculate monthly sales for past 12 months
        monthly_data = []
        for i in range(11, -1, -1):
            month_date = today.replace(day=1) - timedelta(days=30*i)
            month_start = month_date.replace(day=1)
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year+1, month=1, day=1) - timedelta(days=1)
            else:
                month_end = month_start.replace(month=month_start.month+1, day=1) - timedelta(days=1)
            
            month_sales = SalesMaster.objects.filter(
                sales_invoice_no__sales_invoice_date__range=[month_start, month_end]
            ).aggregate(total=Sum('sale_total_amount'))['total'] or 0
            
            monthly_data.append({
                'month': month_start.strftime('%b %Y'),
                'sales': month_sales
            })

        # Monthly Sales Headers
        monthly_headers = ['Month', 'Sales Amount (₹)']
        for col, header in enumerate(monthly_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = sales_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

        # Monthly Sales Data
        for data in monthly_data:
            ws.cell(row=current_row, column=1, value=data['month']).border = thin_border
            ws.cell(row=current_row, column=1).fill = sales_fill
            
            sales_cell = ws.cell(row=current_row, column=2, value=float(data['sales']))
            sales_cell.number_format = currency_format
            sales_cell.border = thin_border
            sales_cell.fill = sales_fill
            
            current_row += 1

        current_row += 2  # Add spacing

        # SECTION 4: Outstanding Receivables (Top Customers)
        ws.merge_cells(f'A{current_row}:F{current_row}')
        title_cell = ws.cell(row=current_row, column=1, value="Outstanding Receivables (Top Customers)")
        title_cell.font = header_font
        title_cell.fill = receivables_fill
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 1

        # Calculate outstanding receivables
        receivables_data = []
        total_receivables = 0
        
        for invoice in SalesInvoiceMaster.objects.all():
            # Get actual invoice total from sales items
            invoice_total = SalesMaster.objects.filter(
                sales_invoice_no=invoice.sales_invoice_no
            ).aggregate(Sum('sale_total_amount'))['sale_total_amount__sum'] or 0
            
            balance = invoice_total - invoice.sales_invoice_paid
            if balance > 0:
                receivables_data.append({
                    'customer': invoice.customerid.customer_name,
                    'amount': balance
                })
                total_receivables += balance

        # Sort by amount descending and take top 10
        receivables_data.sort(key=lambda x: x['amount'], reverse=True)
        top_receivables = receivables_data[:10]

        # Receivables Headers
        receivables_headers = ['Customer', 'Outstanding Amount (₹)']
        for col, header in enumerate(receivables_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = receivables_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

        # Receivables Data
        if top_receivables:
            for data in top_receivables:
                ws.cell(row=current_row, column=1, value=data['customer']).border = thin_border
                ws.cell(row=current_row, column=1).fill = receivables_fill
                
                amount_cell = ws.cell(row=current_row, column=2, value=float(data['amount']))
                amount_cell.number_format = currency_format
                amount_cell.border = thin_border
                amount_cell.fill = receivables_fill
                
                current_row += 1
        else:
            ws.merge_cells(f'A{current_row}:B{current_row}')
            ws.cell(row=current_row, column=1, value="No outstanding receivables").alignment = Alignment(horizontal='center')
            ws.cell(row=current_row, column=1).font = Font(italic=True)
            current_row += 1

        # Total Receivables
        ws.merge_cells(f'A{current_row}:B{current_row}')
        total_cell = ws.cell(row=current_row, column=1, value=f"Total Receivables: ₹{total_receivables:,.2f}")
        total_cell.font = highlight_font
        total_cell.alignment = Alignment(horizontal='right')
        current_row += 2

        # SECTION 5: Outstanding Payables (Top Suppliers)
        ws.merge_cells(f'A{current_row}:F{current_row}')
        title_cell = ws.cell(row=current_row, column=1, value="Outstanding Payables (Top Suppliers)")
        title_cell.font = header_font
        title_cell.fill = payables_fill
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 1

        # Calculate outstanding payables
        payables_data = []
        total_payables = 0
        
        for invoice in InvoiceMaster.objects.all():
            balance = invoice.invoice_total - invoice.invoice_paid
            if balance > 0:
                payables_data.append({
                    'supplier': invoice.supplierid.supplier_name,
                    'amount': balance
                })
                total_payables += balance

        # Sort by amount descending and take top 10
        payables_data.sort(key=lambda x: x['amount'], reverse=True)
        top_payables = payables_data[:10]

        # Payables Headers
        payables_headers = ['Supplier', 'Outstanding Amount (₹)']
        for col, header in enumerate(payables_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = payables_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

        # Payables Data
        if top_payables:
            for data in top_payables:
                ws.cell(row=current_row, column=1, value=data['supplier']).border = thin_border
                ws.cell(row=current_row, column=1).fill = payables_fill
                
                amount_cell = ws.cell(row=current_row, column=2, value=float(data['amount']))
                amount_cell.number_format = currency_format
                amount_cell.border = thin_border
                amount_cell.fill = payables_fill
                
                current_row += 1
        else:
            ws.merge_cells(f'A{current_row}:B{current_row}')
            ws.cell(row=current_row, column=1, value="No outstanding payables").alignment = Alignment(horizontal='center')
            ws.cell(row=current_row, column=1).font = Font(italic=True)
            current_row += 1

        # Total Payables
        ws.merge_cells(f'A{current_row}:B{current_row}')
        total_cell = ws.cell(row=current_row, column=1, value=f"Total Payables: ₹{total_payables:,.2f}")
        total_cell.font = highlight_font
        total_cell.alignment = Alignment(horizontal='right')

        # SECTION 6: Additional Financial Ratios (Bonus Section)
        current_row += 3
        
        ws.merge_cells(f'A{current_row}:F{current_row}')
        title_cell = ws.cell(row=current_row, column=1, value="Financial Ratios & Analysis")
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 1

        # Calculate ratios
        profit_margin = (gross_profit / net_sales * 100) if net_sales > 0 else 0
        collection_efficiency = ((net_sales - total_receivables) / net_sales * 100) if net_sales > 0 else 0
        payment_efficiency = ((net_purchases - total_payables) / net_purchases * 100) if net_purchases > 0 else 0

        ratios_data = [
            ('Gross Profit Margin', f'{profit_margin:.1f}%'),
            ('Collection Efficiency', f'{collection_efficiency:.1f}%'),
            ('Payment Efficiency', f'{payment_efficiency:.1f}%'),
            ('Net Cash Position', f'₹{(total_receivables - total_payables):,.2f}'),
        ]

        ratio_headers = ['Ratio', 'Value']
        for col, header in enumerate(ratio_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

        for ratio, value in ratios_data:
            ws.cell(row=current_row, column=1, value=ratio).border = thin_border
            ws.cell(row=current_row, column=2, value=value).border = thin_border
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
            current_row += 1

        # FIXED: Safe column width adjustment
        # Manually set column widths instead of auto-adjusting
        column_widths = {
            'A': 35,  # Metric names
            'B': 20,  # Amounts
            'C': 15,  # Extra columns for merged cells
            'D': 15,
            'E': 15,
            'F': 15
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"financial_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Financial Excel generation error: {str(e)}")
        print(f"Error details: {error_details}")
        
        from django.http import JsonResponse
        return JsonResponse({
            'error': f'Failed to generate Financial Excel: {str(e)}',
            'details': 'Check the date format. Use DDMM, DD/MM, or YYYY-MM-DD format'
        }, status=500)

# Finance payment function
@login_required
def payment_list(request):
    payments = PaymentMaster.objects.all().order_by('-payment_date')
    
    paginator = Paginator(payments, 20)
    page_number = request.GET.get('page')
    payments = paginator.get_page(page_number)
    
    context = {
        'payments': payments,
        'title': 'Payments'
    }
    return render(request, 'finance/payment_list.html', context)

@login_required

def add_payment(request):
    from django.shortcuts import render, get_object_or_404, redirect
    from django.urls import reverse
    from .models import PaymentMaster
    from .forms import PaymentForm
    if request.method == "POST":
        form = PaymentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('payment_list')
    else:
        form = PaymentForm()
    
    context = {
        'form': form,
        'title': 'Add Payment',
    }
    return render(request, 'finance/payment_form.html', context)

@login_required

# Edit payment
def edit_payment(request, payment_id):
    from django.shortcuts import render, get_object_or_404, redirect
    from django.urls import reverse
    from .models import PaymentMaster
    from .forms import PaymentForm
    payment = get_object_or_404(PaymentMaster, id=payment_id)
    title = "Edit Payment"

    if request.method == "POST":
        form = PaymentForm(request.POST, instance=payment)
        if form.is_valid():
            form.save()
            return redirect('payment_list')
    else:
        form = PaymentForm(instance=payment)

    context = {
        'form': form,
        'title': title,
        'payment': payment,
    }
    return render(request, 'finance/payment_form.html', context)


@login_required
def payment_confirm_delete(request):

    from django.shortcuts import render, get_object_or_404, redirect
    from django.urls import reverse
    from .models import Payment  # make sure Payment model is imported

    payment = get_object_or_404(Payment, id=payment_id)
    title = "Delete Payment"

    if request.method == "POST":
        payment.delete()
        return redirect(reverse('payment_list'))  # redirects back to your payment list page

    context = {
        'payment': payment,
        'title': title,
    }
    return render(request, 'finance/payment_confirm_delete.html', context)


def export_payments_excel(request):
    import openpyxl
    from .models import PaymentMaster
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"

    # Add header row
    ws.append([
        "Payment ID", "Date", "Amount", "Method", "Description", "Reference", "Supplier", "Invoice No"
    ])

    # Fetch all payments
    payments = PaymentMaster.objects.all()

    for payment in payments:
        ws.append([
            payment.payment_id,
            payment.payment_date.strftime("%d-%m-%Y"),
            float(payment.payment_amount),
            payment.payment_method,
            payment.payment_description or "",
            payment.payment_reference or "",
            payment.supplier.pharmacyname if payment.supplier else "",
            payment.invoice.invoice_no if payment.invoice else ""
        ])

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=payments.xlsx'

    wb.save(response)
    return response


@login_required
def export_payments_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from .models import PaymentMaster
    # Create HTTP response with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="payments.pdf"'

    # Create PDF canvas
    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Title
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(200, height - 50, "Payments Report")

    # Table headers
    pdf.setFont("Helvetica-Bold", 12)
    headers = ["ID", "Date", "Amount", "Method", "Supplier", "Invoice No", "Description"]
    x_offsets = [40, 80, 150, 220, 300, 380, 460]
    y = height - 80
    for x, header in zip(x_offsets, headers):
        pdf.drawString(x, y, header)

    # Fetch payments
    payments = PaymentMaster.objects.all()
    pdf.setFont("Helvetica", 10)
    y -= 20

    for payment in payments:
        pdf.drawString(x_offsets[0], y, str(payment.payment_id))
        pdf.drawString(x_offsets[1], y, payment.payment_date.strftime("%d-%m-%Y"))
        pdf.drawString(x_offsets[2], y, f"₹{payment.payment_amount:.2f}")
        pdf.drawString(x_offsets[3], y, payment.payment_method)
        pdf.drawString(x_offsets[4], y, payment.supplier.pharmaname if payment.supplier else "")
        pdf.drawString(x_offsets[5], y, payment.invoice.invoice_no if payment.invoice else "")
        pdf.drawString(x_offsets[6], y, payment.payment_description or "")

        y -= 20
        if y < 50:  # create new page if space is low
            pdf.showPage()
            pdf.setFont("Helvetica", 10)
            y = height - 50

    pdf.save()
    return response




#finance recipte function
@login_required
def receipt_list(request):
    receipts = ReceiptMaster.objects.all().order_by('-receipt_date')
    
    paginator = Paginator(receipts, 20)
    page_number = request.GET.get('page')
    receipts = paginator.get_page(page_number)
    
    context = {
        'receipts': receipts,
        'title': 'Receipts'
    }
    return render(request, 'finance/receipt_list.html', context)



@login_required
def add_receipt(request):
    from django.shortcuts import render, redirect
    from django.contrib.auth.decorators import login_required
    from .models import ReceiptMaster
    from .forms import ReceiptForm
    from datetime import datetime

    title = "Add Receipt"

    if request.method == "POST":
        form = ReceiptForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('receipt_list')  # Redirect to the list view after saving
    else:
        form = ReceiptForm(initial={'receipt_date': datetime.now()})

    context = {
        'form': form,
        'title': title,
    }
    return render(request, 'finance/receipt_form.html', context)



@login_required
def edit_receipt(request, pk):
    return JsonResponse({'status': 'ok'})

@login_required
def delete_receipt(request, pk):
    return JsonResponse({'status': 'ok'})



@login_required
def export_sales_excel(request):
    """
    Export sales data to Excel using OpenPyXL - Handles DDMM date format
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from datetime import datetime, date
    from .models import SalesMaster, SalesInvoiceMaster, CustomerMaster
    from django.db.models import Sum, Count, Avg
    import io

    def parse_date(date_str):
        """Parse date from various formats including DDMM"""
        if not date_str:
            return None
        
        # Handle empty string
        if not date_str.strip():
            return None
            
        # Handle DDMM format (4 digits)
        if len(date_str) == 4 and date_str.isdigit():
            try:
                day = int(date_str[:2])
                month = int(date_str[2:4])
                year = datetime.now().year
                return date(year, month, day)
            except ValueError:
                return None
        
        # Handle DD/MM format
        if len(date_str) == 5 and '/' in date_str:
            try:
                day, month = date_str.split('/')
                year = datetime.now().year
                return date(year, int(month), int(day))
            except ValueError:
                return None
        
        # Handle YYYY-MM-DD format
        if len(date_str) == 10 and '-' in date_str:
            try:
                return datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return None
        
        # Handle DD-MM-YYYY format
        if len(date_str) == 10 and '-' in date_str:
            try:
                return datetime.strptime(date_str, '%d-%m-%Y').date()
            except ValueError:
                return None
        
        return None

    try:
        # Get filter parameters
        start_date_str = request.GET.get('start_date', '')
        end_date_str = request.GET.get('end_date', '')
        customer_id = request.GET.get('customer_id', '')
        
        # Parse dates with DDMM support
        today = date.today()
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
        
        # Set default dates if parsing failed or not provided
        if not start_date:
            start_date = today.replace(day=1)  # First day of current month
            
        if not end_date:
            end_date = today

        print(f"Date range: {start_date} to {end_date}")  # Debug log

        # Get sales data
        sales_data = SalesMaster.objects.filter(
            sales_invoice_no__sales_invoice_date__range=[start_date, end_date]
        ).select_related(
            'sales_invoice_no', 
            'sales_invoice_no__customerid',
            'productid'
        )

        if customer_id:
            sales_data = sales_data.filter(sales_invoice_no__customerid=customer_id)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        normal_font = Font(name='Arial', size=10)
        title_font = Font(name='Arial', size=14, bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        currency_format = '#,##0.00'
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title and headers
        ws.merge_cells('A1:I1')
        ws['A1'] = "PHARMACY SALES REPORT"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:I2')
        ws['A2'] = f"Report Period: {start_date.strftime('%d-%m-%Y')} to {end_date.strftime('%d-%m-%Y')}"
        ws['A2'].font = Font(bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:I3')
        ws['A3'] = f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        ws['A3'].alignment = Alignment(horizontal='center')
        
        if customer_id:
            customer = CustomerMaster.objects.filter(customerid=customer_id).first()
            if customer:
                ws.merge_cells('A4:I4')
                ws['A4'] = f"Customer: {customer.customer_name}"
                ws['A4'].alignment = Alignment(horizontal='center')

        # Column headers
        headers = [
            'S.No', 'Invoice No', 'Date', 'Customer', 'Product', 
            'Batch No', 'Quantity', 'Rate (₹)', 'Amount (₹)'
        ]
        
        header_row = 6
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Data rows
        row_num = header_row + 1
        total_amount = 0
        serial_number = 1
        
        for sale in sales_data:
            # Serial number
            ws.cell(row=row_num, column=1, value=serial_number)
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
            
            # Invoice details
            ws.cell(row=row_num, column=2, value=sale.sales_invoice_no.sales_invoice_no)
            ws.cell(row=row_num, column=3, value=sale.sales_invoice_no.sales_invoice_date.strftime('%d-%m-%Y'))
            
            # Customer
            ws.cell(row=row_num, column=4, value=sale.sales_invoice_no.customerid.customer_name)
            
            # Product details
            ws.cell(row=row_num, column=5, value=sale.product_name)
            ws.cell(row=row_num, column=6, value=sale.product_batch_no)
            
            # Numeric values
            ws.cell(row=row_num, column=7, value=float(sale.sale_quantity))
            ws.cell(row=row_num, column=7).alignment = Alignment(horizontal='center')
            
            ws.cell(row=row_num, column=8, value=float(sale.sale_rate))
            ws.cell(row=row_num, column=8).number_format = currency_format
            
            ws.cell(row=row_num, column=9, value=float(sale.sale_total_amount))
            ws.cell(row=row_num, column=9).number_format = currency_format
            
            # Add borders to all cells
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).border = thin_border
                if col not in [1, 7]:  # Center align serial number and quantity
                    ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='left', vertical='center')
            
            total_amount += float(sale.sale_total_amount)
            serial_number += 1
            row_num += 1

        # Summary section
        if row_num > header_row + 1:
            # Empty row
            row_num += 1
            
            # Total row
            ws.merge_cells(f'A{row_num}:H{row_num}')
            total_cell = ws.cell(row=row_num, column=1, value="TOTAL AMOUNT:")
            total_cell.font = Font(bold=True, size=12)
            total_cell.alignment = Alignment(horizontal='right')
            
            amount_cell = ws.cell(row=row_num, column=9, value=total_amount)
            amount_cell.font = Font(bold=True, size=12)
            amount_cell.number_format = currency_format
            
            # Add border to total row
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).border = thin_border
                ws.cell(row=row_num, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Statistics row
            row_num += 1
            ws.merge_cells(f'A{row_num}:H{row_num}')
            stats_cell = ws.cell(row=row_num, column=1, value=f"Total Records: {serial_number - 1}")
            stats_cell.font = Font(bold=True)
            stats_cell.alignment = Alignment(horizontal='right')

        else:
            # No data message
            ws.merge_cells(f'A{row_num}:I{row_num}')
            ws.cell(row=row_num, column=1, value="No sales data found for the selected period.")
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=1).font = Font(italic=True, color="FF0000")

        # Auto-adjust column widths
        column_widths = {
            'A': 8,   # S.No
            'B': 15,  # Invoice No
            'C': 12,  # Date
            'D': 25,  # Customer
            'E': 35,  # Product
            'F': 15,  # Batch No
            'G': 10,  # Quantity
            'H': 12,  # Rate
            'I': 15   # Amount
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Freeze header row
        ws.freeze_panes = ws['A7']

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        return response

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Excel generation error: {str(e)}")
        print(f"Error details: {error_details}")
        
        from django.http import JsonResponse
        return JsonResponse({
            'error': f'Failed to generate Excel: {str(e)}',
            'details': 'Check the date format. Use DDMM, DD/MM, or YYYY-MM-DD format'
        }, status=500)
# Sale Rate Management Views

@login_required
def add_sale_rate(request):
    if request.method == 'POST':
        form = SaleRateForm(request.POST)
        if form.is_valid():
            rate = form.save()
            messages.success(request, f"Sale rate for {rate.productid.product_name} batch {rate.product_batch_no} added successfully!")
            return redirect('sale_rate_list')
    else:
        form = SaleRateForm()
    
    context = {
        'form': form,
        'title': 'Add Sale Rate'
    }
    return render(request, 'rates/sale_rate_form.html', context)

@login_required
def update_sale_rate(request, pk):
    rate = get_object_or_404(SaleRateMaster, id=pk)
    
    if request.method == 'POST':
        form = SaleRateForm(request.POST, instance=rate)
        if form.is_valid():
            form.save()
            messages.success(request, f"Sale rate for {rate.productid.product_name} batch {rate.product_batch_no} updated successfully!")
            return redirect('sale_rate_list')
    else:
        form = SaleRateForm(instance=rate)
    
    context = {
        'form': form,
        'rate': rate,
        'title': 'Update Sale Rate'
    }
    return render(request, 'rates/sale_rate_form.html', context)

@login_required
def delete_sale_rate(request, pk):
    if not request.user.user_type.lower() in ['admin']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('sale_rate_list')
        
    rate = get_object_or_404(SaleRateMaster, id=pk)
    
    if request.method == 'POST':
        product_name = rate.productid.product_name
        batch_no = rate.product_batch_no
        try:
            rate.delete()
            messages.success(request, f"Sale rate for {product_name} batch {batch_no} deleted successfully!")
        except Exception as e:
            messages.error(request, f"Cannot delete sale rate. Error: {str(e)}")
        return redirect('sale_rate_list')
    
    context = {
        'rate': rate,
        'title': 'Delete Sale Rate'
    }
    return render(request, 'rates/sale_rate_confirm_delete.html', context)

@login_required
def get_product_rates_api(request):
    product_id = request.GET.get('product_id')
    batch_no = request.GET.get('batch_no')
    
    if not product_id:
        return JsonResponse({'error': 'Product ID is required'}, status=400)
    
    try:
        # Get batch-specific rates if batch number is provided
        if batch_no:
            try:
                rate = SaleRateMaster.objects.get(productid=product_id, product_batch_no=batch_no)
                return JsonResponse({
                    'success': True,
                    'rate_A': float(rate.rate_A or 0),
                    'rate_B': float(rate.rate_B or 0),
                    'rate_C': float(rate.rate_C or 0)
                })
            except SaleRateMaster.DoesNotExist:
                pass
        
        # Get all rates for the product
        rates = SaleRateMaster.objects.filter(productid=product_id)
        if rates.exists():
            rate_data = []
            for rate in rates:
                rate_data.append({
                    'batch_no': rate.product_batch_no,
                    'rate_A': float(rate.rate_A or 0),
                    'rate_B': float(rate.rate_B or 0),
                    'rate_C': float(rate.rate_C or 0)
                })
            return JsonResponse({
                'success': True,
                'rates': rate_data
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'No rates found for this product'
            })
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    

    #finance payment pdf and excel view

  
@login_required
def export_payments_pdf(request, payment_type=None, start_date=None, end_date=None):
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.shortcuts import get_object_or_404
    from django.utils import timezone
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from io import BytesIO
    from .models import PaymentMaster, ReceiptMaster, InvoicePaid, SalesInvoicePaid
    from datetime import datetime, timedelta
    
    # Get filter parameters from request
    payment_type = request.GET.get('payment_type', 'all')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    # Set default date range if not provided (last 30 days)
    if not start_date_str:
        start_date = timezone.now().date() - timedelta(days=30)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    
    if not end_date_str:
        end_date = timezone.now().date()
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Create response object
    response = HttpResponse(content_type='application/pdf')
    filename = f"payments_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Create PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch, bottomMargin=1*inch)
    elements = []
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1,  # Center aligned
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=12,
    )
    normal_style = styles['Normal']
    
    # Title
    title = Paragraph("Payments Report", title_style)
    elements.append(title)
    
    # Date range info
    date_info = Paragraph(f"Period: {start_date} to {end_date}", normal_style)
    elements.append(date_info)
    elements.append(Spacer(1, 20))
    
    # Collect data based on payment type
    all_data = []
    total_amount = 0
    
    if payment_type in ['all', 'purchase']:
        # Purchase Payments (InvoicePaid)
        purchase_payments = InvoicePaid.objects.filter(
            payment_date__range=[start_date, end_date]
        ).select_related('ip_invoiceid__supplierid')
        
        for payment in purchase_payments:
            all_data.append({
                'date': payment.payment_date,
                'type': 'Purchase Payment',
                'reference': f"Invoice #{payment.ip_invoiceid.invoice_no}",
                'party': payment.ip_invoiceid.supplierid.supplier_name,
                'amount': payment.payment_amount,
                'mode': payment.payment_mode or 'N/A',
                'ref_no': payment.payment_ref_no or 'N/A'
            })
            total_amount += payment.payment_amount
    
    if payment_type in ['all', 'sales']:
        # Sales Payments (SalesInvoicePaid)
        sales_payments = SalesInvoicePaid.objects.filter(
            sales_payment_date__range=[start_date, end_date]
        ).select_related('sales_ip_invoice_no__customerid')
        
        for payment in sales_payments:
            all_data.append({
                'date': payment.sales_payment_date.date(),
                'type': 'Sales Receipt',
                'reference': f"Sales Invoice #{payment.sales_ip_invoice_no.sales_invoice_no}",
                'party': payment.sales_ip_invoice_no.customerid.customer_name,
                'amount': payment.sales_payment_amount,
                'mode': payment.sales_payment_mode,
                'ref_no': payment.sales_payment_ref_no
            })
            total_amount += payment.sales_payment_amount
    
    if payment_type in ['all', 'other']:
        # Other Payments (PaymentMaster)
        other_payments = PaymentMaster.objects.filter(
            payment_date__range=[start_date, end_date]
        )
        
        for payment in other_payments:
            party_name = payment.supplier.supplier_name if payment.supplier else 'N/A'
            reference = f"Payment #{payment.payment_id}"
            if payment.invoice:
                reference = f"Invoice #{payment.invoice.invoice_no}"
            
            all_data.append({
                'date': payment.payment_date,
                'type': 'Other Payment',
                'reference': reference,
                'party': party_name,
                'amount': payment.payment_amount,
                'mode': payment.payment_method,
                'ref_no': payment.payment_reference or 'N/A'
            })
            total_amount += payment.payment_amount
        
        # Receipts (ReceiptMaster)
        receipts = ReceiptMaster.objects.filter(
            receipt_date__range=[start_date, end_date]
        )
        
        for receipt in receipts:
            party_name = receipt.customer.customer_name if receipt.customer else 'N/A'
            reference = f"Receipt #{receipt.receipt_id}"
            if receipt.sales_invoice:
                reference = f"Sales Invoice #{receipt.sales_invoice.sales_invoice_no}"
            
            all_data.append({
                'date': receipt.receipt_date,
                'type': 'Other Receipt',
                'reference': reference,
                'party': party_name,
                'amount': receipt.receipt_amount,
                'mode': receipt.receipt_method,
                'ref_no': receipt.receipt_reference or 'N/A'
            })
            total_amount += receipt.receipt_amount
    
    # Sort data by date
    all_data.sort(key=lambda x: x['date'])
    
    # Create table data
    table_data = []
    
    # Table headers
    headers = ['Date', 'Type', 'Reference', 'Party', 'Amount (₹)', 'Mode', 'Ref No.']
    table_data.append(headers)
    
    # Add data rows
    for item in all_data:
        row = [
            item['date'].strftime('%d-%m-%Y'),
            item['type'],
            item['reference'],
            item['party'],
            f"₹{item['amount']:,.2f}",
            item['mode'],
            item['ref_no']
        ]
        table_data.append(row)
    
    # Add total row
    if all_data:
        table_data.append([''] * 4 + [f"₹{total_amount:,.2f}"] + [''] * 2)
    
    # Create table
    if table_data:
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),  # Amount column right aligned
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),  # Total row bold
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),  # Total row background
        ]))
        elements.append(table)
    else:
        no_data = Paragraph("No payments found for the selected criteria.", normal_style)
        elements.append(no_data)
    
    # Summary
    elements.append(Spacer(1, 20))
    summary_text = f"Total Payments: {len(all_data)} transactions, Total Amount: ₹{total_amount:,.2f}"
    summary = Paragraph(summary_text, heading_style)
    elements.append(summary)
    
    # Generate PDF
    doc.build(elements)
    
    # Get PDF value and return response
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response

@login_required
def export_receipts_pdf(request):
    """
    Export receipts data to PDF with print styling
    """
    # Get filter parameters from request
    receipt_type = request.GET.get('receipt_type', 'all')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    customer_id = request.GET.get('customer')
    
    # Set default date range if not provided (last 30 days)
    if not start_date_str:
        start_date = timezone.now().date() - timedelta(days=30)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    
    if not end_date_str:
        end_date = timezone.now().date()
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Create response object
    response = HttpResponse(content_type='application/pdf')
    filename = f"receipts_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Create PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        spaceAfter=20,
        alignment=1,  # Center
        textColor=colors.HexColor('#2E86AB'),
    )
    
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6,
        alignment=0,  # Left
        textColor=colors.HexColor('#333333'),
    )
    
    # Company Header (for print)
    company_header = [
        Paragraph("<b>PHARMACY MANAGEMENT SYSTEM</b>", title_style),
        Paragraph("Receipts Report", styles['Heading2']),
    ]
    
    # Title with filters info
    title_text = "RECEIPTS REPORT"
    if receipt_type != 'all':
        title_text += f" - {receipt_type.upper()} RECEIPTS"
    
    elements.extend(company_header)
    
    # Report details
    details_table_data = [
        ['Report Period:', f"{start_date} to {end_date}"],
        ['Generated On:', timezone.now().strftime('%d-%m-%Y %H:%M')],
    ]
    
    if customer_id:
        customer = CustomerMaster.objects.filter(customerid=customer_id).first()
        if customer:
            details_table_data.append(['Customer:', customer.customer_name])
    
    details_table = Table(details_table_data, colWidths=[2*inch, 3*inch])
    details_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F8F9FA')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    
    elements.append(details_table)
    elements.append(Spacer(1, 20))
    
    # Collect data based on filters
    all_data = []
    total_amount = 0
    
    # Sales Invoice Payments
    if receipt_type in ['all', 'sales']:
        sales_payments = SalesInvoicePaid.objects.filter(
            sales_payment_date__range=[start_date, end_date]
        ).select_related('sales_ip_invoice_no__customerid')
        
        if customer_id:
            sales_payments = sales_payments.filter(sales_ip_invoice_no__customerid_id=customer_id)
        
        for payment in sales_payments:
            all_data.append({
                'date': payment.sales_payment_date.date(),
                'type': 'Sales Receipt',
                'reference': f"SI#{payment.sales_ip_invoice_no.sales_invoice_no}",
                'customer': payment.sales_ip_invoice_no.customerid.customer_name,
                'amount': payment.sales_payment_amount,
                'mode': payment.sales_payment_mode,
                'ref_no': payment.sales_payment_ref_no
            })
            total_amount += payment.sales_payment_amount
    
    # Other Receipts
    if receipt_type in ['all', 'other']:
        other_receipts = ReceiptMaster.objects.filter(
            receipt_date__range=[start_date, end_date]
        )
        
        if customer_id:
            other_receipts = other_receipts.filter(customer_id=customer_id)
        
        for receipt in other_receipts:
            party_name = receipt.customer.customer_name if receipt.customer else 'General Receipt'
            reference = f"RCPT#{receipt.receipt_id}"
            if receipt.sales_invoice:
                reference = f"SI#{receipt.sales_invoice.sales_invoice_no}"
            
            all_data.append({
                'date': receipt.receipt_date,
                'type': 'Other Receipt',
                'reference': reference,
                'customer': party_name,
                'amount': receipt.receipt_amount,
                'mode': receipt.receipt_method,
                'ref_no': receipt.receipt_reference or 'N/A'
            })
            total_amount += receipt.receipt_amount
    
    # Sort data by date
    all_data.sort(key=lambda x: x['date'], reverse=True)
    
    # Create table data
    table_data = [['Date', 'Type', 'Reference', 'Customer', 'Amount (₹)', 'Mode', 'Ref No.']]
    
    # Add data rows
    for item in all_data:
        row = [
            item['date'].strftime('%d-%m-%Y'),
            item['type'],
            item['reference'],
            Paragraph(item['customer'], styles['Normal']),
            f"₹{item['amount']:,.2f}",
            item['mode'],
            item['ref_no']
        ]
        table_data.append(row)
    
    # Create table
    if len(table_data) > 1:
        table = Table(table_data, repeatRows=1, colWidths=[0.8*inch, 1*inch, 1*inch, 1.5*inch, 0.8*inch, 0.8*inch, 1*inch])
        table.setStyle(TableStyle([
            # Header style
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            
            # Column alignment
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),  # Amount right aligned
        ]))
        elements.append(table)
        
        # Add total row
        elements.append(Spacer(1, 10))
        total_table_data = [
            ['', '', '', 'TOTAL:', f"₹{total_amount:,.2f}", '', '']
        ]
        total_table = Table(total_table_data, colWidths=[0.8*inch, 1*inch, 1*inch, 1.5*inch, 0.8*inch, 0.8*inch, 1*inch])
        total_table.setStyle(TableStyle([
            ('BACKGROUND', (3, 0), (4, 0), colors.HexColor('#F8F9FA')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (3, 0), (4, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (4, 0), (4, 0), 'RIGHT'),
            ('BOX', (0, 0), (-1, 0), 0.5, colors.black),
        ]))
        elements.append(total_table)
    else:
        no_data = Paragraph("<b>No receipts found for the selected criteria.</b>", styles['Normal'])
        elements.append(no_data)
    
    # Summary section
    elements.append(Spacer(1, 20))
    summary_data = [
        ['Summary', ''],
        ['Total Receipts:', f"{len(all_data)}"],
        ['Total Amount:', f"₹{total_amount:,.2f}"],
        ['Average per Receipt:', f"₹{total_amount/len(all_data):,.2f}" if all_data else '₹0.00'],
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(summary_table)
    
    # Footer for print
    elements.append(Spacer(1, 20))
    footer_text = f"Generated by Pharmacy Management System on {timezone.now().strftime('%d-%m-%Y %H:%M')} - Page 1"
    footer = Paragraph(footer_text, ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=7,
        alignment=1,
        textColor=colors.HexColor('#666666'),
    ))
    elements.append(footer)
    
    # Generate PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response

def export_receipts_excel(request):
    """
    Export receipts data to Excel
    """
    # Get filter parameters from request
    receipt_type = request.GET.get('receipt_type', 'all')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    customer_id = request.GET.get('customer')
    
    # Set default date range if not provided (last 30 days)
    if not start_date_str:
        start_date = timezone.now().date() - timedelta(days=30)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    
    if not end_date_str:
        end_date = timezone.now().date()
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Collect data based on filters
    all_data = []
    total_amount = 0
    
    # Sales Invoice Payments
    if receipt_type in ['all', 'sales']:
        sales_payments = SalesInvoicePaid.objects.filter(
            sales_payment_date__range=[start_date, end_date]
        ).select_related('sales_ip_invoice_no__customerid')
        
        if customer_id:
            sales_payments = sales_payments.filter(sales_ip_invoice_no__customerid_id=customer_id)
        
        for payment in sales_payments:
            all_data.append({
                'Date': payment.sales_payment_date.date(),
                'Type': 'Sales Receipt',
                'Reference': f"SI#{payment.sales_ip_invoice_no.sales_invoice_no}",
                'Customer': payment.sales_ip_invoice_no.customerid.customer_name,
                'Amount': payment.sales_payment_amount,
                'Payment Mode': payment.sales_payment_mode,
                'Reference No': payment.sales_payment_ref_no,
                'Status': 'Completed'
            })
            total_amount += payment.sales_payment_amount
    
    # Other Receipts
    if receipt_type in ['all', 'other']:
        other_receipts = ReceiptMaster.objects.filter(
            receipt_date__range=[start_date, end_date]
        )
        
        if customer_id:
            other_receipts = other_receipts.filter(customer_id=customer_id)
        
        for receipt in other_receipts:
            party_name = receipt.customer.customer_name if receipt.customer else 'General Receipt'
            reference = f"RCPT#{receipt.receipt_id}"
            if receipt.sales_invoice:
                reference = f"SI#{receipt.sales_invoice.sales_invoice_no}"
            
            all_data.append({
                'Date': receipt.receipt_date,
                'Type': 'Other Receipt',
                'Reference': reference,
                'Customer': party_name,
                'Amount': receipt.receipt_amount,
                'Payment Mode': receipt.receipt_method,
                'Reference No': receipt.receipt_reference or 'N/A',
                'Status': 'Completed'
            })
            total_amount += receipt.receipt_amount
    
    # Sort data by date
    all_data.sort(key=lambda x: x['Date'], reverse=True)
    
    # Create DataFrame
    if all_data:
        df = pd.DataFrame(all_data)
        
        # Format date column
        df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%d-%m-%Y')
        
        # Format amount column with Indian currency format
        df['Amount'] = df['Amount'].apply(lambda x: f"₹{x:,.2f}")
        
        # Reorder columns for better presentation
        column_order = ['Date', 'Type', 'Reference', 'Customer', 'Amount', 'Payment Mode', 'Reference No', 'Status']
        df = df[column_order]
    else:
        # Create empty DataFrame with columns if no data
        df = pd.DataFrame(columns=['Date', 'Type', 'Reference', 'Customer', 'Amount', 'Payment Mode', 'Reference No', 'Status'])
    
    # Create response object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"receipts_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Create Excel file using pandas
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        # Write main data
        df.to_excel(writer, sheet_name='Receipts', index=False, startrow=5)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Receipts']
        
        # Add title and header information
        title = "PHARMACY MANAGEMENT SYSTEM - RECEIPTS REPORT"
        if receipt_type != 'all':
            title += f" - {receipt_type.upper()} RECEIPTS"
        
        # Title
        worksheet['A1'] = title
        worksheet['A2'] = f"Report Period: {start_date} to {end_date}"
        worksheet['A3'] = f"Generated On: {timezone.now().strftime('%d-%m-%Y %H:%M')}"
        
        if customer_id:
            customer = CustomerMaster.objects.filter(customerid=customer_id).first()
            if customer:
                worksheet['A4'] = f"Customer: {customer.customer_name}"
        
        # Style the header
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Title styling
        title_cell = worksheet['A1']
        title_cell.font = Font(size=16, bold=True, color='2E86AB')
        title_cell.alignment = Alignment(horizontal='center')
        
        # Header information styling
        for row in [2, 3, 4]:
            cell = worksheet[f'A{row}']
            cell.font = Font(size=11, bold=True)
        
        # Merge title across columns
        worksheet.merge_cells('A1:H1')
        
        # Header styling
        for col in range(1, len(column_order) + 1):
            cell = worksheet.cell(row=6, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Adjust column widths
        column_widths = {
            'A': 12,  # Date
            'B': 15,  # Type
            'C': 18,  # Reference
            'D': 30,  # Customer
            'E': 15,  # Amount
            'F': 15,  # Payment Mode
            'G': 20,  # Reference No
            'H': 12   # Status
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # Add summary section
        if all_data:
            summary_row = len(all_data) + 8
            worksheet[f'D{summary_row}'] = 'SUMMARY'
            worksheet[f'D{summary_row}'].font = Font(bold=True, size=12)
            
            worksheet[f'D{summary_row + 1}'] = 'Total Receipts:'
            worksheet[f'E{summary_row + 1}'] = len(all_data)
            
            worksheet[f'D{summary_row + 2}'] = 'Total Amount:'
            worksheet[f'E{summary_row + 2}'] = f"₹{total_amount:,.2f}"
            worksheet[f'E{summary_row + 2}'].font = Font(bold=True)
            
            worksheet[f'D{summary_row + 3}'] = 'Average per Receipt:'
            worksheet[f'E{summary_row + 3}'] = f"₹{total_amount/len(all_data):,.2f}" if all_data else '₹0.00'
            
            # Style summary section
            for row in range(summary_row, summary_row + 4):
                worksheet[f'D{row}'].font = Font(bold=True)
                worksheet[f'D{row}'].fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    return response


    from django.http import HttpResponse
from django.template.loader import render_to_string
from io import BytesIO
import csv
from .models import ProductMaster

# Simple CSV Export (No external packages needed)
def export_products_excel(request):
    """Export products to CSV/Excel"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products_export.csv"'
    
    writer = csv.writer(response)
    # Write headers
    writer.writerow(['ID', 'Product Name', 'Company', 'Packing', 'Salt', 'Category', 'HSN Code', 'Barcode'])
    
    # Get product data
    products = ProductMaster.objects.all().order_by('productid')
    
    # Write data
    for product in products:
        writer.writerow([
            product.productid,
            product.product_name,
            product.product_company,
            product.product_packing,
            product.product_salt,
            product.product_category,
            product.product_hsn,
            product.product_barcode or 'N/A'
        ])
    
    return response

# PDF Export (If you have reportlab installed)
def export_products_pdf(request):
    """Export products to PDF"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="products_report.pdf"'
        
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        
        # Add title
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, 800, "Products Report")
        
        # Add date
        p.setFont("Helvetica", 10)
        from django.utils import timezone
        p.drawString(100, 780, f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}")
        
        # Get product data
        products = ProductMaster.objects.all().order_by('productid')[:50]  # Limit for one page
        
        # Add table headers
        p.setFont("Helvetica-Bold", 10)
        y_position = 750
        headers = ['ID', 'Product Name', 'Company', 'Packing']
        
        for i, header in enumerate(headers):
            p.drawString(100 + (i * 120), y_position, header)
        
        y_position -= 20
        
        # Add product data
        p.setFont("Helvetica", 8)
        for product in products:
            if y_position < 100:  # Start new page if needed
                p.showPage()
                p.setFont("Helvetica", 8)
                y_position = 750
            
            p.drawString(100, y_position, str(product.productid))
            p.drawString(150, y_position, product.product_name[:20])  # Limit name length
            p.drawString(300, y_position, product.product_company[:15])  # Limit company length
            p.drawString(450, y_position, product.product_packing)
            y_position -= 15
        
        p.save()
        buffer.seek(0)
        response.write(buffer.getvalue())
        buffer.close()
        return response
        
    except ImportError:
        # Fallback to HTML if reportlab is not installed
        return export_products_html(request)

# HTML Export Fallback
def export_products_html(request):
    """Export products as HTML (can be printed as PDF)"""
    products = ProductMaster.objects.all().order_by('productid')
    
    html_content = render_to_string('products_export.html', {
        'products': products,
        'title': 'Products Report'
    })
    
    response = HttpResponse(html_content, content_type='text/html')
    response['Content-Disposition'] = 'attachment; filename="products_report.html"'
    return response