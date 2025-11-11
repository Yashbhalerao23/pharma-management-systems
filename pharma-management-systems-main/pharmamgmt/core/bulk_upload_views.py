import csv
import io
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect
from .models import ProductMaster
from django.core.paginator import Paginator
try:
    from openpyxl import Workbook, load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

@login_required
def bulk_upload_products(request):
    if request.method == 'POST':
        file = request.FILES.get('file')
        
        if not file:
            messages.error(request, 'Please select a file to upload')
            return redirect('bulk_upload_products')
        
        try:
            if file.name.endswith('.csv'):
                products = process_csv_file(file)
            elif file.name.endswith(('.xlsx', '.xls')):
                products = process_excel_file(file)
            else:
                messages.error(request, 'Invalid file format. Please upload CSV or Excel file')
                return redirect('bulk_upload_products')
            
            success_count = 0
            error_count = 0
            errors = []
            
            for idx, product_data in enumerate(products, start=2):
                try:
                    ProductMaster.objects.create(
                        product_name=product_data['product_name'],
                        product_company=product_data['product_company'],
                        product_packing=product_data['product_packing'],
                        product_category=product_data['product_category'],
                        product_salt='N/A',
                        product_hsn='N/A',
                        product_hsn_percent='0',
                        product_barcode=product_data.get('product_barcode', '') or None
                    )
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    errors.append(f"Row {idx}: {str(e)[:100]}")
            
            if success_count > 0:
                messages.success(request, f'Successfully uploaded {success_count} products')
            
            if error_count > 0:
                error_msg = f'{error_count} products failed. ' + '; '.join(errors[:5])
                messages.warning(request, error_msg)
            
            # Redirect to last page with ID sorting to show newly added products
            all_products = ProductMaster.objects.all().order_by('productid')
            paginator = Paginator(all_products, 50)
            last_page = paginator.num_pages
            
            from django.http import HttpResponseRedirect
            return HttpResponseRedirect(f'/products/?sort=productid&page={last_page}')
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('bulk_upload_products')
    
    return render(request, 'products/bulk_upload_products.html')

def process_csv_file(file):
    decoded_file = file.read().decode('utf-8').splitlines()
    reader = csv.DictReader(decoded_file)
    products = []
    
    for row in reader:
        if row.get('product_name') and row.get('product_name').strip():  # Skip empty rows
            products.append({
                'product_name': row['product_name'].strip(),
                'product_company': row['product_company'].strip(),
                'product_packing': row['product_packing'].strip(),
                'product_category': row['product_category'].strip(),
                'product_barcode': row.get('product_barcode', '').strip()
            })
    
    return products

def process_excel_file(file):
    if not EXCEL_SUPPORT:
        raise Exception('Excel support not available. Please install openpyxl: pip install openpyxl')
    
    wb = load_workbook(file)
    ws = wb.active
    products = []
    
    headers = [cell.value for cell in ws[1]]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Skip empty rows
            products.append({
                'product_name': str(row[0]).strip(),
                'product_company': str(row[1]).strip(),
                'product_packing': str(row[2]).strip(),
                'product_category': str(row[3]).strip(),
                'product_barcode': str(row[4] if len(row) > 4 and row[4] else '').strip()
            })
    
    return products

@login_required
def download_product_template(request):
    format_type = request.GET.get('format', 'csv')
    
    headers = ['product_name', 'product_company', 'product_packing', 'product_category', 'product_barcode']
    sample_data = [
        ['Paracetamol 500mg', 'ABC Pharma', '10x10', 'Tablet', '1234567890'],
        ['Amoxicillin 250mg', 'XYZ Labs', '10x10', 'Capsule', '0987654321']
    ]
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="product_template.csv"'
        
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(sample_data)
        
        return response
    
    elif format_type == 'excel':
        if not EXCEL_SUPPORT:
            messages.error(request, 'Excel support not available. Please use CSV format.')
            return redirect('bulk_upload_products')
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Products'
        
        ws.append(headers)
        for row in sample_data:
            ws.append(row)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="product_template.xlsx"'
        
        wb.save(response)
        
        return response
